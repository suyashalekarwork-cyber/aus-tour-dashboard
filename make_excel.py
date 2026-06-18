"""
make_excel.py — build Token_Review.xlsx to send to the PM.

Run:  python make_excel.py

Reads tokens.js (the same tokens shown in the web editor) and writes an Excel
file with one row per token. Flagged tokens are listed first and highlighted,
with a recommended fix pre-filled. The PM reviews, adjusts the Action / Corrected
columns, and sends the file back. You then run import_excel.py on it.
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE      = Path(__file__).parent
TOKENS_JS = BASE / 'tokens.js'
OUT_XLSX  = BASE / 'data' / 'Token_Review.xlsx'

SRC_LABEL = {'B': 'Board', 'C': 'Competitor', 'BC': 'Board + Competitor'}

ISSUE_LABEL = {
    'caps':     'ALL CAPS',
    'lower':    'lowercase start',
    'extra':    'extra word(s)',
    'nonplace': 'not a place / possessive',
    'dupe':     'duplicate',
    'split':    'two places joined',
    None:       'clean',
    '':         'clean',
}

# Visual styles
HEADER_FILL  = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=11)
FLAG_FILL    = PatternFill('solid', fgColor='FFF2CC')   # light amber for flagged rows
EDIT_FILL    = PatternFill('solid', fgColor='E2EFDA')   # light green for editable cols
THIN         = Side(style='thin', color='D9D9D9')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP     = Alignment(wrap_text=True, vertical='top')
TOP          = Alignment(vertical='top')


def load_tokens():
    raw = TOKENS_JS.read_text(encoding='utf-8').strip()
    raw = re.sub(r'^const TOKENS = ', '', raw).rstrip(';')
    return json.loads(raw)


def recommend(tok):
    """Return (action, corrected) pre-filled from the detected issue."""
    issue = tok.get('issue')
    if not issue:
        return 'keep', ''
    if issue == 'split':
        parts = tok.get('split') or []
        return 'split', ' | '.join(parts)
    # caps / lower / extra / dupe / nonplace → rename to the suggestion
    return 'rename', tok.get('suggest', '') or ''


def build():
    tokens = load_tokens()
    # flagged first (already the order in tokens.js), but enforce it
    tokens.sort(key=lambda t: (0 if t.get('issue') else 1, -t.get('count', 0)))

    wb = Workbook()

    # ── Instructions sheet ──────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = 'Instructions'
    info_lines = [
        ('How to review the product tokens', True),
        ('', False),
        ('These are place names pulled from tour itineraries. Some were captured', False),
        ('with mistakes. Please review and fix them on the "Tokens to review" tab.', False),
        ('', False),
        ('For each row, set the ACTION column (green) using the dropdown:', True),
        ('   keep    →  the name is fine, leave it as is', False),
        ('   rename  →  fix the name; type the correct name in "Corrected name(s)"', False),
        ('   split   →  it is two places; put both, separated by  |  e.g.  Perth | Fremantle', False),
        ('   delete  →  not a real place, remove it completely', False),
        ('', False),
        ('Tips:', True),
        ('   - Flagged rows are at the TOP and highlighted in amber, with a', False),
        ('     suggested fix already filled in. Just change it where you disagree.', False),
        ('   - Rows marked "keep" further down are already clean — you can ignore', False),
        ('     them, but feel free to fix any you spot (use the filter on the header).', False),
        ('   - Only change the two green columns (Action, Corrected name(s)).', False),
        ('     Leave the other columns as they are.', False),
        ('', False),
        ('When done, just save and send the file back. Thank you!', True),
    ]
    for i, (text, bold) in enumerate(info_lines, start=1):
        c = ws_info.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
    ws_info.column_dimensions['A'].width = 95

    # ── Tokens sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet('Tokens to review')
    headers = ['Token', 'Issue', 'State', 'Day rows', 'Source',
               'Example context (where it appeared)', 'Action', 'Corrected name(s)']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER

    for tok in tokens:
        action, corrected = recommend(tok)
        ctx = tok.get('ctx') or []
        ctx_str = '\n'.join(f'• {c}' for c in ctx[:3])
        ws.append([
            tok['name'],
            ISSUE_LABEL.get(tok.get('issue'), tok.get('issue')),
            tok.get('state', ''),
            tok.get('count', 0),
            SRC_LABEL.get(tok.get('src'), tok.get('src', '')),
            ctx_str,
            action,
            corrected,
        ])

    # Styling + per-row highlight
    n_rows = ws.max_row
    for r in range(2, n_rows + 1):
        is_flagged = ws.cell(row=r, column=2).value != 'clean'
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.alignment = WRAP_TOP if col == 6 else TOP
            if col in (7, 8):
                cell.fill = EDIT_FILL
            elif is_flagged:
                cell.fill = FLAG_FILL

    # Dropdown for the Action column (G), all data rows.
    # NOTE: showDropDown must stay False/0 — Excel inverts it (True hides the arrow).
    # Enable the input + error popups so the cell is clearly a pick-list.
    dv = DataValidation(type='list', formula1='"keep,rename,split,delete"',
                        allow_blank=True, showDropDown=False)
    dv.showInputMessage = True
    dv.showErrorMessage = True
    dv.promptTitle = 'Pick an action'
    dv.prompt = 'Click the arrow and choose: keep, rename, split, or delete'
    dv.errorTitle = 'Use the dropdown'
    dv.error = 'Please pick one of: keep, rename, split, delete'
    ws.add_data_validation(dv)
    dv.add(f'G2:G{n_rows}')

    # Column widths
    widths = {'A': 34, 'B': 22, 'C': 8, 'D': 9, 'E': 20, 'F': 55, 'G': 12, 'H': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{n_rows}'

    wb.save(OUT_XLSX)

    flagged = sum(1 for t in tokens if t.get('issue'))
    print(f'Wrote {OUT_XLSX.name}: {len(tokens)} tokens ({flagged} flagged, '
          f'{len(tokens) - flagged} clean).')
    print('Send this file to the PM. When it comes back, run:  python import_excel.py')


if __name__ == '__main__':
    build()
