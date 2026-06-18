"""
import_excel.py — apply the PM's reviewed Excel back into the pipeline.

Run:  python import_excel.py                  (uses Token_Review.xlsx)
  or: python import_excel.py "some other.xlsx"

Reads the Action / Corrected columns the PM filled in, merges them into
token_corrections.json (keeping a timestamped backup of the previous version),
then re-runs prepare_data.py so the Itinerary Builder picks up the clean data.
"""
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE             = Path(__file__).parent          # the "pipeline" folder
ROOT             = BASE.parent                     # the "Build and explore" folder
DEFAULT_XLSX     = ROOT / 'data' / 'Token_Review.xlsx'
CORRECTIONS_PATH = ROOT / 'data' / 'token_corrections.json'
BACKUPS_DIR      = ROOT / 'data' / 'corrections_history'
PREPARE_SCRIPT   = BASE / 'prepare_data.py'
SHEET_NAME       = 'Tokens to review'


def load_existing():
    if not CORRECTIONS_PATH.exists():
        return {'rename': {}, 'split': {}, 'delete': []}
    with open(CORRECTIONS_PATH, 'r', encoding='utf-8') as f:
        c = json.load(f)
    return {
        'rename': dict(c.get('rename', {})),
        'split':  dict(c.get('split', {})),
        'delete': list(c.get('delete', [])),
    }


def read_sheet(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f'ERROR: sheet "{SHEET_NAME}" not found in {xlsx_path.name}. '
                 f'Found: {wb.sheetnames}')
    ws = wb[SHEET_NAME]

    # Map header -> column index
    headers = {}
    for col, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col

    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        sys.exit(f'ERROR: could not find a column for {names} in the sheet header.')

    c_token = col('token')
    c_action = col('action')
    c_corr = col('corrected name(s)', 'corrected name', 'corrected')

    rows = []
    for r in range(2, ws.max_row + 1):
        token = ws.cell(row=r, column=c_token).value
        if not token:
            continue
        action = (ws.cell(row=r, column=c_action).value or '').strip().lower()
        corrected = (ws.cell(row=r, column=c_corr).value or '')
        corrected = str(corrected).strip()
        rows.append((str(token).strip(), action, corrected))
    return rows


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        sys.exit(f'ERROR: {xlsx_path} not found. Put the PM\'s file here or pass its path.')

    print(f'Reading {xlsx_path.name} ...')
    rows = read_sheet(xlsx_path)

    corr = load_existing()
    n_rename = n_split = n_delete = n_skip = n_bad = 0

    for token, action, corrected in rows:
        if action in ('', 'keep'):
            n_skip += 1
            continue
        if action == 'rename':
            if corrected and corrected != token:
                corr['rename'][token] = corrected
                corr['delete'] = [d for d in corr['delete'] if d != token]
                n_rename += 1
            else:
                print(f'  ! "{token}": action=rename but no (different) corrected name — skipped')
                n_bad += 1
        elif action == 'split':
            parts = [p.strip() for p in corrected.split('|') if p.strip()]
            if len(parts) >= 2:
                corr['split'][token] = parts
                n_split += 1
            else:
                print(f'  ! "{token}": action=split but needs 2+ names separated by | — skipped')
                n_bad += 1
        elif action == 'delete':
            if token not in corr['delete']:
                corr['delete'].append(token)
            corr['rename'].pop(token, None)
            corr['split'].pop(token, None)
            n_delete += 1
        else:
            print(f'  ! "{token}": unknown action "{action}" — skipped')
            n_bad += 1

    corr['delete'] = sorted(set(corr['delete']))

    # Backup the previous corrections file before overwriting
    BACKUPS_DIR.mkdir(exist_ok=True)
    if CORRECTIONS_PATH.exists():
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        backup = BACKUPS_DIR / f'token_corrections_{ts}_from_excel.json'
        backup.write_text(CORRECTIONS_PATH.read_text(encoding='utf-8'), encoding='utf-8')
        print(f'Backed up previous corrections → {backup.name}')

    with open(CORRECTIONS_PATH, 'w', encoding='utf-8') as f:
        json.dump(corr, f, ensure_ascii=False, indent=2)

    print(f'\nApplied from Excel:  {n_rename} rename, {n_split} split, {n_delete} delete  '
          f'(skipped {n_skip} keep/blank, {n_bad} invalid)')
    print(f'Total now in token_corrections.json:  {len(corr["rename"])} rename, '
          f'{len(corr["split"])} split, {len(corr["delete"])} delete')

    print('\nRe-running prepare_data.py ...')
    result = subprocess.run([sys.executable, str(PREPARE_SCRIPT)],
                            cwd=str(BASE), encoding='utf-8', errors='replace')
    if result.returncode == 0:
        print('\nDone! data.js regenerated. Reload the Itinerary Builder to see the clean data.')
    else:
        print('\nWARNING: prepare_data.py exited with an error — check the output above.')


if __name__ == '__main__':
    main()
