"""
make_excel.py — build Token_Review.xlsx for PM / marketing review.

Run:  python make_excel.py

Three sheets:
  1. "Needs Review"        — PM must action these (~500-800 rows)
  2. "Auto-Fixed"          — obvious fixes applied automatically (reference only)
  3. "Confirmed Clean"     — validated against Australian GeoNames database (reference only)

The PM only needs to work on the "Needs Review" sheet.
"""
import json
import re
from itertools import cycle
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE      = Path(__file__).parent
ROOT      = BASE.parent
TOKENS_JS = ROOT / 'frontend' / 'tokens.js'
OUT_XLSX  = ROOT / 'frontend' / 'Token_Review.xlsx'

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY        = PatternFill('solid', fgColor='1F4E78')   # sheet headers
DARK_GREEN  = PatternFill('solid', fgColor='1E6B3C')   # confirmed-clean header
DARK_ORANGE = PatternFill('solid', fgColor='7F3F00')   # auto-fixed header

WHITE_FONT  = Font(color='FFFFFF', bold=True, size=11)
BOLD_11     = Font(bold=True, size=11)

# Cluster colour bands — alternating pale stripes so a cluster reads as a block
CLUSTER_BANDS = [
    PatternFill('solid', fgColor='EBF5FB'),   # pale blue
    PatternFill('solid', fgColor='FEF9E7'),   # pale yellow
    PatternFill('solid', fgColor='EAFAF1'),   # pale green
    PatternFill('solid', fgColor='FDF2F8'),   # pale pink
    PatternFill('solid', fgColor='F4ECF7'),   # pale purple
]
NO_FILL     = PatternFill(fill_type=None)
EDIT_FILL   = PatternFill('solid', fgColor='D5F5E3')   # light green — editable cols
HINT_FILL   = PatternFill('solid', fgColor='ECECEC')   # light gray — read-only hint
HINT_FONT   = Font(color='555555', italic=True, size=10)  # gray italic — reference only
FLAG_FILL   = PatternFill('solid', fgColor='FFF3CD')   # amber — flagged issue
AUTO_FILL   = PatternFill('solid', fgColor='FEF9E7')   # pale yellow — auto-fixed
CONF_FILL   = PatternFill('solid', fgColor='E8F8F5')   # pale teal — confirmed

THIN  = Side(style='thin',   color='D9D9D9')
MED   = Side(style='medium', color='B0B0B0')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GROUP_TOP   = Border(left=THIN, right=THIN, top=MED,  bottom=THIN)
WRAP_TOP    = Alignment(wrap_text=True, vertical='top')
TOP         = Alignment(vertical='top')

ISSUE_LABEL = {
    'caps':     'ALL CAPS — needs title case',
    'lower':    'Lowercase start',
    'extra':    'Extra word attached',
    'nonplace': 'Not a place (possessive)',
    'dupe':     'Duplicate spelling',
    'split':    'Two places joined with &',
    None:       'Clean — no issue detected',
    '':         'Clean — no issue detected',
}

SRC_LABEL = {'B': 'Board only', 'C': 'Operators only', 'BC': 'Board + Operators'}


# ── Load tokens ───────────────────────────────────────────────────────────────

def load_tokens():
    raw = TOKENS_JS.read_text(encoding='utf-8').strip()
    raw = re.sub(r'^const TOKENS = ', '', raw).rstrip(';')
    return json.loads(raw)


# ── Suggested action ──────────────────────────────────────────────────────────

def recommend(tok):
    """Pre-fill Action + Corrected based on detected issue."""
    issue = tok.get('issue')
    if not issue:
        return 'keep', ''
    if issue == 'split':
        parts = tok.get('split') or []
        return 'split', ' | '.join(parts)
    if issue in ('nonplace', 'dupe') and not tok.get('suggest'):
        return 'delete', ''
    if issue == 'nonplace':
        return 'delete', ''
    return 'rename', tok.get('suggest', '') or ''


# ── Sheet writer helpers ──────────────────────────────────────────────────────

def write_header(ws, headers, fill, font):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER


def add_action_dropdown(ws, n_rows):
    dv = DataValidation(
        type='list',
        formula1='"keep,rename,split,delete"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.showInputMessage = True
    dv.showErrorMessage = True
    dv.promptTitle = 'Choose action'
    dv.prompt     = 'keep · rename · split · delete'
    dv.errorTitle = 'Use the dropdown'
    dv.error      = 'Please pick: keep, rename, split, or delete'
    ws.add_data_validation(dv)
    dv.add(f'G2:G{n_rows}')


# ── Sheet 1: Needs Review ─────────────────────────────────────────────────────

def build_review_sheet(ws, review_tokens):
    """
    Columns: Place name | State(s) | Tours | Example tour | Issue | Suggestion | Action | Correct name

    Tokens are sorted so cluster members sit together.
    Each cluster gets an alternating colour band.
    """
    headers = [
        'Place name',
        'State(s)',
        'Distinct tours',
        'Example tour',
        'Issue',
        'Suggestion (reference only — do not edit)',
        'Action',          # PM edits this
        'Correct name',    # PM edits this (only when Action = rename)
    ]
    write_header(ws, headers, NAVY, WHITE_FONT)

    # Sort: group clusters together (by group name), canonical first within group,
    # then unclassified tokens by count descending.
    def sort_key(t):
        grp  = t.get('group', t['name'])
        is_canonical = (grp == t['name'])
        return (grp.lower(), 0 if is_canonical else 1, -t['count'])

    sorted_tokens = sorted(review_tokens, key=sort_key)

    # Assign cluster band colours
    cluster_colour = {}
    band_cycle = cycle(CLUSTER_BANDS)
    for tok in sorted_tokens:
        grp = tok.get('group', tok['name'])
        # Only assign a band if the cluster has >1 member
        if grp not in cluster_colour:
            cluster_colour[grp] = None  # placeholder
    # Count members per group
    group_sizes = {}
    for tok in sorted_tokens:
        grp = tok.get('group', tok['name'])
        group_sizes[grp] = group_sizes.get(grp, 0) + 1
    final_colours = {}
    for grp, size in group_sizes.items():
        final_colours[grp] = next(band_cycle) if size > 1 else NO_FILL

    prev_group = None
    for tok in sorted_tokens:
        action, corrected = recommend(tok)
        issue_text = ISSUE_LABEL.get(tok.get('issue'), tok.get('issue') or '')
        grp = tok.get('group', tok['name'])

        sample = '; '.join(tok.get('sample_tours', [])[:2])

        ws.append([
            tok['name'],
            tok.get('state', ''),
            tok.get('tours', tok['count']),
            sample,
            issue_text,
            corrected if action in ('rename', 'split') else ('delete' if action == 'delete' else '—'),
            action,
            '',           # Correct name — blank for PM to fill if they disagree
        ])

        r = ws.max_row
        row_fill = final_colours.get(grp, NO_FILL)
        is_top_of_group = (grp != prev_group) and (group_sizes.get(grp, 1) > 1)
        prev_group = grp

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = GROUP_TOP if is_top_of_group else BORDER
            cell.alignment = TOP

            if col_idx in (7, 8):        # editable columns (Action, Correct name) — green
                cell.fill = EDIT_FILL
            elif col_idx == 6:           # Suggestion — gray read-only look
                cell.fill = HINT_FILL
                cell.font = HINT_FONT
            elif tok.get('issue'):       # flagged — amber tint (subtle over cluster band)
                cell.fill = FLAG_FILL if row_fill.fill_type is None else row_fill
            else:
                cell.fill = row_fill

    n_rows = ws.max_row
    add_action_dropdown(ws, n_rows)

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 10   # State(s)
    ws.column_dimensions['C'].width = 13   # Distinct tours
    ws.column_dimensions['D'].width = 40   # Example tour
    ws.column_dimensions['E'].width = 26   # Issue
    ws.column_dimensions['F'].width = 36   # Suggestion
    ws.column_dimensions['G'].width = 10   # Action
    ws.column_dimensions['H'].width = 36   # Correct name

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{n_rows}'
    ws.sheet_properties.tabColor = 'FF0000'   # red tab — needs action


# ── Sheet 2: Auto-Fixed ───────────────────────────────────────────────────────

def build_auto_sheet(ws, auto_tokens):
    """
    Shows tokens being fixed automatically.
    PM can review but does NOT need to action.
    """
    headers = ['Place name', 'Distinct tours', 'Issue', 'Auto-fix applied']
    write_header(ws, headers, DARK_ORANGE, WHITE_FONT)

    for tok in sorted(auto_tokens, key=lambda t: -t['count']):
        issue = tok.get('issue')
        if not issue:
            fix = f"ACCEPTED as-is (used in {tok.get('tours', tok['count'])} distinct tours — very likely real)"
        elif issue == 'nonplace':
            fix = 'DELETE - possessive, not a place name'
        elif issue in ('caps', 'lower'):
            fix = f"RENAME -> {tok.get('suggest', '')}"
        elif issue == 'dupe':
            fix = f"RENAME -> {tok.get('suggest', '')}  (same place, different spelling)"
        elif issue == 'extra':
            fix = f"RENAME -> {tok.get('suggest', '')}  (stripped extra word)"
        else:
            fix = tok.get('suggest', '')

        ws.append([
            tok['name'],
            tok.get('tours', tok['count']),
            ISSUE_LABEL.get(issue, issue or ''),
            fix,
        ])
        r = ws.max_row
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = AUTO_FILL
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = TOP

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 55
    ws.freeze_panes = 'A2'
    ws.sheet_properties.tabColor = 'FFA500'   # orange tab


# ── Sheet 3: Confirmed Clean ──────────────────────────────────────────────────

def build_confirmed_sheet(ws, confirmed_tokens):
    """
    Tokens validated by the Australian GeoNames database.
    These are real places with the correct spelling — no action needed.
    """
    headers = ['Place name', 'Distinct tours', 'Source']
    write_header(ws, headers, DARK_GREEN, WHITE_FONT)

    for tok in sorted(confirmed_tokens, key=lambda t: -t['count']):
        ws.append([
            tok['name'],
            tok.get('tours', tok['count']),
            SRC_LABEL.get(tok.get('src', ''), tok.get('src', '')),
        ])
        r = ws.max_row
        for col in range(1, 4):
            ws.cell(row=r, column=col).fill = CONF_FILL
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = TOP

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.freeze_panes = 'A2'
    ws.sheet_properties.tabColor = '00B050'   # green tab


# ── Instructions sheet ────────────────────────────────────────────────────────

def build_instructions_sheet(ws, counts):
    ws.title = 'Instructions'
    lines = [
        ('Token Review — Instructions', True, 14),
        ('', False, 11),
        (f'This file has {counts["review"]} tokens that need a decision from you.', False, 11),
        (f'({counts["auto"]} fixes were applied automatically, '
         f'{counts["confirmed"]} tokens were confirmed by the government place-name database.)', False, 11),
        ('', False, 11),
        ('WHAT TO DO', True, 12),
        ('1.  Open the red "Needs Review" tab.', False, 11),
        ('2.  For each row, set the ACTION column (green) using the dropdown.', False, 11),
        ('3.  When done, save and send the file back.', False, 11),
        ('', False, 11),
        ('ACTION OPTIONS', True, 12),
        ('   keep    — the name is fine, leave it as-is', False, 11),
        ('   rename  — fix the name; type the correct name in the "Correct name" column', False, 11),
        ('   split   — it is two places joined; put both in "Correct name" separated by  |', False, 11),
        ('             e.g.  Perth | Fremantle', False, 11),
        ('   delete  — not a real place, remove it', False, 11),
        ('', False, 11),
        ('TIPS', True, 12),
        ('   • Rows with the same background colour are the same place with different spellings.', False, 11),
        ('     Keep the top row, rename the rest into it (already pre-filled).', False, 11),
        ('   • The "Suggestion" column is pre-filled — only change it if you disagree.', False, 11),
        ('   • The other two tabs (Auto-Fixed, Confirmed Clean) are for reference only.', False, 11),
        ('     You do NOT need to change anything there.', False, 11),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
    ws.column_dimensions['A'].width = 90


# ── Main ──────────────────────────────────────────────────────────────────────

def build():
    tokens = load_tokens()

    # Bucket each token
    review_tokens    = []   # PM must decide
    auto_tokens      = []   # auto-fixed, reference only
    confirmed_tokens = []   # GeoNames confirmed + clean, reference only

    for tok in tokens:
        if tok.get('resolved'):
            continue   # already in corrections — skip entirely
        if tok.get('auto'):
            auto_tokens.append(tok)
        elif tok.get('geonames') and not tok.get('issue'):
            confirmed_tokens.append(tok)
        else:
            review_tokens.append(tok)

    counts = {
        'review':    len(review_tokens),
        'auto':      len(auto_tokens),
        'confirmed': len(confirmed_tokens),
    }

    wb = Workbook()

    # Instructions (first/default sheet)
    ws_info = wb.active
    build_instructions_sheet(ws_info, counts)

    # Needs Review
    ws_review = wb.create_sheet('Needs Review')
    build_review_sheet(ws_review, review_tokens)

    # Auto-Fixed
    ws_auto = wb.create_sheet('Auto-Fixed')
    build_auto_sheet(ws_auto, auto_tokens)

    # Confirmed Clean
    ws_conf = wb.create_sheet('Confirmed Clean')
    build_confirmed_sheet(ws_conf, confirmed_tokens)

    wb.save(OUT_XLSX)

    print(f'\nWrote {OUT_XLSX.name}')
    print(f'  {"Needs Review":<18}  {counts["review"]:>4} rows  << PM action required')
    print(f'  {"Auto-Fixed":<18}  {counts["auto"]:>4} rows  (reference)')
    print(f'  {"Confirmed Clean":<18}  {counts["confirmed"]:>4} rows  (reference)')
    print(f'\nSend the file to the PM. When it comes back run:  python import_excel.py')


if __name__ == '__main__':
    build()
