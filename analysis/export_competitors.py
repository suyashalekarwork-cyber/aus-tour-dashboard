"""
Export competitor itineraries to a manager-friendly Excel workbook.

Competitors (e.g. Absolute Australia) are kept in their OWN workbook -- separate
from the tourism-board tours in export_excel.py -- and every row carries a
`Source Type` column flagging it as a competitor, so they never get mixed in.

Sheet "Competitors"     — one row per tour; click the tour name to jump to its
                          full day-by-day plan.
Sheet "View Itinerary"  — every competitor tour listed sequentially.

Reads the unified day-level CSV produced by the competitor scrapers
(webscraping/<source>/data/latest/itinerary_days.csv) which already carry a
`source_type` column.

Usage:
    python analysis/export_competitors.py
    python analysis/export_competitors.py --input webscraping/absolute_australia/data/latest/itinerary_days.csv
    python analysis/export_competitors.py --output analysis/competitor_itineraries.xlsx
"""

import argparse
import csv
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    raise SystemExit("openpyxl is not installed. Run: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
# default: every competitor source's latest day-level CSV
DEFAULT_INPUTS = [
    ROOT / "webscraping" / "absolute_australia" / "data" / "latest" / "itinerary_days.csv",
    ROOT / "webscraping" / "veena_world" / "data" / "latest" / "itinerary_days.csv",
    ROOT / "webscraping" / "sotc" / "data" / "latest" / "itinerary_days.csv",
    ROOT / "webscraping" / "make_my_trip" / "data" / "latest" / "itinerary_days.csv",
]
DEFAULT_OUTPUT = ROOT / "analysis" / "competitor_itineraries.xlsx"

# ── Palette (matches export_excel.py, warmer header to read as "competitor") ──
NAV_BG     = "7B2D26"   # deep red — sheet headers & frozen rows
NAV_FG     = "FFFFFF"
TOUR_BG    = "C0504D"   # medium red — tour header rows in View sheet
TOUR_FG    = "FFFFFF"
ALT_BG     = "FBECEB"   # very light red — alternating data rows
LINK_COLOR = "0563C1"

SOURCE_LABEL = {
    "absolute_australia": "Absolute Australia",
    "scott_dunn":         "Scott Dunn",
    "thomas_cook":        "Thomas Cook India",
    "jacada":             "Jacada Travel",
    "kesari":             "Kesari Tours",
    "veena_world":        "Veena World",
    "sotc":               "SOTC",
    "make_my_trip":       "MakeMyTrip",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Data loading & aggregation ────────────────────────────────────────────────

def load_and_aggregate(csv_paths):
    """Load one or more competitor CSVs and aggregate to tour level."""
    tour_meta: dict = {}
    tour_days: dict = defaultdict(list)

    for csv_path in csv_paths:
        if not Path(csv_path).exists():
            print(f"[export_competitors] skip (missing): {csv_path}")
            continue
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                url  = row["tour_url"].strip()
                name = row["tour_name"].strip()
                if url not in tour_meta:
                    tour_meta[url] = {
                        "tour_name":   name,
                        "source":      SOURCE_LABEL.get(row["source"].strip(), row["source"].strip()),
                        "source_type": row.get("source_type", "Competitor").strip() or "Competitor",
                        "total_days":  row["total_days"].strip(),
                        "tour_url":    url,
                    }
                try:
                    day_num = int(row["day_number"])
                except (ValueError, KeyError):
                    day_num = 0
                tour_days[url].append({
                    "day_number": day_num,
                    "location":   row["location"].strip(),
                    "activity":   row["activity"].strip(),
                })

    for url in tour_days:
        tour_days[url].sort(key=lambda d: d["day_number"])

    tours = sorted(
        tour_meta.values(),
        key=lambda t: (t["source"], t["tour_name"].lower()),
    )
    return tours, tour_days


# ── Workbook builder ──────────────────────────────────────────────────────────

def _write_header_row(ws, row: int, labels: list):
    for offset, label in enumerate(labels):
        c = ws.cell(row, 1 + offset)
        c.value     = label
        c.font      = Font(bold=True, color=NAV_FG, size=10)
        c.fill      = _fill(NAV_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[row].height = 22


def build_workbook(tours: list, tour_days: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 2: "View Itinerary" (built first to know anchor rows) ────────────
    ws2 = wb.active
    ws2.title = "View Itinerary"
    ws2.column_dimensions["A"].width = 7
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 85
    ws2.freeze_panes = "A2"
    _write_header_row(ws2, 1, ["Day", "Location", "What You'll Do"])

    tour_anchor: dict = {}
    cur = 2
    for tour in tours:
        url  = tour["tour_url"]
        days = tour_days.get(url, [])
        tour_anchor[url] = cur

        header_text = (
            f"  {tour['tour_name']}   |   "
            f"{tour['source']}  ({tour['source_type']})   |   "
            f"{tour['total_days']} days"
        )
        ws2.merge_cells(f"A{cur}:C{cur}")
        c = ws2.cell(cur, 1)
        c.value     = header_text
        c.font      = Font(bold=True, color=TOUR_FG, size=10)
        c.fill      = _fill(TOUR_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()
        ws2.row_dimensions[cur].height = 20
        cur += 1

        for idx, day in enumerate(days):
            bg = ALT_BG if idx % 2 == 1 else "FFFFFF"
            for col, val in [(1, day["day_number"]), (2, day["location"]), (3, day["activity"])]:
                c = ws2.cell(cur, col)
                c.value     = val
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = Alignment(vertical="top", wrap_text=(col == 3),
                                        indent=(1 if col != 1 else 0))
            ws2.row_dimensions[cur].height = 15
            cur += 1
        cur += 1   # blank separator

    # ── Sheet 1: "Competitors" ────────────────────────────────────────────────
    # Columns: # | Tour Name | Source | Source Type | Duration | Website
    ws1 = wb.create_sheet("Competitors", 0)
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 42
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 11
    ws1.column_dimensions["F"].width = 12
    ws1.freeze_panes = "A2"

    _write_header_row(ws1, 1, [
        "#",
        "Tour Name  (click to view days)",
        "Source",
        "Source Type",
        "Duration",
        "Website",
    ])
    ws1.auto_filter.ref = f"A1:F{len(tours) + 1}"

    for idx, tour in enumerate(tours):
        row        = idx + 2
        bg         = ALT_BG if idx % 2 == 1 else "FFFFFF"
        anchor_row = tour_anchor.get(tour["tour_url"], 1)

        c = ws1.cell(row, 1)
        c.value = idx + 1
        c.fill = _fill(bg); c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=10)

        c = ws1.cell(row, 2)
        c.value     = tour["tour_name"]
        c.hyperlink = f"#'View Itinerary'!A{anchor_row}"
        c.font      = Font(color=LINK_COLOR, underline="single", size=10)
        c.fill      = _fill(bg); c.border = _border()
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for col_off, val in enumerate([
            tour["source"],
            tour["source_type"],
            f"{tour['total_days']} days",
        ], 3):
            c = ws1.cell(row, col_off)
            c.value = val
            c.fill = _fill(bg); c.border = _border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.font = Font(size=10)

        c = ws1.cell(row, 6)
        c.fill = _fill(bg); c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if tour["tour_url"].startswith("http"):
            c.value     = "View online"
            c.hyperlink = tour["tour_url"]
            c.font      = Font(color=LINK_COLOR, underline="single", size=10)
        else:
            c.value = ""; c.font = Font(size=10)

        ws1.row_dimensions[row].height = 16

    wb.active = ws1
    return wb


# ── Entry point ───────────────────────────────────────────────────────────────

def export(input_paths, output_path: Path = DEFAULT_OUTPUT) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tours, tour_days = load_and_aggregate(input_paths)
    if not tours:
        raise SystemExit("[export_competitors] no competitor rows found — run the scraper first")

    total_days = sum(len(v) for v in tour_days.values())
    print(f"[export_competitors] {len(tours)} tours | {total_days} day rows")
    print("[export_competitors] building workbook ...")
    wb = build_workbook(tours, tour_days)
    wb.save(output_path)
    print(f"[export_competitors] saved -> {output_path}")
    return output_path


def main():
    ap = argparse.ArgumentParser(
        description="Export competitor itineraries to a manager-ready Excel file."
    )
    ap.add_argument("--input", action="append",
                    help="Competitor day-level CSV (repeatable). Defaults to all known competitors.")
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output .xlsx path")
    args = ap.parse_args()
    inputs = [Path(p) for p in args.input] if args.input else DEFAULT_INPUTS
    export(inputs, Path(args.output))


if __name__ == "__main__":
    main()
