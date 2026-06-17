"""
Convert every real data CSV in the project to a sibling .xlsx file.

"Real" = under data/, analysis/, product_app/, or webscraping/*/data/ —
excluding virtualenvs (.venv) and _archive snapshots.

Each CSV becomes a .xlsx of the same name (same folder) with a single sheet,
a bold frozen header row, and an autofilter. Values are read as text so
nothing is silently reinterpreted (leading zeros, long ids, etc.).

Usage:
    python analysis/csv_to_xlsx.py            # convert, keep the CSVs
    python analysis/csv_to_xlsx.py --delete   # convert, then remove the CSVs
    python analysis/csv_to_xlsx.py --dry-run  # just list what would happen
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

# Folders to scan, and folder names to skip anywhere in the path.
INCLUDE_DIRS = ["data", "analysis", "product_app", "webscraping"]
EXCLUDE_PARTS = {".venv", "_archive", ".git", "node_modules"}

HDR_BG = "1F3864"
HDR_FG = "FFFFFF"


def find_csvs() -> list[Path]:
    found: list[Path] = []
    for d in INCLUDE_DIRS:
        base = ROOT / d
        if not base.exists():
            continue
        for p in base.rglob("*.csv"):
            if EXCLUDE_PARTS & set(p.parts):
                continue
            found.append(p)
    return sorted(set(found))


def convert(csv_path: Path) -> Path:
    xlsx_path = csv_path.with_suffix(".xlsx")
    # Read everything as text; don't turn blanks into NaN.
    df = pd.read_csv(
        csv_path,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    )

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]

        # Style + freeze the header, add an autofilter.
        for col_idx in range(1, df.shape[1] + 1):
            c = ws.cell(1, col_idx)
            c.font = Font(bold=True, color=HDR_FG)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        if df.shape[1] >= 1:
            last_col = get_column_letter(df.shape[1])
            ws.auto_filter.ref = f"A1:{last_col}{df.shape[0] + 1}"

        # Reasonable column widths (capped) based on content length.
        for col_idx, col_name in enumerate(df.columns, 1):
            sample = df[col_name].astype(str)
            width = max(len(str(col_name)), int(sample.str.len().head(200).max() or 0))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)

    return xlsx_path


def main() -> None:
    ap = argparse.ArgumentParser(description="Convert project CSVs to .xlsx")
    ap.add_argument("--delete", action="store_true", help="remove each CSV after a successful convert")
    ap.add_argument("--dry-run", action="store_true", help="list files, change nothing")
    args = ap.parse_args()

    csvs = find_csvs()
    print(f"[csv_to_xlsx] found {len(csvs)} CSV file(s)")

    converted = deleted = failed = 0
    for csv_path in csvs:
        rel = csv_path.relative_to(ROOT).as_posix()
        if args.dry_run:
            print(f"  would convert {rel} -> {csv_path.with_suffix('.xlsx').name}"
                  + ("  (+delete csv)" if args.delete else ""))
            continue
        try:
            convert(csv_path)
            converted += 1
            if args.delete:
                csv_path.unlink()
                deleted += 1
            print(f"  ok  {rel}")
        except Exception as exc:  # noqa: BLE001 - report and keep going
            failed += 1
            print(f"  FAIL {rel}: {exc}")

    if not args.dry_run:
        print(f"[csv_to_xlsx] converted={converted} deleted={deleted} failed={failed}")


if __name__ == "__main__":
    main()
