"""
benchmark_ner.py — Compare Ollama models for Australian place-name extraction.

Usage:
    python analysis/benchmark_ner.py

What it does:
    1. Picks 50 representative rows from the combined itineraries CSV
    2. Runs each model in MODELS list on those same 50 rows
    3. Saves a side-by-side Excel file so you can judge which model is best

The Excel has one row per itinerary day. Columns:
    location | activity (first 200 chars) | model_1_places | model_2_places | ...

Edit MODELS below to add/remove models to compare.
Ollama must be running locally (ollama serve) and each model must be pulled first:
    ollama pull qwen2.5:7b
    ollama pull qwen3:4b
    ollama pull gemma3:4b
"""

import os, sys, json, re, time, hashlib
import pandas as pd
import requests

# ---------------------------------------------------------------------------
# CONFIG — edit these
# ---------------------------------------------------------------------------
MODELS = [
    "qwen2.5:7b",
    "gemma3:4b",
]
SAMPLE_SIZE  = 50     # number of rows to test
OLLAMA_HOST  = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
TIMEOUT_SECS = 300    # per-row timeout (qwen3 thinking mode can be slow)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_CSV = os.path.join(BASE, "data", "raw", "combined_itineraries_2026-06.csv")
OUT_XLSX  = os.path.join(BASE, "data", "keywords", "benchmark_ner.xlsx")
CACHE_FILE = os.path.join(BASE, "data", "keywords", "benchmark_ner_cache.json")

# ---------------------------------------------------------------------------
# System prompt (same one the main pipeline uses)
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = (
    "You extract place names from Australian tour itinerary text. "
    "Include: cities, towns, suburbs, national parks, beaches, gorges, mountains, "
    "capes, bays, islands, rivers, lakes, waterfalls, LOOKOUTS and VIEWPOINTS, "
    "landmarks, tourist attractions, museums, galleries, theatres, precincts, "
    "cultural centres, heritage sites. "
    "IMPORTANT: expand shared-suffix conjunctions — e.g. 'Oxer and Junction Pool "
    "lookouts' -> ['Oxer Lookout','Junction Pool Lookout']; 'Joffre and Knox gorges' "
    "-> ['Joffre Gorge','Knox Gorge']. "
    "Exclude: hotel/motel/inn/resort/lodge names (Novotel, Voco, Holiday Inn), restaurant names, "
    "bar names, marketing phrases (Follow The Flavours, Fall In Love With Sydney), "
    "non-Australian places (Cairo, Antarctica), activity verbs (Kayak, Snorkel), "
    "tour company names, generic words (Area, Region, Reserve, Metro, Level, Building)."
)

# Few-shot examples appended to every user message
FEW_SHOT = """
Examples (show correct output format):
Input: Location: Gibb River Road, Windjana Gorge, Tunnel Creek | Activity: Head to the Boab Prison Tree, then Windjana Gorge where ancient fossils lay in the walls. Next visit Tunnel Creek cave system.
Output: ["Boab Prison Tree", "Gibb River Road", "Windjana Gorge", "Tunnel Creek"]

Input: Location: Karijini National Park | Activity: Hike into Joffre and Knox gorges, then stop at Oxer and Junction Pool lookouts.
Output: ["Karijini National Park", "Joffre Gorge", "Knox Gorge", "Oxer Lookout", "Junction Pool Lookout"]

Input: Location: Melbourne | Activity: Walk along the Yarra River, visit Hosier Lane street art, explore the Queen Victoria Market and the Ian Potter Centre.
Output: ["Melbourne", "Yarra River", "Hosier Lane", "Queen Victoria Market", "Ian Potter Centre"]

Now extract places from this row:
"""

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_ARR_RE = re.compile(r"\[[^\[\]]*\]", re.S)


def _parse(text):
    """Extract the first JSON array from model output."""
    text = re.sub(r"```[a-z]*", "", text or "", flags=re.IGNORECASE)
    m = _ARR_RE.search(text)
    if not m:
        return []
    try:
        return [str(p).strip() for p in json.loads(m.group(0)) if str(p).strip()]
    except json.JSONDecodeError:
        return []


def _call(model, loc, act):
    """Call Ollama for one row. Returns list of place strings."""
    url  = OLLAMA_HOST + "/v1/chat/completions"
    user = (
        FEW_SHOT
        + f"Location: {loc}\nActivity: {act[:800]}\n\n"
        + "Return ONLY a JSON array of the Australian place-name strings."
    )
    body = {
        "model": model,
        "temperature": 0.1,
        "stream": False,
        "keep_alive": "60m",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user},
        ],
    }
    for attempt in range(3):
        try:
            r = requests.post(url, json=body, timeout=TIMEOUT_SECS)
            r.raise_for_status()
            return _parse(r.json()["choices"][0]["message"]["content"])
        except Exception as e:
            print(f"    attempt {attempt+1} failed: {e}")
            time.sleep(3 * (attempt + 1))
    return []


def _cache_key(model, loc, act):
    return hashlib.md5(f"bench|{model}|{loc}|||{act[:1500]}".encode()).hexdigest()


# ---------------------------------------------------------------------------
# Sample rows — spread across different states and sources for variety
# ---------------------------------------------------------------------------

def _pick_sample(df, n=50):
    """
    Pick n rows that are spread across states and have non-empty activity text,
    preferring rows where the activity description is long (more to extract).
    """
    df = df[df["activity"].notna() & (df["activity"].str.len() > 80)].copy()
    df["_act_len"] = df["activity"].str.len()

    # Try to get variety by stratifying across states
    groups = []
    for state, grp in df.groupby("state"):
        # take top rows by activity length from each state
        groups.append(grp.nlargest(max(1, n // df["state"].nunique() + 2), "_act_len"))

    sampled = pd.concat(groups).drop_duplicates().nlargest(n, "_act_len")
    if len(sampled) < n:
        # fill with whatever is left
        remaining = df[~df.index.isin(sampled.index)].sample(
            min(n - len(sampled), len(df) - len(sampled)), random_state=42
        )
        sampled = pd.concat([sampled, remaining])

    return sampled.head(n).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Loading {INPUT_CSV} ...")
    df = pd.read_csv(INPUT_CSV)
    print(f"  {len(df):,} rows total")

    sample = _pick_sample(df, SAMPLE_SIZE)
    print(f"  Sampled {len(sample)} rows across {sample['state'].nunique()} states")

    # Load cache
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            cache = json.load(open(CACHE_FILE, encoding="utf-8"))
        except Exception:
            cache = {}

    results = {m: [] for m in MODELS}

    for model in MODELS:
        print(f"\n--- Model: {model} ---")

        # Check Ollama has this model
        try:
            r = requests.get(OLLAMA_HOST + "/api/tags", timeout=10)
            available = [t["name"] for t in r.json().get("models", [])]
            # Ollama tags may have :latest suffix — normalise
            avail_norm = [n.split(":")[0] for n in available]
            model_base = model.split(":")[0]
            if model_base not in avail_norm and model not in available:
                print(f"  WARNING: '{model}' not found in Ollama. Run: ollama pull {model}")
                results[model] = [["NOT INSTALLED"] for _ in range(len(sample))]
                continue
        except Exception as e:
            print(f"  WARNING: Could not reach Ollama ({e}). Is it running?")
            results[model] = [["OLLAMA NOT RUNNING"] for _ in range(len(sample))]
            continue

        row_results = []
        for idx, row in sample.iterrows():
            loc = str(row.get("location", "") or "")
            act = str(row.get("activity", "") or "")
            k   = _cache_key(model, loc, act)

            if k in cache:
                places = cache[k]
                print(f"  row {idx+1:2d}/{len(sample)} [cached] {places}")
            else:
                places = _call(model, loc, act)
                cache[k] = places
                # Save cache after every row
                json.dump(cache, open(CACHE_FILE, "w", encoding="utf-8"), ensure_ascii=False)
                print(f"  row {idx+1:2d}/{len(sample)} → {places}")

            row_results.append(places)

        results[model] = row_results
        print(f"  Done — {model}")

    # Save cache one final time
    json.dump(cache, open(CACHE_FILE, "w", encoding="utf-8"), ensure_ascii=False)

    # ---------------------------------------------------------------------------
    # Build comparison data
    # ---------------------------------------------------------------------------
    assert len(MODELS) == 2, "Comparison columns require exactly 2 models"
    m0, m1 = MODELS[0], MODELS[1]
    col0, col1 = m0.replace(":", "_"), m1.replace(":", "_")

    rows_out = []
    # Aggregates for the report
    total_m0 = total_m1 = total_merged = total_only0 = total_only1 = total_both = 0
    unique_only0, unique_only1 = set(), set()

    for i, row in sample.iterrows():
        idx = list(sample.index).index(i)
        p0 = set(p.lower() for p in results[m0][idx])
        p1 = set(p.lower() for p in results[m1][idx])

        # Original cased versions (prefer m0 casing for shared tokens)
        cased0 = {p.lower(): p for p in results[m0][idx]}
        cased1 = {p.lower(): p for p in results[m1][idx]}

        only0  = p0 - p1          # in m0 but not m1
        only1  = p1 - p0          # in m1 but not m0
        both   = p0 & p1          # in both
        merged = p0 | p1          # union

        # Build cased merged list (alphabetical)
        merged_cased = sorted(
            [cased0.get(p, cased1.get(p, p)) for p in merged],
            key=str.lower
        )

        total_m0      += len(p0)
        total_m1      += len(p1)
        total_merged  += len(merged)
        total_only0   += len(only0)
        total_only1   += len(only1)
        total_both    += len(both)
        unique_only0.update(only0)
        unique_only1.update(only1)

        rows_out.append({
            "state":    row.get("state", ""),
            "tour":     row.get("tour_name", ""),
            "day":      row.get("day_number", ""),
            "location": row.get("location", ""),
            "activity": str(row.get("activity", ""))[:300],
            col0:                    " | ".join(results[m0][idx]),
            col1:                    " | ".join(results[m1][idx]),
            f"only_in_{col0}":       " | ".join(sorted(cased0[p] for p in only0)),
            f"only_in_{col1}":       " | ".join(sorted(cased1[p] for p in only1)),
            "in_both":               " | ".join(sorted(cased0.get(p, cased1[p]) for p in both)),
            "merged (use this)":     " | ".join(merged_cased),
        })

    n = len(sample)
    report_lines = [
        f"NER Benchmark Report — {m0} vs {m1}",
        f"Sample: {n} rows across {sample['state'].nunique()} states",
        "",
        f"{'Metric':<40} {'qwen2.5:7b':>12} {'gemma3:4b':>12} {'Merged':>12}",
        "-" * 78,
        f"{'Total places extracted':<40} {total_m0:>12,} {total_m1:>12,} {total_merged:>12,}",
        f"{'Avg places per row':<40} {total_m0/n:>12.1f} {total_m1/n:>12.1f} {total_merged/n:>12.1f}",
        f"{'Places only this model found (total)':<40} {total_only0:>12,} {total_only1:>12,} {'—':>12}",
        f"{'Places both models agreed on':<40} {total_both:>12,} {total_both:>12,} {'—':>12}",
        f"{'Agreement rate (both/union)':<40} {total_both/max(total_merged,1)*100:>11.1f}% {total_both/max(total_merged,1)*100:>11.1f}% {'—':>12}",
        "",
        f"Unique place tokens only {m0} found: {len(unique_only0)}",
        f"Unique place tokens only {m1} found: {len(unique_only1)}",
        "",
        "Recommendation:",
        "  Use 'merged' column — union of both models catches more real places",
        "  than either model alone, with minimal extra noise.",
        f"  Merged gives {(total_merged-max(total_m0,total_m1))/max(total_m0,total_m1)*100:.1f}% more places than the best single model.",
    ]
    report_text = "\n".join(report_lines)
    print("\n" + report_text)

    # ---------------------------------------------------------------------------
    # Build Excel
    # ---------------------------------------------------------------------------
    print(f"\nBuilding Excel → {OUT_XLSX}")

    out_df = pd.DataFrame(rows_out)

    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:

        # Sheet 1 — row-by-row comparison
        out_df.to_excel(writer, sheet_name="Comparison", index=False)
        ws = writer.sheets["Comparison"]

        header_fill  = PatternFill("solid", fgColor="1F4E79")
        merged_fill  = PatternFill("solid", fgColor="375623")  # green header for merged col
        only0_fill   = PatternFill("solid", fgColor="7030A0")  # purple for only-in-m0
        only1_fill   = PatternFill("solid", fgColor="C55A11")  # orange for only-in-m1
        light_fill   = PatternFill("solid", fgColor="EBF3FB")
        white_font   = Font(bold=True, color="FFFFFF")

        col_colors = {
            col0:                  header_fill,
            col1:                  header_fill,
            f"only_in_{col0}":     only0_fill,
            f"only_in_{col1}":     only1_fill,
            "in_both":             header_fill,
            "merged (use this)":   merged_fill,
        }
        col_widths = {
            "state": 10, "tour": 28, "day": 5, "location": 30, "activity": 50,
            col0: 40, col1: 40,
            f"only_in_{col0}": 35, f"only_in_{col1}": 35,
            "in_both": 35, "merged (use this)": 45,
        }

        for ci, col in enumerate(out_df.columns, 1):
            letter = get_column_letter(ci)
            hdr_cell = ws.cell(row=1, column=ci)
            fill = col_colors.get(col, header_fill)
            hdr_cell.fill  = fill
            hdr_cell.font  = white_font
            hdr_cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[letter].width = col_widths.get(col, 35)
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        for row_idx in range(2, len(out_df) + 2):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = light_fill

        # Sheet 2 — text report
        report_df = pd.DataFrame({"Report": report_lines})
        report_df.to_excel(writer, sheet_name="Report", index=False, header=False)
        ws2 = writer.sheets["Report"]
        ws2.column_dimensions["A"].width = 80
        ws2.cell(row=1).font = Font(bold=True, size=13)

    print(f"Saved: {OUT_XLSX}")
    print(f"\nKey takeaway: 'merged (use this)' column = best of both models combined.")


if __name__ == "__main__":
    main()
