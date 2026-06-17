"""
Export competitor itineraries to a manager-friendly Excel workbook.

Competitors are tour *operators* (source_type == "Competitor") scraped into a
location-block schema -- one row per day block:

    source | source_type | tour_name | tour_url | total_days |
    day_number | location | place | description | experiences | scrape_date

  - location    : the block's main title (e.g. "Sydney")
  - place        : "At a Glance"
  - description : the block's "At a Glance" overview text followed by a
                  "More Experiences in This Area:" list of experience titles
  - experiences : the same titles, "; "-joined (a filterable column)

Sheets:
  "Tours"          — one row per competitor tour; click the name to jump to its
                     full day-by-day plan in the View Itinerary sheet.
  "View Itinerary" — every tour listed sequentially: Day / Location / Place /
                     Description.

Usage:
    python analysis/export_competitors_excel.py
    python analysis/export_competitors_excel.py --input webscraping/scott_dunn/data/latest/itinerary_days.csv
    python analysis/export_competitors_excel.py --output analysis/competitor_itineraries.xlsx
"""

import argparse
import csv
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    raise SystemExit("openpyxl is not installed. Run: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "webscraping" / "scott_dunn" / "data" / "latest" / "itinerary_days.csv"
DEFAULT_OUTPUT = ROOT / "analysis" / "competitor_itineraries.xlsx"

# ── Palette (shared with export_excel.py) ─────────────────────────────────────
NAV_BG     = "1F3864"   # dark navy  — sheet headers
NAV_FG     = "FFFFFF"
TOUR_BG    = "2E75B6"   # medium blue — tour header rows
TOUR_FG    = "FFFFFF"
ALT_BG     = "EBF3FB"   # very light blue — alternating rows
LINK_COLOR = "0563C1"

SOURCE_LABEL = {
    "scott_dunn":         "Scott Dunn",
    "absolute_australia": "Absolute Australia",
    "thomas_cook":        "Thomas Cook India",
    "jacada":             "Jacada Travel",
    "kesari":             "Kesari Tours",
    "veena_world":        "Veena World",
    "sotc":               "SOTC",
    "make_my_trip":       "MakeMyTrip",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Data loading ──────────────────────────────────────────────────────────────
def load(csv_path: Path):
    tour_meta: dict = {}
    tour_rows: dict = defaultdict(list)

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            url = (row.get("tour_url") or "").strip()
            if not url:
                continue
            if url not in tour_meta:
                src = (row.get("source") or "").strip()
                tour_meta[url] = {
                    "tour_url":   url,
                    "tour_name":  (row.get("tour_name") or "").strip(),
                    "source":     SOURCE_LABEL.get(src, src),
                    "total_days": (row.get("total_days") or "").strip(),
                }
            tour_rows[url].append({
                "day_number":  (row.get("day_number") or "").strip(),
                "location":    (row.get("location") or "").strip(),
                "place":       (row.get("place") or "").strip(),
                "description": (row.get("description") or "").strip(),
                "experiences": (row.get("experiences") or "").strip(),
            })

    for url, meta in tour_meta.items():
        # count distinct experiences mentioned across the tour
        exps = set()
        for r in tour_rows[url]:
            exps.update(e.strip() for e in r["experiences"].split(";") if e.strip())
        meta["n_exp"] = len(exps)
        meta["n_blocks"] = len(tour_rows[url])

    tours = sorted(tour_meta.values(),
                   key=lambda t: (t["source"], t["tour_name"].lower()))
    return tours, tour_rows


# ── Workbook ──────────────────────────────────────────────────────────────────
def _header_row(ws, row: int, labels: list):
    for off, label in enumerate(labels):
        c = ws.cell(row, 1 + off)
        c.value     = label
        c.font      = Font(bold=True, color=NAV_FG, size=10)
        c.fill      = _fill(NAV_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[row].height = 22


def _approx_row_height(text: str, width_chars: int) -> float:
    """Rough wrapped-row height so long descriptions are readable."""
    lines = 0
    for para in (text or "").split("\n"):
        lines += max(1, -(-len(para) // max(1, width_chars)))  # ceil division
    return min(420, max(15, lines * 14 + 4))


def build_workbook(tours: list, tour_rows: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 2: View Itinerary (built first to know anchor rows) ─────────────
    ws2 = wb.active
    ws2.title = "View Itinerary"
    ws2.column_dimensions["A"].width = 7      # Day
    ws2.column_dimensions["B"].width = 28     # Location
    ws2.column_dimensions["C"].width = 16     # Place
    ws2.column_dimensions["D"].width = 110    # Description
    ws2.freeze_panes = "A2"
    _header_row(ws2, 1, ["Day", "Location", "Place", "Description"])

    anchor: dict = {}
    cur = 2
    for tour in tours:
        url  = tour["tour_url"]
        rows = tour_rows.get(url, [])
        anchor[url] = cur

        ws2.merge_cells(f"A{cur}:D{cur}")
        c = ws2.cell(cur, 1)
        c.value     = (f"  {tour['tour_name']}   |   {tour['source']}   "
                       f"(Competitor)   |   {tour['total_days']} days   |   "
                       f"{tour['n_exp']} experiences")
        c.font      = Font(bold=True, color=TOUR_FG, size=10)
        c.fill      = _fill(TOUR_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()
        ws2.row_dimensions[cur].height = 20
        cur += 1

        for i, r in enumerate(rows):
            bg = ALT_BG if i % 2 == 1 else "FFFFFF"
            for col, val, wrap in [
                (1, r["day_number"], False),
                (2, r["location"],   True),
                (3, r["place"],      True),
                (4, r["description"], True),
            ]:
                c = ws2.cell(cur, col)
                c.value     = val
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = Alignment(
                    vertical="top",
                    horizontal="center" if col == 1 else "left",
                    wrap_text=wrap,
                    indent=(1 if col in (2, 3, 4) else 0),
                )
                c.font = Font(size=10, bold=(col in (1, 2)))
            ws2.row_dimensions[cur].height = _approx_row_height(r["description"], 95)
            cur += 1
        cur += 1   # blank separator between tours

    # ── Sheet 1: Tours ────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Tours", 0)
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 44
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 11
    ws1.column_dimensions["E"].width = 13
    ws1.column_dimensions["F"].width = 12
    ws1.freeze_panes = "A2"
    _header_row(ws1, 1, ["#", "Tour Name  (click to view days)", "Competitor",
                         "Duration", "Experiences", "Website"])
    ws1.auto_filter.ref = f"A1:F{len(tours) + 1}"

    for idx, tour in enumerate(tours):
        row = idx + 2
        bg  = ALT_BG if idx % 2 == 1 else "FFFFFF"
        a   = anchor.get(tour["tour_url"], 1)

        c = ws1.cell(row, 1)
        c.value = idx + 1
        c.fill = _fill(bg); c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center"); c.font = Font(size=10)

        c = ws1.cell(row, 2)
        c.value     = tour["tour_name"]
        c.hyperlink = f"#'View Itinerary'!A{a}"
        c.font      = Font(color=LINK_COLOR, underline="single", size=10)
        c.fill = _fill(bg); c.border = _border()
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for col_off, val in enumerate([
            tour["source"],
            f"{tour['total_days']} days",
            tour["n_exp"],
        ], 3):
            c = ws1.cell(row, col_off)
            c.value = val
            c.fill = _fill(bg); c.border = _border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.font = Font(size=10)

        c = ws1.cell(row, 6)
        c.fill = _fill(bg); c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if tour["tour_url"].startswith("http"):
            c.value = "View online"
            c.hyperlink = tour["tour_url"]
            c.font = Font(color=LINK_COLOR, underline="single", size=10)
        else:
            c.value = ""; c.font = Font(size=10)

        ws1.row_dimensions[row].height = 16

    wb.active = ws1
    return wb


def export(input_path: Path = DEFAULT_INPUT, output_path: Path = DEFAULT_OUTPUT) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[export_competitors] loading  {input_path}")
    tours, tour_rows = load(input_path)
    total_rows = sum(len(v) for v in tour_rows.values())
    print(f"[export_competitors] {len(tours)} tours | {total_rows} block rows")

    wb = build_workbook(tours, tour_rows)
    wb.save(output_path)
    print(f"[export_competitors] saved -> {output_path}")
    return output_path


def main():
    ap = argparse.ArgumentParser(description="Export competitor itineraries to Excel.")
    ap.add_argument("--input",  default=str(DEFAULT_INPUT))
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = ap.parse_args()
    export(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
