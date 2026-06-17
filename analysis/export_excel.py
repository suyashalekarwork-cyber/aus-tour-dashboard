"""
Export combined itineraries to a manager-friendly Excel workbook.

Sheet "Tours"          — one row per tour; click the tour name to jump to its
                         full day-by-day plan in the View Itinerary sheet.
Sheet "View Itinerary" — every tour listed sequentially with full day plans.

Usage (standalone):
    python analysis/export_excel.py
    python analysis/export_excel.py --input data/raw/combined_itineraries_latest.csv
    python analysis/export_excel.py --output analysis/itineraries_for_manager.xlsx
"""

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl is not installed. Run: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT  = ROOT / "data" / "raw" / "combined_itineraries_latest.csv"
DEFAULT_OUTPUT = ROOT / "analysis" / "itineraries_for_manager.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
NAV_BG     = "1F3864"   # dark navy  — sheet headers & frozen rows
NAV_FG     = "FFFFFF"
TOUR_BG    = "2E75B6"   # medium blue — tour header rows in View sheet
TOUR_FG    = "FFFFFF"
ALT_BG     = "EBF3FB"   # very light blue — alternating data rows
LINK_COLOR = "0563C1"   # standard Excel hyperlink blue

SOURCE_LABEL = {
    "global_journeys":      "Global Journeys",
    "inside_australia":     "Inside Australia",
    "australia_com":        "Australia.com",
    "discover_tasmania":    "Discover Tasmania",
    "queensland_com":       "Queensland.com",
    "sydney_com":           "Sydney.com",
    "westernaustralia_com": "Tourism Western Australia",
    "visitvictoria_com":    "Visit Victoria",
}

# Australian state/territory codes — anything else is international
AU_STATES = {"NSW", "VIC", "QLD", "WA", "SA", "TAS", "NT", "ACT"}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Data loading & aggregation ────────────────────────────────────────────────

def load_and_aggregate(csv_path: Path):
    """
    Load CSV and aggregate to tour level.
    Tours with any day tagged outside AU_STATES (e.g. South Pacific,
    International) are excluded entirely.

    Returns:
        tours     — sorted list of AU-only tour-level dicts
        tour_days — dict  tour_url -> sorted list of day dicts
    """
    tour_meta: dict = {}
    tour_days: dict = defaultdict(list)

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            # Competitors live in their own workbook (export_competitors.py);
            # keep them out of the manager tour inventory.
            if row.get("source_type", "").strip() == "Competitor":
                continue

            url  = row["tour_url"].strip()
            name = row["tour_name"].strip()

            if url not in tour_meta:
                tour_meta[url] = {
                    "tour_name":  name,
                    "source":     SOURCE_LABEL.get(row["source"].strip(), row["source"].strip()),
                    "total_days": row["total_days"].strip(),
                    "tour_url":   url,
                    "states":     set(),
                    "cities":     [],   # insertion-ordered unique list
                    "has_intl":   False,
                }

            meta  = tour_meta[url]
            state = row["state"].strip()

            if state:
                meta["states"].add(state)
                if state not in AU_STATES:
                    meta["has_intl"] = True

            city = row["city"].strip()
            if city and city not in meta["cities"]:
                meta["cities"].append(city)

            try:
                day_num = int(row["day_number"])
            except (ValueError, KeyError):
                day_num = 0

            tour_days[url].append({
                "day_number": day_num,
                "location":   row["location"].strip(),
                "activity":   row["activity"].strip(),
            })

    # Drop tours that include any non-Australian destination
    excluded = sum(1 for t in tour_meta.values() if t["has_intl"])
    tour_meta = {url: t for url, t in tour_meta.items() if not t["has_intl"]}
    if excluded:
        print(f"[export_excel] excluded {excluded} international/South Pacific tours")

    # Sort days within each tour
    for url in tour_days:
        tour_days[url].sort(key=lambda d: d["day_number"])

    # Sort tours: source order, then alphabetically by name
    source_order = {
        "Global Journeys": 0, "Inside Australia": 1, "Australia.com": 2,
        "Discover Tasmania": 3, "Queensland.com": 4, "Sydney.com": 5,
        "Tourism Western Australia": 6, "Visit Victoria": 7,
    }
    tours = sorted(
        tour_meta.values(),
        key=lambda t: (source_order.get(t["source"], 9), t["tour_name"].lower()),
    )

    # Finalise states (AU only) and cities strings
    for t in tours:
        t["states"] = ", ".join(sorted(t["states"] & AU_STATES))
        t["cities"] = ", ".join(t["cities"][:6])   # cap at 6 to keep cells readable

    return tours, tour_days

# ── Workbook builder ──────────────────────────────────────────────────────────

def _write_header_row(ws, row: int, labels: list):
    for offset, label in enumerate(labels):
        c = ws.cell(row, 1 + offset)
        c.value     = label
        c.font      = Font(bold=True, color=NAV_FG, size=10)
        c.fill      = _fill(NAV_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[row].height = 22


def build_workbook(tours: list, tour_days: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 2: "View Itinerary" ─────────────────────────────────────────────
    # Build first so we know the exact row each tour lands on, then use those
    # row numbers as hyperlink targets from Sheet 1.

    ws2 = wb.active
    ws2.title = "View Itinerary"

    ws2.column_dimensions["A"].width = 7    # Day #
    ws2.column_dimensions["B"].width = 30   # Location
    ws2.column_dimensions["C"].width = 85   # Activity

    ws2.freeze_panes = "A2"
    _write_header_row(ws2, 1, ["Day", "Location", "What You'll Do"])

    tour_anchor: dict = {}   # tour_url -> row number of its blue header
    cur = 2

    for tour in tours:
        url  = tour["tour_url"]
        days = tour_days.get(url, [])

        tour_anchor[url] = cur

        # Blue tour header row (merged across all three columns)
        header_text = (
            f"  {tour['tour_name']}   |   "
            f"{tour['source']}   |   "
            f"{tour['total_days']} days"
        )
        ws2.merge_cells(f"A{cur}:C{cur}")
        c = ws2.cell(cur, 1)
        c.value     = header_text
        c.font      = Font(bold=True, color=TOUR_FG, size=10)
        c.fill      = _fill(TOUR_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()
        ws2.row_dimensions[cur].height = 20
        cur += 1

        # Day rows
        for idx, day in enumerate(days):
            bg = ALT_BG if idx % 2 == 1 else "FFFFFF"
            for col, val in [
                (1, day["day_number"]),
                (2, day["location"]),
                (3, day["activity"]),
            ]:
                c = ws2.cell(cur, col)
                c.value     = val
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = Alignment(
                    vertical="top",
                    wrap_text=(col == 3),
                    indent=(1 if col != 1 else 0),
                )
            ws2.row_dimensions[cur].height = 15
            cur += 1

        cur += 1   # blank separator between tours

    # ── Sheet 1: "Tours" ──────────────────────────────────────────────────────
    # Columns: # | Tour Name | Source | Duration | States | Cities | Website

    ws1 = wb.create_sheet("Tours", 0)   # first/leftmost tab

    ws1.column_dimensions["A"].width = 5    # #
    ws1.column_dimensions["B"].width = 42   # Tour Name (hyperlink)
    ws1.column_dimensions["C"].width = 24   # Source
    ws1.column_dimensions["D"].width = 11   # Duration
    ws1.column_dimensions["E"].width = 20   # States
    ws1.column_dimensions["F"].width = 38   # Cities
    ws1.column_dimensions["G"].width = 12   # Website

    ws1.freeze_panes = "A2"

    _write_header_row(ws1, 1, [
        "#",
        "Tour Name  (click to view days)",
        "Source",
        "Duration",
        "States",
        "Cities",
        "Website",
    ])

    # AutoFilter on all 7 columns
    ws1.auto_filter.ref = f"A1:G{len(tours) + 1}"

    for idx, tour in enumerate(tours):
        row        = idx + 2
        bg         = ALT_BG if idx % 2 == 1 else "FFFFFF"
        anchor_row = tour_anchor.get(tour["tour_url"], 1)

        # A — sequence number
        c           = ws1.cell(row, 1)
        c.value     = idx + 1
        c.fill      = _fill(bg)
        c.border    = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font      = Font(size=10)

        # B — Tour Name as internal hyperlink to its section in View Itinerary
        c           = ws1.cell(row, 2)
        c.value     = tour["tour_name"]
        c.hyperlink = f"#'View Itinerary'!A{anchor_row}"
        c.font      = Font(color=LINK_COLOR, underline="single", size=10)
        c.fill      = _fill(bg)
        c.border    = _border()
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # C–F — Source, Duration, States, Cities
        for col_off, val in enumerate([
            tour["source"],
            f"{tour['total_days']} days",
            tour["states"],
            tour["cities"],
        ], 3):
            c           = ws1.cell(row, col_off)
            c.value     = val
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.font      = Font(size=10)

        # G — "View online" external link to source URL
        c           = ws1.cell(row, 7)
        c.fill      = _fill(bg)
        c.border    = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if tour["tour_url"] and tour["tour_url"].startswith("http"):
            c.value     = "View online"
            c.hyperlink = tour["tour_url"]
            c.font      = Font(color=LINK_COLOR, underline="single", size=10)
        else:
            c.value = ""
            c.font  = Font(size=10)

        ws1.row_dimensions[row].height = 16

    wb.active = ws1
    return wb

# ── Entry point ───────────────────────────────────────────────────────────────

def export(input_path: Path = DEFAULT_INPUT, output_path: Path = DEFAULT_OUTPUT) -> Path:
    """Core export function — callable from the pipeline."""
    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[export_excel] loading  {input_path.name}")
    tours, tour_days = load_and_aggregate(input_path)
    total_days = sum(len(v) for v in tour_days.values())
    print(f"[export_excel] {len(tours)} tours | {total_days} day rows")

    print("[export_excel] building workbook ...")
    wb = build_workbook(tours, tour_days)
    wb.save(output_path)
    print(f"[export_excel] saved -> {output_path}")
    return output_path


def main():
    ap = argparse.ArgumentParser(
        description="Export combined itineraries to a manager-ready Excel file."
    )
    ap.add_argument("--input",  default=str(DEFAULT_INPUT),  help="Path to combined CSV")
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output .xlsx path")
    args = ap.parse_args()
    export(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
