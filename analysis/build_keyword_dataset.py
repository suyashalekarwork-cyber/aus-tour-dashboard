"""
Build a keyword dataset from itinerary activity descriptions.

Layers:
  1. Rules-based place extraction (existing extractor.py)
  2. spaCy NER for places missed by regex (GPE, LOC, FAC, EVENT)
  3. Named-activity pattern (e.g. "Daintree River Cruise", "Mimbi Caves Guided Tour")
  4. Allowlist — custom phrases from data/keyword_allowlist.txt (e.g. "Boab Prison Tree")

Manual control files (edit these to tune output):
  data/config/keyword_blacklist.txt  — tags to always remove (generic phrases)
  data/config/keyword_allowlist.txt  — custom places to always include if found in text
  data/config/keyword_synonyms.txt   — canonical name mappings (e.g. Ayers Rock → Uluru)

Output:
  data/keywords/keyword_dataset_2026-06.csv   — full tagged dataset
  data/keywords/keyword_review.csv            — all unique tags with frequency + example, for manual review
"""

import re
import sys
import os
import json
import hashlib
import time
import pandas as pd
import spacy
from collections import Counter, defaultdict

# Allow importing from product_app and this script's own directory (analysis/)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.dirname(__file__))
from product_app.extractor import extract_products, extract_themes, clean_text
from xlsx_io import write_xlsx_safe

# ---------------------------------------------------------------------------
# Manual config loaders
# ---------------------------------------------------------------------------

def _load_lines(path):
    """Load non-empty, non-comment lines from a text file. Returns lowercase set."""
    if not os.path.exists(path):
        return set()
    with open(path, encoding="utf-8") as f:
        lines = [l.strip() for l in f if l.strip() and not l.strip().startswith("#")]
    return set(l.lower() for l in lines)


def _load_allowlist(path):
    """Load allowlist lines, preserving original casing for output."""
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        lines = [l.strip() for l in f if l.strip() and not l.strip().startswith("#")]
    return lines  # list of original-cased phrases


# ---------------------------------------------------------------------------
# Layer 3 — named activity patterns
# ---------------------------------------------------------------------------
_ACTIVITY_TYPES = (
    r"(?:Tour|Cruise|Safari|Walk|Trek|Flight|Experience|Workshop|Tasting"
    r"|Ceremony|Performance|Show|Swim|Dive|Drive|Journey|Expedition)"
)
_ACTIVITY_RE = re.compile(
    r"\b([A-Z][a-zA-Z''\-]+(?:\s+[A-Z][a-zA-Z''\-]+){1,5}\s+" + _ACTIVITY_TYPES + r")\b"
)

# spaCy entity types we trust as useful place references
_SPACY_KEEP = {"GPE", "LOC", "FAC", "EVENT"}

# ORG entities are kept only when they contain a venue word —
# spaCy (en_core_web_lg) can still mislabel museums, galleries, theatres, hotels as ORG
_VENUE_WORDS = re.compile(
    r"\b(museum|gallery|galleries|theatre|theater|cinema|arena|stadium|"
    r"centre|center|precinct|institute|aquarium|zoo|planetarium|"
    r"hotel|resort|lodge|retreat|motel|inn)\b",
    re.IGNORECASE,
)

# Short noise words / generic phrases spaCy sometimes mis-tags
_SPACY_NOISE = {
    "today", "morning", "evening", "afternoon", "night", "day",
    "australia", "australian",
    "a", "the", "an",
    "national park", "conservation park", "marine park",
    "indigenous", "aboriginal",
    "outback",
}

_LEADING_ARTICLE  = re.compile(r"^(?:the|a|an)\s+", re.IGNORECASE)
_TRAILING_POSSESS = re.compile(r"[‘’’]s?$", re.IGNORECASE)  # strip "Australia’s" → "Australia"
_PARTIAL_PLACE    = re.compile(r"^(?:Lake|Mount|Mt|Cape|Port|Gulf)\s+\w{1,2}$")  # e.g. "Lake St"
_PREFIX_PREP      = re.compile(r"^(?:at|from|to|in|on|of)\s+", re.IGNORECASE)
_SENTENCE_CONNECTOR = re.compile(r"\b(and|for|before|after|with|your|its)\b", re.IGNORECASE)


def _is_sentence_like(token: str) -> bool:
    """True when a location token looks like prose rather than a place name."""
    return len(token.split()) > 5 and bool(_SENTENCE_CONNECTOR.search(token))

# ---------------------------------------------------------------------------
# Location tokeniser
# ---------------------------------------------------------------------------

# Words/phrases that prefix or suffix a place name — strip these, keep the place
_LOC_STRIP_PREFIX = re.compile(
    r"^(?:arrive\s+|depart\s+|departing\s+|arriving\s+|leaving\s+"
    r"|welcome\s+to\s+|farewell\s+from\s+|fly\s+to\s+|fly\s+from\s+"
    r"|drive\s+to\s+|travel\s+to\s+|transfer\s+to\s+|head\s+to\s+"
    r"|return\s+to\s+|back\s+to\s+)",
    re.IGNORECASE,
)
_LOC_STRIP_SUFFIX = re.compile(
    r"\s+(?:free\s+time|airport|transfer|overnight|day\s+trip)$",
    re.IGNORECASE,
)
# Journey separator: spaced dash (but NOT tight dash, which is part of place names)
_JOURNEY_SEP = re.compile(r"\s+-\s+")

# Things that look like activities embedded in the location field
_LOC_ACTIVITY_RE = re.compile(
    r"\b(?:cruise|tour|flight|drive|walk|trek|safari|experience|show|performance)\b",
    re.IGNORECASE,
)


def tokenize_location(raw: str) -> list[str]:
    """
    Split a raw location string into individual clean place tokens.

    Rules (applied in order):
      1. Split on journey separator  ` - `  (spaced dash = route, e.g. "Hobart - Strahan")
      2. Split each chunk on `,` or `/`
      3. For each token: strip navigation prefixes (Arrive, Depart, Welcome to…)
         and suffixes (Free Time, Airport…)
      4. Drop tokens that look like embedded activities (contain "Cruise", "Tour", etc.)
      5. Drop tokens that are too short or all-lowercase (noise)
    Returns a list of clean place name strings.
    """
    if not raw or not isinstance(raw, str):
        return []

    # Step 1 — split on journey separator (spaced dash)
    chunks = _JOURNEY_SEP.split(raw)

    tokens = []
    for chunk in chunks:
        # Step 2 — split on comma or slash
        for part in re.split(r"[,/]", chunk):
            part = part.strip()
            if not part:
                continue

            # Step 3 — strip navigation noise
            part = _LOC_STRIP_PREFIX.sub("", part).strip()
            part = _LOC_STRIP_SUFFIX.sub("", part).strip()
            part = _LEADING_ARTICLE.sub("", part).strip()

            if not part or len(part) < 3:
                continue

            # Step 4 — skip tokens that are really activities embedded in location
            if _LOC_ACTIVITY_RE.search(part):
                continue

            # Step 5 — skip pure lowercase tokens (generic noise like "overnight")
            if part == part.lower():
                continue

            tokens.append(part)

    # Deduplicate while preserving order
    seen = set()
    result = []
    for t in tokens:
        key = t.lower()
        if key not in seen:
            seen.add(key)
            result.append(t)

    # Drop tokens that look like prose sentences — these will be handled by location NER instead
    result = [t for t in result if not _is_sentence_like(t)]

    return result


def _extract_spacy_entities(doc):
    results = []
    for ent in doc.ents:
        is_venue_org = ent.label_ == "ORG" and _VENUE_WORDS.search(ent.text)
        if ent.label_ not in _SPACY_KEEP and not is_venue_org:
            continue
        text = _LEADING_ARTICLE.sub("", ent.text.strip()).strip()
        text = _TRAILING_POSSESS.sub("", text).strip()
        if len(text) < 4:
            continue
        if text.lower() in _SPACY_NOISE:
            continue
        if " " not in text and text[0].islower():
            continue
        if _PARTIAL_PLACE.match(text):  # skip "Lake St", "Mt Rd" etc.
            continue
        results.append(text)
    return results


def _extract_activity_tags(text):
    return _ACTIVITY_RE.findall(text or "")


def _allowlist_matches(text, allowlist):
    """Return allowlist phrases that appear in text (case-insensitive)."""
    text_lower = text.lower()
    return [phrase for phrase in allowlist if phrase.lower() in text_lower]


def _clean_tag(tag: str) -> str:
    """Strip possessives, leading prepositions, and normalise case."""
    tag = _TRAILING_POSSESS.sub("", tag).strip()   # "Australia's" → "Australia"
    tag = _PREFIX_PREP.sub("", tag).strip()         # "At Manning Creek" → "Manning Creek"
    if tag and (tag == tag.lower() or tag == tag.upper()):
        tag = tag.title()                           # "east coast" → "East Coast"
    return tag


def _load_synonyms(path):
    """Load synonym map from file. Each line: old_name -> canonical_name"""
    synonyms = {}
    if not os.path.exists(path):
        return synonyms
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "->" in line:
                old, new = line.split("->", 1)
                synonyms[old.strip().lower()] = new.strip()
    return synonyms


def _apply_synonyms(tags, synonyms):
    """Replace tags with canonical names, re-deduplicating afterwards."""
    if not synonyms:
        return tags
    seen = {}
    for tag in tags:
        canonical = synonyms.get(tag.lower(), tag)
        key = canonical.lower()
        if key not in seen:
            seen[key] = canonical
    return sorted(seen.values(), key=lambda x: x.lower())


def _apply_blacklist(tags, blacklist):
    """
    Remove tags that match the blacklist.

    Matching rules:
      - Single-word blacklist entries: exact match only.
        e.g. "Creek" blocks the tag "Creek" but NOT "Tunnel Creek".
      - Multi-word blacklist entries: exact match OR the tag is entirely composed of
        the blacklisted phrase (no extra specific name attached).
        e.g. "National Park" blocks "National Park" but NOT "Kakadu National Park".
    """
    single_word_bl = {b for b in blacklist if " " not in b}
    multi_word_bl  = {b for b in blacklist if " " in b}

    result = []
    for tag in tags:
        tag_lower = tag.lower()
        # Exact match for any entry
        if tag_lower in blacklist:
            continue
        # Multi-word: drop only if the tag IS exactly that phrase (already handled above)
        # — no endswith/startswith substring matching to avoid false positives like
        # "Tunnel Creek" being dropped because "Creek" is blacklisted.
        result.append(tag)
    return result


def _merge_tags(*tag_lists):
    """Merge multiple tag lists, clean and deduplicate case-insensitively, return sorted."""
    seen = {}
    for lst in tag_lists:
        for tag in lst:
            tag = _clean_tag(tag.strip())
            if not tag or len(tag) <= 2:
                continue
            key = tag.lower()
            if key not in seen:
                seen[key] = tag
    return sorted(seen.values(), key=lambda x: x.lower())


# ---------------------------------------------------------------------------
# Review CSV builder
# ---------------------------------------------------------------------------

def _build_review_csv(df, output_path):
    """
    Write data/keyword_review.csv with:
      tag, tag_type, frequency, example_tour, example_day, example_snippet
    So the user can open it in Excel and decide what to blacklist/allowlist.
    """
    rows = []
    for _, r in df.iterrows():
        tour = r.get("tour_name", "")
        day  = r.get("day_number", "")
        act  = str(r.get("activity", ""))[:120]

        url = r.get("tour_url", "")
        for col, tag_type in [("place_tags", "place"), ("activity_tags", "activity"), ("theme_tags", "theme")]:
            cell = r.get(col)
            if pd.isna(cell):
                continue
            for tag in str(cell).split("; "):
                tag = tag.strip()
                if tag:
                    rows.append({"tag": tag, "tag_type": tag_type,
                                 "example_tour": tour, "example_url": url,
                                 "example_day": day,
                                 "example_snippet": act})

    if not rows:
        return

    review_df = pd.DataFrame(rows)

    # Count frequency per tag
    freq = review_df.groupby("tag").size().rename("frequency")

    # One example per tag (first occurrence)
    examples = review_df.drop_duplicates("tag").set_index("tag")[
        ["tag_type", "example_tour", "example_url", "example_day", "example_snippet"]
    ]

    result = examples.join(freq).reset_index()
    result = result.sort_values(["tag_type", "frequency"], ascending=[True, False])
    result = result[["tag", "tag_type", "frequency", "example_tour", "example_url", "example_day", "example_snippet"]]

    result.to_csv(output_path, index=False)
    print(f"Review file saved to {output_path}")
    print(f"  {len(result):,} unique tags — open in Excel to review, then add unwanted tags to keyword_blacklist.txt")
    return result


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _build_excel(df, review_df, output_path):
    """
    Write a multi-sheet Excel workbook:
      Sheet 1 - Dataset      : full tagged itinerary rows
      Sheet 2 - Place Tags   : unique place tags, frequency, example
      Sheet 3 - Activity Tags: unique activity tags, frequency, example
      Sheet 4 - Theme Tags   : unique theme tags, frequency, example
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    def _style_header(ws, header_color="1F4E79"):
        fill = PatternFill("solid", fgColor=header_color)
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"

    def _autofit(ws, max_width=60):
        for col_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, max_width)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # Sheet 1 — full dataset (key columns first, original text last)
        display_cols = [
            "tour_name", "day_number", "location", "location_tokens",
            "place_tags", "activity_tags", "theme_tags", "keyword_count",
            "all_keywords", "has_content", "activity",
            "source", "tour_url", "price", "state", "city", "city_type",
        ]
        display_cols = [c for c in display_cols if c in df.columns]
        df[display_cols].to_excel(writer, sheet_name="Dataset", index=False)
        ws = writer.sheets["Dataset"]
        _style_header(ws)
        # Wrap text for tag columns, wider width
        for col_idx, col_name in enumerate(display_cols, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in ("place_tags", "activity_tags", "all_keywords", "activity"):
                ws.column_dimensions[col_letter].width = 55
                for cell in ws[col_letter]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_name in ("tour_name", "location"):
                ws.column_dimensions[col_letter].width = 35
            else:
                ws.column_dimensions[col_letter].width = 18
        ws.row_dimensions[1].height = 30

        # Sheets 2–4 — tag summaries (one per tag type)
        tag_colors = {
            "place":    ("Place Tags",    "2E75B6"),
            "activity": ("Activity Tags", "375623"),
            "theme":    ("Theme Tags",    "7030A0"),
        }
        for tag_type, (sheet_name, color) in tag_colors.items():
            subset = review_df[review_df["tag_type"] == tag_type].copy()
            subset = subset[["tag", "frequency", "example_tour", "example_day", "example_snippet"]]
            subset.columns = ["Keyword", "Frequency", "Example Tour", "Example Day", "Example Snippet"]
            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_header(ws, header_color=color)
            _autofit(ws)
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["E"].width = 80
            for cell in ws["E"]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    print(f"Excel workbook saved to {output_path}")
    print(f"  Sheets: Dataset | Place Tags | Activity Tags | Theme Tags")


# ---------------------------------------------------------------------------
# Gemini Flash NER (optional — activated by setting GEMINI_API_KEY env var)
# ---------------------------------------------------------------------------

_GEMINI_SYSTEM = (
    "You extract place names from Australian tour itinerary text. "
    "Include: cities, towns, suburbs, national parks, beaches, gorges, mountains, "
    "capes, bays, islands, rivers, lakes, waterfalls, LOOKOUTS and VIEWPOINTS, "
    "landmarks, tourist attractions, museums, galleries, theatres, precincts, "
    "cultural centres, heritage sites. "
    "IMPORTANT: expand shared-suffix conjunctions — e.g. 'Oxer and Junction Pool "
    "lookouts' -> ['Oxer Lookout','Junction Pool Lookout']; 'Joffre and Knox gorges' "
    "-> ['Joffre Gorge','Knox Gorge']. "
    "Exclude: hotel/motel/inn/resort/lodge names (Novotel, Voco, Holiday Inn), restaurant names, "
    "bar names, marketing phrases (Follow The Flavours, Fall In Love With Sydney), "
    "non-Australian places (Cairo, Antarctica), activity verbs (Kayak, Snorkel), "
    "tour company names, generic words (Area, Region, Reserve, Metro, Level, Building)."
)

_GEMINI_ROW_PROMPT = """\
For each numbered row, extract Australian place names from the Location and Activity text.
Return one JSON array per row on its own line, same order as input. No other text.

{rows}"""


def _build_gemini_client():
    """Initialise google-genai client. Raises SystemExit if GEMINI_API_KEY is not set."""
    try:
        from google import genai
    except ImportError:
        raise SystemExit("google-genai not installed. Run: pip install google-genai")
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not key:
        # Fall back to a key file in the repo root: "Gemini API Key = AQ.xxxx"
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        kf = os.path.join(base, "gemini_apikey.txt")
        if os.path.exists(kf):
            raw = open(kf, encoding="utf-8").read()
            after = raw.split("=", 1)[1] if "=" in raw else raw
            toks = [t for t in after.split() if t]
            key = toks[0].strip() if toks else ""
    if not key:
        raise SystemExit(
            "GEMINI_API_KEY not set and no gemini_apikey.txt found.\n"
            "  1. Get a free key at: https://aistudio.google.com/app/apikey\n"
            "  2. Set it: $env:GEMINI_API_KEY = 'your-key-here'  (or put it in gemini_apikey.txt)\n"
            "  3. Re-run the script."
        )
    return genai.Client(api_key=key)


def _gemini_call(pairs, client):
    """Send one batch of (location, activity) pairs to Gemini. Returns list of place lists."""
    from google.genai import types
    rows_text = "\n\n".join(
        f"Row {i+1}:\n  Location: {loc}\n  Activity: {act[:2000]}"
        for i, (loc, act) in enumerate(pairs)
    )
    prompt = _GEMINI_ROW_PROMPT.format(rows=rows_text)
    try:
        response = client.models.generate_content(
            model=os.environ.get("GEMINI_MODEL", "gemini-flash-latest"),
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction=_GEMINI_SYSTEM,
                temperature=0.1,
            ),
        )
        raw_lines = [l.strip() for l in response.text.strip().splitlines() if l.strip()]
        results = []
        for line in raw_lines:
            # Strip markdown fences if Gemini wraps output in ```json
            line = line.strip("`").strip()
            if line.lower().startswith("json"):
                line = line[4:].strip()
            try:
                arr = json.loads(line)
                results.append([str(p).strip() for p in arr if str(p).strip()])
            except json.JSONDecodeError:
                results.append([])
        # Pad to match batch length in case Gemini returned fewer lines
        while len(results) < len(pairs):
            results.append([])
        return results[:len(pairs)]
    except Exception as e:
        print(f"  [gemini] API error: {e}")
        return [[] for _ in pairs]


def _run_gemini_ner(locations, activities, cache_path, batch_size=15, rpm=15):
    """
    Extract places for all rows using Gemini Flash.
    Results are cached in cache_path (JSON) so reruns skip already-processed rows.
    Returns a list of place-name lists, one per row.
    """
    client = _build_gemini_client()

    cache = {}
    if os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            cache = json.load(f)

    def _key(loc, act):
        # v2 = Gemini-first config (full text + lookout/conjunction prompt). The
        # version tag busts the old 300-char cache so every row re-extracts.
        return hashlib.md5(f"v2|||{loc}|||{act[:2000]}".encode()).hexdigest()

    all_results = [None] * len(locations)
    uncached_idx = []

    for i, (loc, act) in enumerate(zip(locations, activities)):
        k = _key(loc, act)
        if k in cache:
            all_results[i] = cache[k]
        else:
            uncached_idx.append(i)

    cached_count = len(locations) - len(uncached_idx)
    print(f"  [gemini] {cached_count:,} rows from cache | {len(uncached_idx):,} rows to process via API")

    if not uncached_idx:
        return all_results

    delay = 60.0 / rpm
    total_batches = (len(uncached_idx) + batch_size - 1) // batch_size

    for batch_num, start in enumerate(range(0, len(uncached_idx), batch_size), 1):
        idx_slice = uncached_idx[start : start + batch_size]
        pairs = [(locations[i], activities[i]) for i in idx_slice]

        batch_results = _gemini_call(pairs, client)

        for i, places in zip(idx_slice, batch_results):
            all_results[i] = places
            cache[_key(locations[i], activities[i])] = places

        # Save cache after every batch so progress is never lost
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)

        if batch_num % 10 == 0 or batch_num == total_batches:
            print(f"  [gemini] {batch_num}/{total_batches} batches complete")

        if batch_num < total_batches:
            time.sleep(delay)

    return all_results


# ---------------------------------------------------------------------------
# NuNER Zero NER (local, no API key — set NER_BACKEND=nunerzero to activate)
# ---------------------------------------------------------------------------

_NUNER_LABELS = [
    "Australian city or town",
    "suburb or neighbourhood",
    "national park or nature reserve",
    "beach, bay, or coastal feature",
    "mountain, gorge, or canyon",
    "island, lake, or river",
    "tourist attraction or landmark",
    "museum, gallery, or theatre",
    "cultural or heritage site",
    "road, highway, or touring route",
]


def _run_nunerzero_ner(locations, activities):
    """
    Extract places using NuNER Zero (DeBERTa-v3-large backbone, zero-shot NER).
    Returns a list of place-name lists, one per row.
    """
    try:
        from gliner import GLiNER
    except ImportError:
        raise SystemExit("gliner not installed. Run: pip install gliner")

    print("  Loading NuNER Zero model (numind/NuNER_Zero) — first run downloads ~700MB ...")
    model = GLiNER.from_pretrained("numind/NuNER_Zero")
    print(f"  Running NuNER Zero NER on {len(locations):,} rows ...")

    texts = [
        f"{loc} {act[:300]}".strip()
        for loc, act in zip(locations, activities)
    ]

    BATCH_SIZE = 32
    all_results = []

    for start in range(0, len(texts), BATCH_SIZE):
        batch = texts[start : start + BATCH_SIZE]
        batch_entities = model.batch_predict_entities(
            batch, _NUNER_LABELS, threshold=0.45
        )
        for entities in batch_entities:
            places = list(dict.fromkeys(
                e["text"].strip() for e in entities if e["text"].strip()
            ))
            all_results.append(places)

        done = min(start + BATCH_SIZE, len(texts))
        if done % 500 == 0 or done == len(texts):
            print(f"  [nunerzero] {done:,}/{len(texts):,} rows done")

    return all_results


# ---------------------------------------------------------------------------
# Ollama (local, no API key, no quotas — set NER_BACKEND=ollama to activate)
# ---------------------------------------------------------------------------
# Runs against a local Ollama server (default qwen2.5:7b at localhost:11434).
# Per-row calls (reliable: a parse miss costs one row, never a bulk wipe),
# robust JSON-array parsing, retry, and an auto-abort if the server dies so we
# never silently write a mostly-empty dataset. Caches every row for resume.

_OLLAMA_ARR = re.compile(r"\[[^\[\]]*\]", re.S)


def _parse_place_array(text):
    """Extract the first JSON array of place strings from a model reply.
    Robust to ``` fences and pretty-printed/multiline arrays. Returns a list."""
    text = re.sub(r"```[a-z]*", "", text or "", flags=re.IGNORECASE)
    m = _OLLAMA_ARR.search(text)
    if not m:
        return []
    try:
        return [str(p).strip() for p in json.loads(m.group(0)) if str(p).strip()]
    except json.JSONDecodeError:
        return []


def _ollama_call(url, model, loc, act):
    """One row -> list of places. Raises on transport/HTTP failure so the
    caller can distinguish a genuine empty result from a dead server."""
    import requests
    user = (f"Location: {loc}\nActivity: {act[:1500]}\n\n"
            "Return ONLY a JSON array of the Australian place-name strings for this day.")
    body = {
        "model": model, "temperature": 0.1, "stream": False, "keep_alive": "60m",
        "messages": [
            {"role": "system", "content": _GEMINI_SYSTEM},
            {"role": "user", "content": user},
        ],
    }
    last = None
    for attempt in range(3):
        try:
            r = requests.post(url, json=body, timeout=300)
            r.raise_for_status()
            return _parse_place_array(r.json()["choices"][0]["message"]["content"])
        except Exception as e:  # noqa: BLE001 — transient; retry then raise
            last = e
            time.sleep(2 * (attempt + 1))
    raise RuntimeError(f"Ollama call failed after retries: {last!r}")


def _run_ollama_ner(locations, activities, cache_path):
    model = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")
    host  = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
    url   = host + "/v1/chat/completions"

    cache = {}
    if os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path, encoding="utf-8"))
        except Exception:
            cache = {}

    def key(loc, act):
        return hashlib.md5(f"ollama|{model}|{loc}|||{act[:1500]}".encode()).hexdigest()

    results = [None] * len(locations)
    todo = []
    for i, (loc, act) in enumerate(zip(locations, activities)):
        k = key(loc, act)
        if k in cache:
            results[i] = cache[k]
        else:
            todo.append(i)

    print(f"  [ollama] model={model} | {len(locations) - len(todo):,} cached | "
          f"{len(todo):,} rows to process")

    def _save():
        json.dump(cache, open(cache_path, "w", encoding="utf-8"), ensure_ascii=False)

    consec_fail = 0
    done = 0
    for i in todo:
        loc, act = locations[i], activities[i]
        try:
            places = _ollama_call(url, model, loc, act)
            consec_fail = 0
        except Exception as e:  # noqa: BLE001
            consec_fail += 1
            print(f"  [ollama] row {i} failed ({consec_fail} in a row): {e}")
            if consec_fail >= 6:
                _save()
                raise SystemExit(
                    "Ollama failed 6 rows in a row — aborting to avoid a mostly-empty "
                    "dataset. Progress is cached; fix the server and re-run to resume."
                )
            results[i] = []  # not cached -> a later resume will retry this row
            continue
        results[i] = places
        cache[key(loc, act)] = places
        done += 1
        if done % 25 == 0:
            _save()
            print(f"  [ollama] {done:,}/{len(todo):,} processed ...")

    _save()
    return results


# ---------------------------------------------------------------------------
# n8n Cloud NER (company proxy — set NER_BACKEND=cloud to activate)
# ---------------------------------------------------------------------------
# Uses the company n8n webhook that proxies to Claude. No external API key needed.
# Also used as a targeted fallback in NER_BACKEND=lookup+cloud mode (only fires
# for rows where the allowlist scan returns 0 place matches).

_N8N_URL    = "https://gtxn8n.yourbestwayhome.com.au/webhook/openai-response-proxy"
_N8N_KEY    = os.environ.get("N8N_KEY", "Gtx1234*")
_N8N_SYSTEM = (
    "You extract Australian place names from tour itinerary text. "
    "Return ONLY a comma-separated list of place names, nothing else. "
    "Include: cities, towns, national parks, beaches, gorges, mountains, capes, bays, "
    "islands, rivers, lakes, waterfalls, lookouts, landmarks, tourist attractions, "
    "museums, galleries, cultural centres, heritage sites. "
    "Exclude: hotel names, restaurant names, tour company names, generic words, "
    "non-Australian places, activity verbs."
)


def _n8n_call(loc, act):
    """Call the n8n Claude proxy for one row. Returns list of place strings."""
    import requests as _req
    user = f"Location: {loc}\nActivity: {act[:2000]}"
    for attempt in range(3):
        try:
            r = _req.post(
                _N8N_URL,
                headers={"N8N-KEY": _N8N_KEY, "Content-Type": "application/json"},
                json={"system": _N8N_SYSTEM, "user": user},
                timeout=45,
            )
            r.raise_for_status()
            text = r.json()["output"][0]["content"][0]["text"].strip()
            return [p.strip() for p in text.split(",") if p.strip()]
        except Exception:
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
    return []


def _run_n8n_ner(locations, activities, cache_path):
    """Run n8n cloud NER on all rows. Caches every 25 rows for safe resume."""
    cache = {}
    if os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path, encoding="utf-8"))
            print(f"  Resuming from cache: {len(cache)} rows already done.")
        except Exception:
            pass

    results = [None] * len(locations)
    todo = [
        (i, loc, act)
        for i, (loc, act) in enumerate(zip(locations, activities))
        if str(i) not in cache
    ]

    def _save():
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache, f)

    for done, (i, loc, act) in enumerate(todo, 1):
        places = _n8n_call(loc, act)
        cache[str(i)] = places
        results[i] = places
        if done % 25 == 0:
            _save()
            print(f"  [n8n cloud] {done:,}/{len(todo):,} processed ...")

    for i in range(len(locations)):
        if results[i] is None:
            results[i] = cache.get(str(i), [])

    _save()
    return results


# ---------------------------------------------------------------------------
# Cerebras Cloud NER (free — set NER_BACKEND=cerebras to activate)
# ---------------------------------------------------------------------------
# Primary : Cerebras llama-3.3-70b  (1 M tokens/day free, 30 RPM)
# Fallback: Groq   llama-3.1-8b-instant (500 K tokens/day free, ~14 RPM)
# Auto-switches to fallback when the primary hits its daily quota.
# Caches every 25 rows to cache_path; safe to stop and resume any time.

_CEREBRAS_URL      = "https://api.cerebras.ai/v1/chat/completions"
_GEMINI_COMPAT_URL = "https://generativelanguage.googleapis.com/v1beta/openai/chat/completions"
_GROQ_URL          = "https://api.groq.com/openai/v1/chat/completions"


def _load_cloud_key(pattern, env_var):
    """Return an API key from env var or by scanning the key files in the repo root."""
    key = os.environ.get(env_var, "").strip()
    if key:
        return key
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for fn in ("gemini_apikey.txt", "cerebras_apikey.txt", "groq_apikey.txt"):
        p = os.path.join(base, fn)
        if not os.path.exists(p):
            continue
        m = re.search(pattern, open(p, encoding="utf-8").read())
        if m:
            return m.group(0)
    return ""


def _cloud_ner_call(url, model, key, loc, act):
    """
    One row → (places, status) via any OpenAI-compatible endpoint.
    status: "ok" | "daily_limit" | "rate_limit" | "error"
    """
    import requests
    user = (f"Location: {loc}\nActivity: {act[:1500]}\n\n"
            "Return ONLY a JSON array of the Australian place-name strings for this day.")
    body = {
        "model": model, "temperature": 0.1,
        "messages": [
            {"role": "system", "content": _GEMINI_SYSTEM},
            {"role": "user",   "content": user},
        ],
    }
    for attempt in range(4):
        try:
            r = requests.post(
                url,
                headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
                json=body, timeout=60,
            )
            if r.status_code == 429:
                body_lower = r.text.lower()
                if any(w in body_lower for w in ("daily", "per day", "quota", "exceeded your", "monthly")):
                    return [], "daily_limit"
                wait = float(r.headers.get("retry-after", min(2 ** attempt, 30)))
                print(f"    rate-limited, waiting {wait:.0f}s ...")
                time.sleep(wait + 1)
                return [], "rate_limit"
            r.raise_for_status()
            return _parse_place_array(r.json()["choices"][0]["message"]["content"]), "ok"
        except Exception as e:  # noqa: BLE001
            if attempt < 3:
                time.sleep(2 * (attempt + 1))
    return [], "error"


def _run_cerebras_ner(locations, activities, cache_path):
    # Scan all key files for every available key — multiple accounts = stacked daily quota.
    # Chain order: all Cerebras keys → all Gemini keys → all Groq keys.
    # When one account hits its daily limit the next picks up automatically.
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _txt = "".join(
        open(os.path.join(base, fn), encoding="utf-8").read()
        for fn in ("gemini_apikey.txt", "cerebras_apikey.txt", "groq_apikey.txt")
        if os.path.exists(os.path.join(base, fn))
    )
    def _all_keys(pattern, env_var):
        seen, out = set(), []
        env = os.environ.get(env_var, "").strip()
        if env and env not in seen:
            seen.add(env); out.append(env)
        for k in re.findall(pattern, _txt):
            if k not in seen:
                seen.add(k); out.append(k)
        return out

    cerebras_keys = _all_keys(r"csk-[A-Za-z0-9_\-]+",       "CEREBRAS_API_KEY")
    gemini_keys   = _all_keys(r"(?:AIzaSy[A-Za-z0-9_\-]{30,}|AQ\.[A-Za-z0-9_\-]{20,})", "GEMINI_API_KEY")
    groq_keys     = _all_keys(r"gsk_[A-Za-z0-9_\-]+",       "GROQ_API_KEY")

    providers = []
    for i, key in enumerate(cerebras_keys, 1):
        providers.append({
            "name": f"cerebras-{i}", "url": _CEREBRAS_URL, "key": key,
            "model": os.environ.get("CEREBRAS_MODEL", "gpt-oss-120b"),
            "delay": 2.1,
        })
    for i, key in enumerate(gemini_keys, 1):
        providers.append({
            "name": f"gemini-{i}", "url": _GEMINI_COMPAT_URL, "key": key,
            "model": os.environ.get("GEMINI_COMPAT_MODEL", "gemini-2.0-flash"),
            "delay": 4.1,
        })
    for i, key in enumerate(groq_keys, 1):
        providers.append({
            "name": f"groq-{i}", "url": _GROQ_URL, "key": key,
            "model": os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile"),
            "delay": 4.5,
        })

    if not providers:
        raise SystemExit(
            "No keys found. Add csk-... (Cerebras) or gsk_... (Groq) to gemini_apikey.txt\n"
            "or set CEREBRAS_API_KEY / GROQ_API_KEY."
        )

    cache = {}
    if os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path, encoding="utf-8"))
        except Exception:
            cache = {}

    def cache_key(loc, act):
        return hashlib.md5(f"cerebras_v1|{loc}|||{act[:1500]}".encode()).hexdigest()

    results = [None] * len(locations)
    todo = []
    for i, (loc, act) in enumerate(zip(locations, activities)):
        k = cache_key(loc, act)
        if k in cache:
            results[i] = cache[k]
        else:
            todo.append(i)

    cached = len(locations) - len(todo)
    chain = " → ".join(f"{p['name']}({p['model']})" for p in providers)
    print(f"  [cloud-ner] chain: {chain}")
    print(f"  [cloud-ner] {cached:,} cached | {len(todo):,} to process")

    if not todo:
        return results

    def _save():
        json.dump(cache, open(cache_path, "w", encoding="utf-8"), ensure_ascii=False)

    prov_idx    = 0
    consec_fail = 0
    done        = 0

    for i in todo:
        if prov_idx >= len(providers):
            print("  [cerebras] All providers hit their daily limits. Re-run tomorrow to continue.")
            break

        p = providers[prov_idx]
        loc, act = locations[i], activities[i]

        while True:
            places, status = _cloud_ner_call(p["url"], p["model"], p["key"], loc, act)

            if status == "daily_limit":
                print(f"  [{p['name']}] Daily quota reached ({done:,} rows processed today). ", end="")
                prov_idx += 1
                if prov_idx >= len(providers):
                    print("No more providers. Saving progress.")
                    _save()
                    return [r if r is not None else [] for r in results]
                p = providers[prov_idx]
                print(f"Switching to [{p['name']}].")
                continue

            if status == "rate_limit":
                continue  # already waited inside _cloud_ner_call

            if status == "error":
                consec_fail += 1
                print(f"  [{p['name']}] row {i} error ({consec_fail} consecutive)")
                if consec_fail >= 6:
                    _save()
                    raise SystemExit("6 consecutive API failures — saved progress; re-run to resume.")
                results[i] = []
                break

            # ok
            consec_fail     = 0
            results[i]      = places
            cache[cache_key(loc, act)] = places
            done           += 1
            if done % 25 == 0:
                _save()
                print(f"  [{p['name']}] {done:,}/{len(todo):,} processed ...")
            time.sleep(p["delay"])
            break

    _save()
    return [r if r is not None else [] for r in results]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_keyword_dataset(input_csv: str, output_csv: str, review_csv: str,
                           blacklist_path: str, allowlist_path: str,
                           synonyms_path: str = "", output_xlsx: str = ""):
    print(f"Loading {input_csv}...")
    df = pd.read_csv(input_csv)
    print(f"  {len(df):,} rows")

    # Load manual config
    blacklist = _load_lines(blacklist_path)
    allowlist = _load_allowlist(allowlist_path)
    synonyms  = _load_synonyms(synonyms_path)
    print(f"  Blacklist: {len(blacklist)} entries  |  Allowlist: {len(allowlist)} entries  |  Synonyms: {len(synonyms)} entries")

    # Build allowlist regex patterns (case-insensitive, word-boundary enforced).
    # Lookarounds on \w prevent substring false matches, e.g. "Ross" inside
    # "aCROSS"/"Cross", or "Cook" inside "Cooktown". More robust than \b for
    # phrases that may begin/end with punctuation (apostrophes, parentheses).
    allowlist_patterns = [(phrase, re.compile(r"(?<!\w)" + re.escape(phrase) + r"(?!\w)", re.IGNORECASE))
                          for phrase in allowlist]

    activities = df["activity"].fillna("").astype(str).tolist()
    locations  = df["location"].fillna("").astype(str).tolist()

    ner_backend = os.environ.get("NER_BACKEND", "spacy").lower().strip()
    # Options: "spacy" (default), "nunerzero", "gemini", "ollama", "cerebras"

    external_ner_results = None

    if ner_backend == "nunerzero":
        print("NuNER Zero selected (NER_BACKEND=nunerzero) ...")
        external_ner_results = _run_nunerzero_ner(locations, activities)
        docs = loc_docs = nlp = None

    elif ner_backend == "gemini":
        print("Gemini Flash NER selected (NER_BACKEND=gemini) ...")
        cache_path = os.path.join(os.path.dirname(output_csv), "gemini_ner_cache.json")
        external_ner_results = _run_gemini_ner(locations, activities, cache_path)
        print("  Gemini NER complete.")
        docs = loc_docs = nlp = None

    elif ner_backend == "ollama":
        print("Ollama (local) NER selected (NER_BACKEND=ollama) ...")
        cache_path = os.path.join(os.path.dirname(output_csv), "ollama_ner_cache.json")
        external_ner_results = _run_ollama_ner(locations, activities, cache_path)
        print("  Ollama NER complete.")
        docs = loc_docs = nlp = None

    elif ner_backend == "cerebras":
        print("Cerebras Cloud NER selected (NER_BACKEND=cerebras) ...")
        cache_path = os.path.join(os.path.dirname(output_csv), "cerebras_ner_cache.json")
        external_ner_results = _run_cerebras_ner(locations, activities, cache_path)
        print("  Cerebras NER complete.")
        docs = loc_docs = nlp = None

    elif ner_backend == "cloud":
        print("n8n Cloud NER selected (NER_BACKEND=cloud) ...")
        cache_path = os.path.join(os.path.dirname(output_csv), "n8n_ner_cache.json")
        external_ner_results = _run_n8n_ner(locations, activities, cache_path)
        print("  n8n cloud NER complete.")
        docs = loc_docs = nlp = None

    elif ner_backend in ("lookup", "lookup+cloud"):
        if ner_backend == "lookup+cloud":
            print("Hybrid Lookup+Cloud mode (NER_BACKEND=lookup+cloud).")
            print("  Lookup handles all rows; n8n cloud fallback for zero-match rows only.")
        else:
            print("Keyword Lookup mode selected (NER_BACKEND=lookup).")
            print("  Uses allowlist + location tokeniser only. No LLM or spaCy needed.")
        print(f"  Allowlist: {len(allowlist)} phrases | Blacklist: {len(blacklist)} phrases | Synonyms: {len(synonyms)} pairs")
        external_ner_results = None
        docs = loc_docs = nlp = None

    else:
        print("Loading spaCy model (set NER_BACKEND=nunerzero for NuNER Zero) ...")
        nlp = spacy.load("en_core_web_lg")
        print("Running spaCy NER on activity text (batch mode)...")
        docs = list(nlp.pipe(activities, batch_size=128, disable=["tagger", "parser", "lemmatizer"]))
        print("  Activity NER done.")
        print("Running spaCy NER on location text (batch mode)...")
        loc_docs = list(nlp.pipe(locations, batch_size=128, disable=["tagger", "parser", "lemmatizer"]))
        print("  Location NER done.")

    place_tags_col      = []
    activity_tags_col   = []
    theme_tags_col      = []
    all_keywords_col    = []
    keyword_count_col   = []
    has_content_col     = []
    location_tokens_col = []

    use_external       = external_ner_results is not None
    is_lookup          = ner_backend in ("lookup", "lookup+cloud")
    use_cloud_fallback = ner_backend == "lookup+cloud"
    cloud_fallback_count = 0

    for i, (loc, act) in enumerate(zip(locations, activities)):
        act_clean = clean_text(act)
        loc_clean = clean_text(loc)
        has_content = bool(act_clean)
        combined = (loc_clean + " " + act_clean).strip()

        # Location tokeniser — clean individual place tokens from the location field
        loc_tokens = tokenize_location(loc)
        loc_tokens_clean = _apply_blacklist(
            _apply_synonyms(_merge_tags(loc_tokens), synonyms), blacklist
        )

        # Layer 4 — allowlist scan (runs for all backends)
        allow_hits = [phrase for phrase, pat in allowlist_patterns if pat.search(combined)]

        if is_lookup:
            # Lookup mode: allowlist + location tokens only — no LLM, no spaCy
            spacy_places     = []
            loc_spacy_places = []
            regex_places     = []
            themes           = extract_themes(loc_clean, act_clean)
            place_inputs     = [loc_tokens_clean, allow_hits]

        elif use_external:
            # External LLM (Gemini/Ollama/Cerebras): trust model output as sole place source
            spacy_places     = external_ner_results[i]
            loc_spacy_places = []
            regex_places     = []
            themes           = extract_themes(loc_clean, act_clean)
            place_inputs     = [spacy_places, allow_hits]

        else:
            # spaCy (default): merge all layers
            doc     = docs[i]
            loc_doc = loc_docs[i]
            spacy_places     = _extract_spacy_entities(doc)
            loc_spacy_places = _extract_spacy_entities(loc_doc)
            loc_for_extractor = ", ".join(loc_tokens_clean) if loc_tokens_clean else loc_clean
            regex_places     = [p for p in extract_products(loc_for_extractor, act_clean) if p]
            themes           = extract_themes(loc_clean, act_clean)
            place_inputs     = [loc_tokens_clean, regex_places, spacy_places, loc_spacy_places, allow_hits]

        # Layer 3 — named activity patterns (all backends)
        act_activities = _extract_activity_tags(act_clean)

        place_tags = _apply_blacklist(
            _apply_synonyms(_merge_tags(*place_inputs), synonyms),
            blacklist,
        )

        # Hybrid fallback: if lookup found nothing, call n8n cloud for this row
        if use_cloud_fallback and not place_tags and has_content:
            cloud_places = _n8n_call(loc, act)
            cloud_places = _apply_blacklist(_apply_synonyms(_merge_tags(cloud_places), synonyms), blacklist)
            if cloud_places:
                place_tags = cloud_places
                cloud_fallback_count += 1

        activity_tags = _apply_blacklist(_apply_synonyms(_merge_tags(act_activities), synonyms), blacklist)
        themes        = [t for t in themes if t.lower() not in blacklist]
        all_kw        = _merge_tags(place_tags, activity_tags, themes)

        location_tokens_col.append(" | ".join(loc_tokens_clean))
        place_tags_col.append("; ".join(place_tags))
        activity_tags_col.append("; ".join(activity_tags))
        theme_tags_col.append("; ".join(themes))
        all_keywords_col.append("; ".join(all_kw))
        keyword_count_col.append(len(all_kw))
        has_content_col.append(has_content)

        if (i + 1) % 1000 == 0:
            print(f"  Processed {i + 1:,}/{len(df):,} rows...")

    df["location_tokens"] = location_tokens_col
    df["place_tags"]      = place_tags_col
    df["activity_tags"]   = activity_tags_col
    df["theme_tags"]      = theme_tags_col
    df["all_keywords"]    = all_keywords_col
    df["keyword_count"]   = keyword_count_col
    df["has_content"]     = has_content_col

    for col in ["location_tokens", "place_tags", "activity_tags", "theme_tags", "all_keywords"]:
        df[col] = df[col].replace("", pd.NA)

    df.to_csv(output_csv, index=False)
    print(f"\nSaved tagged dataset to {output_csv}")

    # Build review file
    review_df = _build_review_csv(df, review_csv)

    # Build Excel workbook (resilient if the file is open in Excel)
    if output_xlsx and review_df is not None:
        write_xlsx_safe(lambda p: _build_excel(df, review_df, p), output_xlsx)

    # Summary stats
    total = len(df)
    with_content  = df["has_content"].sum()
    with_keywords = (df["keyword_count"] > 0).sum()
    avg_kw = df[df["keyword_count"] > 0]["keyword_count"].mean()

    print(f"\n--- Summary ---")
    print(f"Total rows:         {total:,}")
    print(f"Rows with content:  {with_content:,} ({with_content/total*100:.1f}%)")
    print(f"Rows with keywords: {with_keywords:,} ({with_keywords/total*100:.1f}%)")
    print(f"Avg keywords/row:   {avg_kw:.1f} (where > 0)")
    if use_cloud_fallback:
        print(f"Cloud fallback used:{cloud_fallback_count:,} rows (zero-match rows rescued by n8n)")

    print(f"\n--- Top 20 place tags ---")
    place_counts = Counter()
    for tags in df["place_tags"].dropna():
        for t in str(tags).split("; "):
            if t:
                place_counts[t] += 1
    for tag, count in place_counts.most_common(20):
        print(f"  {count:4d}  {tag}")

    print(f"\n--- Top 15 activity tags ---")
    act_counts = Counter()
    for tags in df["activity_tags"].dropna():
        for t in str(tags).split("; "):
            if t:
                act_counts[t] += 1
    for tag, count in act_counts.most_common(15):
        print(f"  {count:4d}  {tag}")


if __name__ == "__main__":
    base          = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_csv     = os.path.join(base, "data", "raw",      "combined_itineraries_2026-06.csv")
    output_csv    = os.path.join(base, "data", "keywords", "keyword_dataset_2026-06.csv")
    output_xlsx   = os.path.join(base, "data", "keywords", "keyword_dataset_2026-06.xlsx")
    review_csv    = os.path.join(base, "data", "keywords", "keyword_review.csv")
    blacklist_path  = os.path.join(base, "data", "config",   "keyword_blacklist.txt")
    allowlist_path  = os.path.join(base, "data", "config",   "keyword_allowlist.txt")
    synonyms_path   = os.path.join(base, "data", "config",   "keyword_synonyms.txt")
    build_keyword_dataset(input_csv, output_csv, review_csv, blacklist_path, allowlist_path,
                          synonyms_path, output_xlsx)
