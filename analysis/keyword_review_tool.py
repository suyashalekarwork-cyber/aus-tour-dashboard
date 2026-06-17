"""
Keyword Review Tool — two commands:

  python analysis/keyword_review_tool.py export
      Reads data/keywords/keyword_review.csv and produces
      data/keywords/keyword_review.xlsx  — ready to fill in.

  python analysis/keyword_review_tool.py apply
      Reads your decisions from keyword_review.xlsx and
      appends them to the three config files:
        data/config/keyword_blacklist.txt
        data/config/keyword_synonyms.txt
        data/config/keyword_allowlist.txt

Excel columns (Review sheet):
  Tag | Type | Freq | Action ▾ | Synonym → | Example Tour | URL | Day | Snippet

  Action dropdown: blacklist · synonym · allowlist  (blank = keep as-is)
  Synonym →: fill in only when Action = synonym  (e.g. "Uluru")
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

sys.path.insert(0, str(Path(__file__).resolve().parent))
from xlsx_io import write_xlsx_safe

ROOT         = Path(__file__).resolve().parent.parent
REVIEW_CSV   = ROOT / "data" / "keywords" / "keyword_review.csv"
REVIEW_XLSX  = ROOT / "data" / "keywords" / "keyword_review.xlsx"
BLACKLIST    = ROOT / "data" / "config" / "keyword_blacklist.txt"
SYNONYMS     = ROOT / "data" / "config" / "keyword_synonyms.txt"
ALLOWLIST    = ROOT / "data" / "config" / "keyword_allowlist.txt"
CHANGELOG    = ROOT / "data" / "config" / "keyword_changelog.txt"

# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ROW_ALT    = "F2F7FF"
BL_FILL    = "FFD7D7"   # red tint   — blacklist
SYN_FILL   = "FFF3CD"   # amber tint — synonym
AL_FILL    = "D6F0D6"   # green tint — allowlist

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text):
    cell.value     = text
    cell.font      = Font(bold=True, color=HEADER_FG, size=10)
    cell.fill      = _fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()


# ── EXPORT ────────────────────────────────────────────────────────────────────

def export():
    if not REVIEW_CSV.exists():
        raise SystemExit(f"Source not found: {REVIEW_CSV}\nRun build_keyword_dataset.py first.")

    rows = []
    with open(REVIEW_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    print(f"[export] {len(rows):,} tags  ->  {REVIEW_XLSX.name}")

    wb = openpyxl.Workbook()

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws_help = wb.active
    ws_help.title = "How to Use"
    instructions = [
        ("How to use this workbook", True),
        ("", False),
        ("1. Go to the 'Review' sheet.", False),
        ("2. For every tag you want to clean, click the 'Action' cell and pick from the dropdown:", False),
        ("   • blacklist  — tag will be removed from all future extractions", False),
        ("   • synonym    — tag will be replaced by the canonical name you enter in 'Synonym →'", False),
        ("   • allowlist  — tag will always be kept (even if spaCy misses it)", False),
        ("   • (blank)    — leave tag as-is, no change", False),
        ("", False),
        ("3. For synonym rows, fill in the 'Synonym →' column with the target name.", False),
        ("   Example:  Tag = 'Ayers Rock'   Synonym → = 'Uluru'", False),
        ("", False),
        ("4. Save the file (Ctrl+S).", False),
        ("", False),
        ("5. Run the apply command to push your decisions to the config files:", False),
        ("   python analysis/keyword_review_tool.py apply", False),
        ("", False),
        ("Colour guide:", True),
        ("   Red    = blacklist", False),
        ("   Amber  = synonym", False),
        ("   Green  = allowlist", False),
        ("", False),
        ("Config files updated by 'apply':", True),
        (f"   {BLACKLIST.relative_to(ROOT)}", False),
        (f"   {SYNONYMS.relative_to(ROOT)}", False),
        (f"   {ALLOWLIST.relative_to(ROOT)}", False),
    ]
    ws_help.column_dimensions["A"].width = 70
    for i, (text, bold) in enumerate(instructions, 1):
        c = ws_help.cell(i, 1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(wrap_text=True)

    # ── Review sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Review", 0)

    col_widths = {
        "A": 38,   # Tag
        "B": 10,   # Type
        "C": 8,    # Freq
        "D": 12,   # Action
        "E": 28,   # Synonym →
        "F": 32,   # Example Tour
        "G": 42,   # URL
        "H": 6,    # Day
        "I": 60,   # Snippet
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    headers = ["Tag", "Type", "Freq", "Action", "Synonym →",
               "Example Tour", "URL", "Day", "Snippet"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(1, col), h)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"

    # Dropdown validation on the Action column (D)
    dv = DataValidation(
        type="list",
        formula1='"blacklist,synonym,allowlist"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid action",
        error="Choose: blacklist, synonym, allowlist, or leave blank.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"D2:D{len(rows) + 1}"

    # Conditional formatting (whole row colour based on Action cell in col D)
    max_row = len(rows) + 1
    range_ref = f"A2:I{max_row}"

    for action, hex_color in [("blacklist", BL_FILL), ("synonym", SYN_FILL), ("allowlist", AL_FILL)]:
        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(
                formula=[f'$D2="{action}"'],
                fill=_fill(hex_color),
            ),
        )

    # Data rows
    for i, row in enumerate(rows):
        r = i + 2
        bg = ROW_ALT if i % 2 == 1 else "FFFFFF"
        values = [
            row.get("tag", ""),
            row.get("tag_type", ""),
            int(row.get("frequency", 0)),
            "",                                  # Action — user fills
            "",                                  # Synonym → — user fills
            row.get("example_tour", ""),
            row.get("example_url", ""),
            row.get("example_day", ""),
            row.get("example_snippet", ""),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(r, col, value=val)
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="top", wrap_text=(col == 9), indent=(1 if col != 3 else 0))
            c.font      = Font(size=10)
            if col == 3:
                c.alignment = Alignment(horizontal="center", vertical="top")
            if col == 7 and str(val).startswith("http"):
                c.hyperlink = val
                c.font      = Font(color="0563C1", underline="single", size=10)
        ws.row_dimensions[r].height = 15

    wb.active = ws
    saved = write_xlsx_safe(lambda p: wb.save(p), REVIEW_XLSX)
    print(f"[export] saved  ->  {saved}")
    print(f"         Open the file, fill in the 'Action' column, then run:")
    print(f"         python analysis/keyword_review_tool.py apply")


# ── APPLY ─────────────────────────────────────────────────────────────────────

def _write_changelog(added_bl, added_syn, added_al):
    """Append a timestamped entry to keyword_changelog.txt for every new decision applied."""
    if not (added_bl or added_syn or added_al):
        return
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(CHANGELOG, "a", encoding="utf-8") as f:
        f.write(f"\n=== {timestamp} ===\n")
        for tag in added_bl:
            f.write(f"  [blacklist]  {tag}\n")
        for entry in added_syn:
            f.write(f"  [synonym]    {entry}\n")
        for tag in added_al:
            f.write(f"  [allowlist]  {tag}\n")
        total = len(added_bl) + len(added_syn) + len(added_al)
        f.write(f"  --- {total} new entries ---\n")
    print(f"[apply] Changelog updated -> {CHANGELOG.name}")


def _load_existing(path: Path) -> set:
    """Return lowercase set of non-comment, non-blank lines already in a config file."""
    if not path.exists():
        return set()
    existing = set()
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                existing.add(line.lower())
    return existing


def apply():
    if not REVIEW_XLSX.exists():
        raise SystemExit(f"Not found: {REVIEW_XLSX}\nRun 'export' first.")

    print(f"[apply] Reading: {REVIEW_XLSX}")
    wb   = openpyxl.load_workbook(REVIEW_XLSX, data_only=True)
    if "Review" not in wb.sheetnames:
        raise SystemExit("'Review' sheet not found in the workbook.")
    ws   = wb["Review"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Find column indices from header row
    header = [c.value for c in ws[1]]
    try:
        col_tag    = header.index("Tag")
        col_action = header.index("Action")
        col_syn    = header.index("Synonym →")
    except ValueError as e:
        raise SystemExit(f"Expected column not found: {e}")

    new_blacklist = []
    new_synonyms  = []
    new_allowlist = []

    action_count = sum(1 for row in rows if row[col_action] and str(row[col_action]).strip())
    print(f"[apply] {len(rows):,} total rows | {action_count} rows have an Action filled in")

    for row in rows:
        tag    = str(row[col_tag]).strip()    if row[col_tag]    else ""
        action = str(row[col_action]).strip().lower() if row[col_action] else ""
        target = str(row[col_syn]).strip()    if row[col_syn]    else ""

        if not tag or not action:
            continue

        if action == "blacklist":
            new_blacklist.append(tag)
        elif action == "synonym":
            if not target:
                print(f"  [warn] synonym action but no target for '{tag}' — skipped")
                continue
            new_synonyms.append((tag, target))
        elif action == "allowlist":
            new_allowlist.append(tag)

    def _append(path, lines, formatter):
        existing = _load_existing(path)
        added = []
        with open(path, "a", encoding="utf-8") as f:
            for item in lines:
                key = formatter(item).lower()
                if key not in existing:
                    f.write(formatter(item) + "\n")
                    added.append(formatter(item))
                    existing.add(key)
        return added

    added_bl  = _append(BLACKLIST,  new_blacklist,  lambda x: x)
    added_syn = _append(SYNONYMS,   new_synonyms,   lambda x: f"{x[0]} -> {x[1]}")
    added_al  = _append(ALLOWLIST,  new_allowlist,  lambda x: x)

    print(f"[apply] blacklist  : {len(added_bl):>4} new entries added to {BLACKLIST.name}")
    print(f"[apply] synonyms   : {len(added_syn):>4} new entries added to {SYNONYMS.name}")
    print(f"[apply] allowlist  : {len(added_al):>4} new entries added to {ALLOWLIST.name}")

    total = len(added_bl) + len(added_syn) + len(added_al)
    if total == 0:
        print("[apply] Nothing new to add (all selected tags were already in config files).")
    else:
        print(f"\n[apply] Done. {total} entries written.")
        print("        Re-run build_keyword_dataset.py to regenerate a clean dataset.")
        print("        python analysis/build_keyword_dataset.py")

    _write_changelog(added_bl, added_syn, added_al)


# ── Update (apply + rebuild + export in one command) ──────────────────────────

def update():
    """One-command workflow: apply decisions -> rebuild dataset -> export fresh Excel."""
    import importlib.util, sys as _sys

    print("=" * 60)
    print("Step 1/3 — Applying your review decisions ...")
    print("=" * 60)
    apply()

    print()
    print("=" * 60)
    print("Step 2/3 — Rebuilding keyword dataset ...")
    print("=" * 60)

    # Dynamically import and call build_keyword_dataset so we stay in one process
    build_path = ROOT.parent / "analysis" / "build_keyword_dataset.py" \
        if not (ROOT / "analysis" / "build_keyword_dataset.py").exists() \
        else ROOT / "analysis" / "build_keyword_dataset.py"

    spec = importlib.util.spec_from_file_location("build_kw", build_path)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    input_csv    = str(ROOT / "data" / "raw"      / "combined_itineraries_2026-06.csv")
    output_csv   = str(ROOT / "data" / "keywords" / "keyword_dataset_2026-06.csv")
    output_xlsx  = str(ROOT / "data" / "keywords" / "keyword_dataset_2026-06.xlsx")
    review_csv   = str(ROOT / "data" / "keywords" / "keyword_review.csv")
    blacklist_p  = str(ROOT / "data" / "config"   / "keyword_blacklist.txt")
    allowlist_p  = str(ROOT / "data" / "config"   / "keyword_allowlist.txt")
    synonyms_p   = str(ROOT / "data" / "config"   / "keyword_synonyms.txt")

    mod.build_keyword_dataset(input_csv, output_csv, review_csv,
                              blacklist_p, allowlist_p, synonyms_p, output_xlsx)

    print()
    print("=" * 60)
    print("Step 3/3 — Exporting fresh review Excel ...")
    print("=" * 60)
    export()

    print()
    print("=" * 60)
    print("Done! Open data/keywords/keyword_review.xlsx to continue reviewing.")
    print("=" * 60)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Keyword review helper."
    )
    ap.add_argument(
        "command",
        choices=["export", "apply", "update"],
        help=(
            "'export' creates keyword_review.xlsx  |  "
            "'apply'  reads it and updates config files  |  "
            "'update' does apply + rebuild + export in one go"
        ),
    )
    args = ap.parse_args()

    if args.command == "export":
        export()
    elif args.command == "apply":
        apply()
    else:
        update()


if __name__ == "__main__":
    main()
