"""
Export the combined itineraries CSV to a debug-friendly Excel workbook.

Unlike export_excel.py (the styled manager view), this is a plain dump for
checking/debugging the data on Windows:

    Sheet "Summary"   — one row per source: tours + day-row counts
    Sheet "<source>"  — every raw day row for that source, all columns,
                        with a frozen header row and an AutoFilter

Usage:
    python analysis/export_debug_excel.py
    python analysis/export_debug_excel.py --input data/raw/combined_itineraries_latest.csv
    python analysis/export_debug_excel.py --output analysis/itineraries_all_sources.xlsx
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "data" / "raw" / "combined_itineraries_latest.csv"
DEFAULT_OUTPUT = ROOT / "analysis" / "itineraries_all_sources.xlsx"

HEADER_BG = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# Excel tab names: <=31 chars, no : \ / ? * [ ]
_MAX_WIDTH = {"activity": 90, "location": 32, "tour_name": 38, "tour_url": 50}
_DEFAULT_WIDTH = 14


def _style_sheet(ws, df):
    """Freeze the header, bold it, add an AutoFilter, and size columns."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
    for col_idx, name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        cell = ws.cell(1, col_idx)
        cell.fill = HEADER_BG
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[letter].width = _MAX_WIDTH.get(name, _DEFAULT_WIDTH)


def export(input_path: Path, output_path: Path) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")

    df = pd.read_csv(input_path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    print(f"[debug_excel] loaded {len(df)} rows from {input_path.name}")

    # Summary: tours + day rows per source
    summary = (
        df.groupby("source")
        .agg(tours=("tour_url", "nunique"), day_rows=("tour_url", "size"))
        .reset_index()
        .sort_values("day_rows", ascending=False)
    )
    total = pd.DataFrame([{
        "source": "TOTAL",
        "tours": df["tour_url"].nunique(),
        "day_rows": len(df),
    }])
    summary = pd.concat([summary, total], ignore_index=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        _style_sheet(xl.sheets["Summary"], summary)

        for source, sub in df.groupby("source"):
            tab = source[:31]
            sub = sub.sort_values(["tour_name", "day_number"], key=lambda s: s if s.name != "day_number"
                                  else pd.to_numeric(s, errors="coerce"))
            sub.to_excel(xl, sheet_name=tab, index=False)
            _style_sheet(xl.sheets[tab], sub)
            print(f"[debug_excel]   {source:22} {sub['tour_url'].nunique():>4} tours  {len(sub):>5} rows")

    print(f"[debug_excel] saved -> {output_path}")
    return output_path


def main():
    ap = argparse.ArgumentParser(description="Dump combined itineraries to a debug Excel workbook.")
    ap.add_argument("--input", default=str(DEFAULT_INPUT), help="Path to combined CSV")
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output .xlsx path")
    args = ap.parse_args()
    export(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
