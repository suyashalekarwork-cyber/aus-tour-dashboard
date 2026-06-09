"""
Streamlit dashboard — Australian Tour Itineraries.

Run:
    cd Itinenaries_analysis
    streamlit run analysis/dashboard.py
"""

import html
import io
import re
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

# ── Constants ─────────────────────────────────────────────────────────────────

ROOT     = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "data" / "combined_itineraries_latest.csv"

AU_STATES   = ["NSW", "VIC", "QLD", "WA", "SA", "TAS", "NT", "ACT"]
INTL_STATES = {"South Pacific", "International"}

SOURCE_LABELS = {
    "global_journeys":  "Global Journeys",
    "inside_australia": "Inside Australia",
    "australia_com":    "Australia.com",
}
TOURS_PER_PAGE = 50

# ── Data loading ──────────────────────────────────────────────────────────────

@st.cache_data
def load_data():
    """Return (tours_df, days_df). International tours are excluded."""
    raw = pd.read_csv(CSV_PATH, encoding="utf-8-sig", dtype=str)
    raw["day_number"] = pd.to_numeric(raw["day_number"], errors="coerce").fillna(0).astype(int)
    raw["total_days"] = pd.to_numeric(raw["total_days"], errors="coerce").fillna(0).astype(int)
    raw["source"]     = raw["source"].map(SOURCE_LABELS).fillna(raw["source"])

    intl_urls = set(raw.loc[raw["state"].isin(INTL_STATES), "tour_url"])
    days = raw[~raw["tour_url"].isin(intl_urls)].copy()

    def _au_states(s):
        return ", ".join(sorted(v for v in s.dropna().unique() if v in AU_STATES))

    def _cities(s):
        seen, out = set(), []
        for v in s.dropna():
            if v and v not in seen:
                seen.add(v); out.append(v)
        return ", ".join(out[:6])

    tours = (
        days.groupby("tour_url", as_index=False)
        .agg(
            tour_name  = ("tour_name",  "first"),
            source     = ("source",     "first"),
            total_days = ("total_days", "first"),
            price      = ("price",      "first"),
            states     = ("state",      _au_states),
            cities     = ("city",       _cities),
        )
    )

    def _price_num(p):
        if pd.isna(p) or str(p).strip() == "":
            return 0
        d = re.sub(r"[^\d]", "", str(p))
        return int(d) if d else 0

    tours["price_num"]  = tours["price"].apply(_price_num)
    tours["price_disp"] = tours.apply(
        lambda r: r["price"] if (not pd.isna(r["price"]) and str(r["price"]).strip()) else "",
        axis=1,
    )

    return tours, days


# ── Per-tour Excel generation ─────────────────────────────────────────────────

def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_tour_excel(tour: pd.Series, days: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Itinerary"

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = tour["tour_name"]
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = _fill("1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value     = (
        f"  {tour['source']}   |   {tour['total_days']} days"
        f"   |   {tour['price_disp']}   |   {tour['states']}"
    )
    c.font      = Font(bold=True, color="FFFFFF", size=10)
    c.fill      = _fill("2E75B6")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    for col, label in [(1, "Day"), (2, "Location"), (3, "What You'll Do")]:
        c = ws.cell(4, col)
        c.value     = label
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = "A5"

    tour_days = days[days["tour_url"] == tour["tour_url"]].sort_values("day_number")
    for idx, (_, day) in enumerate(tour_days.iterrows()):
        r  = idx + 5
        bg = "EBF3FB" if idx % 2 == 1 else "FFFFFF"
        for col, val in [(1, day["day_number"]), (2, day["location"]), (3, day["activity"])]:
            c = ws.cell(r, col)
            c.value     = val
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="top", wrap_text=(col == 3), indent=(1 if col != 1 else 0))
        ws.row_dimensions[r].height = 15

    wb.save(buf)
    return buf.getvalue()


# ── Detail table (HTML so text wraps fully) ───────────────────────────────────

def _day_table_html(tour_day_rows: pd.DataFrame) -> str:
    rows = ""
    for i, (_, row) in enumerate(tour_day_rows.iterrows()):
        bg  = "#EBF3FB" if i % 2 == 1 else "#FFFFFF"
        day = html.escape(str(int(row["Day"])) if row["Day"] else "")
        loc = html.escape(str(row["Location"]) if pd.notna(row["Location"]) else "")
        act = html.escape(str(row["What You'll Do"]) if pd.notna(row["What You'll Do"]) else "")
        rows += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:8px 10px;border:1px solid #dde;text-align:center;'
            f'vertical-align:top;white-space:nowrap;color:#555;font-weight:600">{day}</td>'
            f'<td style="padding:8px 10px;border:1px solid #dde;vertical-align:top;'
            f'white-space:nowrap;min-width:160px">{loc}</td>'
            f'<td style="padding:8px 10px;border:1px solid #dde;vertical-align:top;'
            f'word-wrap:break-word;line-height:1.5">{act}</td>'
            f'</tr>'
        )
    return (
        '<div style="overflow-x:auto">'
        '<table style="width:100%;border-collapse:collapse;font-size:13.5px">'
        '<thead><tr style="background:#1F3864;color:#fff">'
        '<th style="padding:8px 10px;text-align:center;width:48px">Day</th>'
        '<th style="padding:8px 10px;text-align:left;min-width:160px">Location</th>'
        '<th style="padding:8px 10px;text-align:left">What You\'ll Do</th>'
        '</tr></thead>'
        f'<tbody>{rows}</tbody>'
        '</table></div>'
    )


# ── App ───────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Australian Tour Itineraries",
        page_icon="🇦🇺",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    tours, days = load_data()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## 🇦🇺 Tour Finder")
        st.divider()

        search = st.text_input(
            "Search by tour name", placeholder="e.g. Kimberley, Tasman, reef…",
            help="Searches tour names (case-insensitive)",
        )

        city_search = st.text_input(
            "Search by city", placeholder="e.g. Cairns, Darwin, Perth…",
            help="Searches cities visited on each tour (case-insensitive)",
        )

        min_d = int(tours["total_days"].min() or 1)
        max_d = int(tours["total_days"].max() or 30)
        dur_range = st.slider("Duration (days)", min_d, max_d, (min_d, max_d))

        st.markdown("**State / Territory**")
        _sa_col, _sc_col = st.columns(2)
        if _sa_col.button("Select All", use_container_width=True, key="_sel_all"):
            for _s in AU_STATES:
                st.session_state[f"state_{_s}"] = True
        if _sc_col.button("Clear All", use_container_width=True, key="_sel_clear"):
            for _s in AU_STATES:
                st.session_state[f"state_{_s}"] = False
        sel_states = [s for s in AU_STATES if st.checkbox(s, value=True, key=f"state_{s}")]

        st.divider()

    # ── Filter ────────────────────────────────────────────────────────────────
    filt = tours.copy()
    if search.strip():
        filt = filt[filt["tour_name"].str.contains(search.strip(), case=False, na=False)]
    if city_search.strip():
        filt = filt[filt["cities"].str.contains(city_search.strip(), case=False, na=False)]
    if sel_states:
        filt = filt[filt["states"].apply(
            lambda s: any(code in (s or "").split(", ") for code in sel_states)
        )]
    else:
        filt = filt.iloc[:0]
    filt = filt[filt["total_days"].between(dur_range[0], dur_range[1])]

    # ── CSV download ──────────────────────────────────────────────────────────
    with st.sidebar:
        if len(filt) > 0:
            csv_bytes = (
                filt[["tour_name", "total_days", "states", "cities", "price_disp", "tour_url"]]
                .rename(columns={"price_disp": "price", "total_days": "duration_days"})
                .to_csv(index=False)
                .encode()
            )
            st.download_button(
                label=f"⬇ Download {len(filt)} tours (CSV)",
                data=csv_bytes,
                file_name="filtered_tours.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── Main area ─────────────────────────────────────────────────────────────
    st.markdown("# Australian Tour Itineraries")

    m1, m2, _a, _b = st.columns(4)
    m1.metric("Matching Tours",   len(filt))
    m2.metric("Total in Dataset", len(tours))

    st.divider()

    if len(filt) == 0:
        st.info("No tours match your filters — try broadening your search.")
        return

    # ── Pagination state ──────────────────────────────────────────────────────
    total_pages = max(1, (len(filt) - 1) // TOURS_PER_PAGE + 1)
    filter_sig = (search, city_search, tuple(sorted(sel_states)), dur_range)
    if st.session_state.get("_fsig") != filter_sig:
        st.session_state["_fsig"] = filter_sig
        st.session_state["_page"] = 0
    page = max(0, min(int(st.session_state.get("_page", 0)), total_pages - 1))

    start      = page * TOURS_PER_PAGE
    end        = start + TOURS_PER_PAGE
    page_tours = filt.iloc[start:end].reset_index(drop=True)

    showing_end = min(end, len(filt))
    st.markdown(
        f"**Showing {start+1}–{showing_end} of {len(filt)} tours**"
        + (f"  ·  Page {page+1} of {total_pages}" if total_pages > 1 else "")
        + "  —  *Click any row to view its full itinerary below*"
    )

    # ── Tour table (selectable rows, no Source column) ────────────────────────
    display_df = page_tours[["tour_name", "states", "cities", "total_days"]].copy()
    display_df.columns = ["Tour Name", "States", "Cities", "Days"]

    event = st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        height=700,
        on_select="rerun",
        selection_mode="single-row",
        column_config={
            "Tour Name": st.column_config.TextColumn("Tour Name", width="large"),
            "States":    st.column_config.TextColumn("States",    width="small"),
            "Cities":    st.column_config.TextColumn("Cities",    width="medium"),
            "Days":      st.column_config.NumberColumn("Days",    width="small", format="%d"),
        },
    )

    # ── Detail panel ──────────────────────────────────────────────────────────
    sel_rows = event.selection.rows if hasattr(event, "selection") and event.selection else []
    if sel_rows:
        tour = page_tours.iloc[sel_rows[0]]
        st.divider()

        with st.container(border=True):
            st.markdown(f"### {tour['tour_name']}")

            c1, c2, c3, c4 = st.columns(4)
            c1.markdown(f"**Source:** {tour['source']}")
            c2.markdown(f"**Duration:** {tour['total_days']} days")
            c3.markdown(f"**Price:** {tour['price_disp'] or '—'}")
            c4.markdown(f"**States:** {tour['states'] or 'N/A'}")

            if tour["cities"]:
                st.caption(f"Cities visited: {tour['cities']}")

            # HTML table — text wraps fully on every row
            tour_day_rows = (
                days[days["tour_url"] == tour["tour_url"]]
                .sort_values("day_number")
                [["day_number", "location", "activity"]]
                .rename(columns={
                    "day_number": "Day",
                    "location":   "Location",
                    "activity":   "What You'll Do",
                })
                .reset_index(drop=True)
            )
            st.markdown(_day_table_html(tour_day_rows), unsafe_allow_html=True)
            st.markdown("")   # breathing room after table

            btn1, btn2, _spacer = st.columns([1, 1, 4])
            if _OPENPYXL:
                excel_bytes = make_tour_excel(tour, days)
                safe_name   = re.sub(r'[\\/*?:"<>|]', "", tour["tour_name"])[:60]
                btn1.download_button(
                    label="⬇ Download Excel",
                    data=excel_bytes,
                    file_name=f"{safe_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{tour['tour_url']}",
                    use_container_width=True,
                )
            if str(tour["tour_url"]).startswith("http"):
                btn2.link_button("🔗 View online", tour["tour_url"], use_container_width=True)

    # ── Pagination controls ───────────────────────────────────────────────────
    if total_pages > 1:
        st.divider()
        p_prev, p_label, p_next = st.columns([1, 3, 1])
        if p_prev.button("← Prev", disabled=(page == 0), use_container_width=True):
            st.session_state["_page"] = page - 1
            st.rerun()
        p_label.markdown(
            f"<p style='text-align:center;padding-top:10px'>Page {page+1} of {total_pages}</p>",
            unsafe_allow_html=True,
        )
        if p_next.button("Next →", disabled=(page == total_pages - 1), use_container_width=True):
            st.session_state["_page"] = page + 1
            st.rerun()


if __name__ == "__main__":
    main()
