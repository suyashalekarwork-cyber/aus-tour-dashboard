"""
geo_extract_experiment.py — API-grounded place extraction (50-row experiment).

Tests an alternative to LLM-only place extraction: instead of asking an LLM to
*invent* the place list from raw itinerary prose, we

    1. generate cheap candidate phrases from the text (rule-based, no LLM),
    2. GROUND each candidate against free place databases — local gazetteers
       first (Geoscience Australia, Wikidata, All the Places), then live APIs
       (Nominatim -> Photon -> Overpass -> Wikipedia) only for the long tail,
    3. let an LLM ADJUDICATE each candidate using that real-world evidence and
       label it PLACE / BLACKLIST / SYNONYM,
    4. compare the resulting place lists against the Ollama and NotebookLM
       outputs we already have, on the SAME 50-row sample.

The point is NOT to score each API individually — it is to see whether this
candidate -> grounding -> LLM flow extracts the *correct* places we want.

Usage:
    python analysis/geo_extract_experiment.py                 # n8n cloud LLM (free)
    python analysis/geo_extract_experiment.py --backend ollama
    python analysis/geo_extract_experiment.py --no-shortcut   # run LLM even on gazetteer hits

Everything is cached to data/keywords/geo_api_cache.json, so re-runs are
instant and make zero API calls.

Reference datasets (optional — the run degrades gracefully if absent):
    data/reference/ga_gazetteer.csv     Geoscience Australia gazetteer  (env GEO_GA_GAZETTEER)
    data/reference/wikidata_au.json     Wikidata AU places/attractions  (env GEO_WIKIDATA)
    data/reference/alltheplaces.json    All the Places GeoJSON POIs      (env GEO_ALLTHEPLACES)
"""

import os
import re
import sys
import json
import time
import argparse

import pandas as pd

# benchmark_ner gives us the EXACT same 50-row sampler + the Ollama cache key,
# so our rows line up 1:1 with the Ollama / NotebookLM outputs we compare against.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import benchmark_ner as bn  # noqa: E402  (defines _pick_sample, _cache_key)
from xlsx_io import write_xlsx_safe  # noqa: E402

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE         = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH     = os.path.join(BASE, "data", "raw", "combined_itineraries_2026-06.csv")
OLLAMA_CACHE = os.path.join(BASE, "data", "keywords", "benchmark_ner_cache.json")
NLM_FILE     = os.path.join(BASE, "Gemini_Noteboolm", "testing_files",
                            "notebooklm_extraction_50rows.txt")

REF_DIR      = os.path.join(BASE, "data", "reference")
# National gazetteer: GeoNames Australia dump (AU.txt, headerless TSV) by default;
# a CSV with a name column also works (e.g. the Geoscience Australia gazetteer).
GA_GAZ       = os.environ.get("GEO_GA_GAZETTEER", os.path.join(REF_DIR, "AU.txt"))
WIKIDATA     = os.environ.get("GEO_WIKIDATA",     os.path.join(REF_DIR, "wikidata_au.json"))
# AllThePlaces: directory of per-brand _au.geojson files OR a single merged JSON.
ALLTHEPLACES = os.environ.get("GEO_ALLTHEPLACES", os.path.join(BASE, "alltheplaces.xyz", "output"))
# DCCEEW CAPAD protected areas (directory of per-state xlsx files).
CAPAD_DIR    = os.environ.get("GEO_CAPAD", os.path.join(BASE, "dcceew_data"))

GEO_CACHE    = os.path.join(BASE, "data", "keywords", "geo_api_cache.json")
OUT_XLSX     = os.path.join(BASE, "data", "keywords", "geo_extract_comparison.xlsx")
OUT_CSV      = os.path.join(BASE, "data", "keywords", "geo_extract_candidates.csv")

SAMPLE_SIZE  = 50
OLLAMA_BENCH_MODEL = "qwen2.5:7b"   # model whose answers live in benchmark_ner_cache.json

# Australia bounding box (lon/lat) — generous, includes external territories' mainland spread.
AU_BBOX = (112.0, -44.0, 154.0, -9.0)   # (min_lon, min_lat, max_lon, max_lat)

# ---------------------------------------------------------------------------
# n8n cloud proxy (free Claude proxy — same endpoint the pipeline uses)
# ---------------------------------------------------------------------------
N8N_URL = "https://gtxn8n.yourbestwayhome.com.au/webhook/openai-response-proxy"
N8N_KEY = os.environ.get("N8N_KEY", "Gtx1234*")
OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")

NOMINATIM_UA = "TurtleDownUnder-geo-experiment/1.0 (claude-data@turtledownunder.com.au)"


def _requests():
    """Import requests lazily with a friendly error (it isn't in root requirements.txt)."""
    try:
        import requests
        return requests
    except ImportError:
        sys.exit("This experiment needs the 'requests' package. Run: pip install requests")


# ===========================================================================
# STAGE 2 — candidate generation (rule-based, no LLM)
# ===========================================================================

# Words that are capitalised at the start of a sentence (or in marketing blocks) but are
# NOT places. Used to strip leading/trailing noise and reject single-word junk candidates.
STOP_WORDS = {
    # navigation / itinerary verbs
    "leave", "drive", "driving", "visit", "visiting", "enjoy", "enjoying", "explore",
    "exploring", "discover", "discovering", "depart", "departing", "departure", "arrive",
    "arriving", "arrival", "head", "heading", "continue", "continuing", "travel",
    "travelling", "board", "return", "returning", "see", "seeing", "experience", "spend",
    "spending", "relax", "relaxing", "meet", "join", "walk", "set", "begin", "start",
    "finish", "end", "wake", "rise", "cruise", "sail", "swim", "ride", "hike", "climb",
    "scale", "wander", "venture", "traverse", "chase", "choose", "hop", "fly", "float",
    "drift", "glide", "soar", "jump", "cycle", "pop", "grab", "tuck", "nestle", "delve",
    "book", "step", "watch", "sip", "savour", "sample", "taste", "learn", "try", "listen",
    "admire", "feel", "get", "make", "escape", "spot", "marvel", "laze", "pause", "waste",
    "hang", "gaze", "freshen", "refresh", "sleep", "indulge", "become", "uncover", "swap",
    # time / generic
    "today", "tonight", "tomorrow", "overnight", "morning", "afternoon", "evening",
    "breakfast", "lunch", "dinner", "optional", "free", "day", "days", "night", "nights",
    "tour", "tours", "transfer", "check", "checkin", "welcome", "stop", "take", "stay",
    "staying", "journey", "then", "next", "later", "this", "that", "your", "our", "you",
    "after", "before", "upon", "early", "here", "there", "when", "from", "for", "with",
    "come", "however", "want", "even", "pro", "few", "no", "all", "into", "onto", "out",
    # weekdays / months (full + abbrev)
    "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
    "january", "february", "march", "april", "may", "june", "july", "august",
    "september", "october", "november", "december",
    "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "sept", "oct", "nov", "dec",
    # marketing / deal-block labels
    "deal", "deals", "discount", "special", "rate", "package", "promo", "code", "bonus",
    "offer", "valid", "terms", "save", "off", "sale", "list", "event", "food", "drink",
    "kids", "kid", "adult", "adults", "children", "child", "week", "weekly", "hour",
    "hours", "pack", "per", "duo", "combo", "madness", "affordable", "luxury", "locals",
    "local", "must", "unmissable", "how", "only", "more", "one", "want", "prefer",
    # nationality / language adjectives (not places on their own)
    "australian", "australia", "aussie", "aussies", "french", "asian", "aboriginal",
}
# Lowercase connector words allowed *inside* a place phrase.
CONNECTORS = {"of", "the", "and", "on", "de", "du", "la", "le", "von", "der", "by", "at", "in"}

# Title-cased run, allowing the connectors above to bridge two capitalised tokens.
# NOTE: '.' is deliberately NOT in the token class — a full stop must end a phrase,
# otherwise phrases merge across sentence boundaries ("Rottnest Island. It's ...").
_PHRASE_RE = re.compile(
    r"[A-Z][\w'’&-]*"
    r"(?:\s+(?:of|the|and|on|de|du|la|le|von|der|by|at|in)\s+[A-Z][\w'’&-]*"
    r"|\s+[A-Z][\w'’&-]*)*"
)

# Lightweight location-field tokenizer (mirrors build_keyword_dataset.tokenize_location
# without pulling in spaCy/product_app at import time).
_JOURNEY_SEP = re.compile(r"\s+-\s+")
_NAV_PREFIX = re.compile(
    r"^(arrive(?:\s+in)?|arrival(?:\s+in)?|depart(?:\s+from)?|departure|welcome\s+to|"
    r"to|via|en\s+route\s+to|drive\s+to|fly\s+to|transfer\s+to|overnight\s+in)\s+",
    re.I,
)
_NAV_SUFFIX = re.compile(r"\s+(airport|free\s+time|at\s+leisure|region|area|surrounds)$", re.I)
_ACTIVITYISH = re.compile(r"\b(cruise|tour|flight|transfer|walk|hike|safari|drive)\b", re.I)


def tokenize_location(raw):
    """Split a raw location string into clean place tokens."""
    if not raw or not isinstance(raw, str):
        return []
    out = []
    for chunk in _JOURNEY_SEP.split(raw):
        for tok in re.split(r"[,/]", chunk):
            t = tok.strip()
            t = _NAV_PREFIX.sub("", t).strip()
            t = _NAV_SUFFIX.sub("", t).strip()
            if len(t) < 3:
                continue
            if t.islower():            # all-lowercase => noise
                continue
            if _ACTIVITYISH.search(t):  # embedded activity, not a place
                continue
            out.append(t)
    return out


_STOP_OR_CONN = STOP_WORDS | CONNECTORS | {"am", "pm"}
MAX_PHRASE_WORDS = 6   # real place names rarely exceed this; longer = marketing prose


def _is_allcaps(tok):
    letters = [c for c in tok if c.isalpha()]
    return len(letters) > 1 and tok.upper() == tok


def _dropword(tok):
    """A leading/trailing token to strip: stop word, connector, or ALL-CAPS label."""
    return tok.lower() in _STOP_OR_CONN or _is_allcaps(tok)


def _trim_phrase(phrase):
    """Strip leading/trailing stop/connector/label words and reject obvious non-place noise."""
    # Drop a phrase that is entirely ALL-CAPS (DEAL, TOUR, SPECIAL RATE ...).
    if _is_allcaps(phrase.replace(" ", "")):
        return ""
    toks = phrase.split()
    while toks and _dropword(toks[0]):
        toks.pop(0)
    while toks and _dropword(toks[-1]):
        toks.pop()
    if not toks or len(toks) > MAX_PHRASE_WORDS:
        return ""
    # Drop a single bare token that is a stop word or a contraction/possessive ("It's", "Perth's").
    if len(toks) == 1:
        t = toks[0]
        if t.lower() in STOP_WORDS or re.search(r"['’][a-z]", t):
            return ""
    out = " ".join(toks).strip(" .,&-")
    if len(out) < 4:                      # too short to be a useful place name
        return ""
    return out


def extract_candidates(location, activity):
    """Return a de-duplicated, order-preserving list of candidate place phrases for one row."""
    seen, out = set(), []

    def _add(p):
        p = (p or "").replace("�", "").strip()   # drop mojibake replacement chars
        if not p:
            return
        key = p.lower()
        if key not in seen:
            seen.add(key)
            out.append(p)

    for tok in tokenize_location(location):
        _add(tok)

    text = activity if isinstance(activity, str) else ""
    for m in _PHRASE_RE.finditer(text):
        _add(_trim_phrase(m.group(0)))

    return out


# ===========================================================================
# STAGE 3a — local gazetteers (download/query once, match in-memory)
# ===========================================================================

def _norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _in_au(lon, lat):
    try:
        lon, lat = float(lon), float(lat)
    except (TypeError, ValueError):
        return False
    return AU_BBOX[0] <= lon <= AU_BBOX[2] and AU_BBOX[1] <= lat <= AU_BBOX[3]


def load_ga_gazetteer(path):
    """National gazetteer -> {normalised_name: feature_code}.

    Handles two formats:
      * GeoNames dump (AU.txt): headerless TSV, cols [1]=name [2]=asciiname
        [3]=alternatenames (comma-sep) [7]=feature_code.
      * CSV with a name column (e.g. the Geoscience Australia gazetteer).
    """
    if not os.path.exists(path):
        print(f"  [skip] national gazetteer not found at {path}")
        return {}

    gaz = {}
    if path.lower().endswith(".txt"):  # GeoNames TSV
        try:
            with open(path, encoding="utf-8") as f:
                for line in f:
                    parts = line.rstrip("\n").split("\t")
                    if len(parts) < 8:
                        continue
                    code = parts[7]
                    for nm in (parts[1], parts[2]):
                        n = _norm(nm)
                        if n:
                            gaz.setdefault(n, code)
                    for alt in parts[3].split(","):
                        n = _norm(alt)
                        if n:
                            gaz.setdefault(n, code)
            print(f"  national gazetteer (GeoNames AU): {len(gaz):,} names")
            return gaz
        except Exception as e:
            print(f"  [skip] could not read GeoNames dump: {e}")
            return {}

    try:
        df = pd.read_csv(path, dtype=str, low_memory=False)
    except Exception as e:
        print(f"  [skip] could not read gazetteer CSV: {e}")
        return {}
    name_col = next((c for c in df.columns
                     if c.lower() in ("name", "place_name", "authority_name",
                                      "feature_name", "gazetteer_name")), None) or df.columns[0]
    type_col = next((c for c in df.columns
                     if c.lower() in ("feature", "feature_type", "category", "class")), None)
    for _, r in df.iterrows():
        nm = _norm(r.get(name_col, ""))
        if nm:
            gaz.setdefault(nm, r.get(type_col, "") if type_col else "")
    print(f"  national gazetteer (CSV): {len(gaz):,} names")
    return gaz


def load_wikidata(path):
    """Load (or fetch+cache) Wikidata AU places/attractions -> set of normalised names."""
    if os.path.exists(path):
        try:
            data = json.load(open(path, encoding="utf-8"))
            names = {_norm(n) for n in data.get("names", data if isinstance(data, list) else [])}
            print(f"  Wikidata: {len(names):,} names (cached)")
            return names
        except Exception as e:
            print(f"  [skip] could not read Wikidata cache: {e}")
            return set()
    # Best-effort one-time SPARQL fetch; split by type to keep responses small.
    # Use CSV format (10× smaller than JSON) and labels-only (no aliases) to avoid
    # the Wikidata endpoint's ~700KB response truncation bug.
    print("  Wikidata cache missing — attempting SPARQL fetch (by type, CSV)...")
    # Ten AU place/attraction types — one query per type keeps each response small.
    TYPES = [
        ("tourist attractions", "wd:Q570116"),
        ("archaeological sites",  "wd:Q839954"),
        ("museums",               "wd:Q33506"),
        ("national parks",        "wd:Q46169"),
        ("monuments",             "wd:Q4989906"),
        ("lakes",                 "wd:Q23397"),
        ("mountains",             "wd:Q8502"),
        ("beaches",               "wd:Q40080"),
        ("parks",                 "wd:Q22698"),
        ("heritage sites",        "wd:Q207524"),
    ]
    query_tmpl = """
SELECT DISTINCT ?label WHERE {{
  ?item wdt:P17 wd:Q408 ; wdt:P31/wdt:P279* {type} .
  ?item rdfs:label ?label . FILTER(LANG(?label)="en")
}}
"""
    try:
        import csv as _csv
        import io as _io
        import re as _re
        requests = _requests()
        names = set()
        for type_label, type_qid in TYPES:
            query = query_tmpl.format(type=type_qid)
            r = requests.get(
                "https://query.wikidata.org/sparql",
                params={"query": query, "format": "csv"},
                headers={"User-Agent": NOMINATIM_UA, "Accept": "text/csv"},
                timeout=120,
            )
            r.raise_for_status()
            content = r.content.decode("utf-8", errors="replace")
            count_before = len(names)
            ct = r.headers.get("Content-Type", "")
            if "xml" in ct or content.lstrip().startswith("<"):
                # Wikidata fell back to XML despite format=csv — extract text from literals
                for m in _re.finditer(r"<literal[^>]*>([^<]+)</literal>", content):
                    names.add(_norm(m.group(1)))
            else:
                reader = _csv.reader(_io.StringIO(content))
                next(reader, None)  # skip header
                for row in reader:
                    if row:
                        names.add(_norm(row[0]))
            print(f"    {type_label}: +{len(names)-count_before} names")
            time.sleep(1)
        names.discard("")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        json.dump({"names": sorted(names)}, open(path, "w", encoding="utf-8"))
        print(f"  Wikidata: {len(names):,} names total (fetched + cached)")
        return names
    except Exception as e:
        print(f"  [skip] Wikidata fetch failed ({e}); continuing without it")
        return set()


def load_alltheplaces(path):
    """All the Places GeoJSON -> {normalised_name: brand/category}, AU only.

    Accepts either:
      - a directory of per-brand *_au.geojson files (the alltheplaces.xyz/output layout), or
      - a single merged GeoJSON/JSON file.
    """
    if not os.path.exists(path):
        print(f"  [skip] All the Places not found at {path}")
        return {}

    atp = {}

    def _ingest_features(feats, au_only=False):
        for f in feats:
            props = f.get("properties", f) if isinstance(f, dict) else {}
            geom  = f.get("geometry") if isinstance(f, dict) else None
            if au_only:
                pass  # file is already AU-specific, skip country check
            else:
                country = str(props.get("addr:country") or props.get("country") or "").upper()
                in_au = country in ("AU", "AUS", "AUSTRALIA")
                if not in_au and geom and geom.get("type") == "Point":
                    lon, lat = (geom.get("coordinates") or [None, None])[:2]
                    in_au = _in_au(lon, lat)
                if country and not in_au:
                    return
            for key in ("name", "brand", "official_name", "alt_name"):
                nm = _norm(props.get(key, ""))
                if nm:
                    atp.setdefault(nm, props.get("brand") or props.get("category") or "")

    if os.path.isdir(path):
        import glob as _glob
        au_files = sorted(_glob.glob(os.path.join(path, "*_au.geojson")))
        if not au_files:
            print(f"  [skip] No *_au.geojson files found in {path}")
            return {}
        for fp in au_files:
            try:
                data = json.load(open(fp, encoding="utf-8"))
                feats = data.get("features", data if isinstance(data, list) else [])
                _ingest_features(feats, au_only=True)
            except Exception:
                pass
        print(f"  All the Places: {len(atp):,} AU names ({len(au_files)} files)")
    else:
        try:
            data = json.load(open(path, encoding="utf-8"))
        except Exception as e:
            print(f"  [skip] could not read All the Places: {e}")
            return {}
        feats = data.get("features", data if isinstance(data, list) else [])
        _ingest_features(feats, au_only=False)
        print(f"  All the Places: {len(atp):,} AU names")

    return atp


def load_capad(capad_dir):
    """DCCEEW CAPAD protected areas -> {normalised_name: 'protected_area'}.

    Reads the 'Detailed List' sheet from each per-state capad-2024-terrestrial-*.xlsx
    and extracts the Name column.  Returns {} gracefully if openpyxl is missing or the
    directory is absent.
    """
    if not os.path.exists(capad_dir):
        print(f"  [skip] CAPAD dir not found at {capad_dir}")
        return {}
    try:
        import openpyxl as _xl
        import glob as _glob
    except ImportError:
        print("  [skip] CAPAD needs openpyxl — run: pip install openpyxl")
        return {}

    capad = {}
    files = sorted(_glob.glob(os.path.join(capad_dir, "capad-2024-terrestrial-*.xlsx")))
    if not files:
        print(f"  [skip] No capad-2024-terrestrial-*.xlsx in {capad_dir}")
        return {}

    for fp in files:
        try:
            wb = _xl.load_workbook(fp, read_only=True, data_only=True)
            if "Detailed List" not in wb.sheetnames:
                continue
            ws = wb["Detailed List"]
            header = None
            name_idx = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    if any(str(c or "").strip() == "Name" for c in row):
                        header = [str(c or "").strip() for c in row]
                        name_idx = header.index("Name")
                    continue
                if name_idx is None:
                    break
                v = row[name_idx]
                if v and str(v).strip():
                    nm = _norm(str(v).strip())
                    if nm:
                        capad.setdefault(nm, "protected_area")
        except Exception:
            pass

    print(f"  CAPAD (protected areas): {len(capad):,} names ({len(files)} state files)")
    return capad


# ===========================================================================
# STAGE 3b — live per-candidate APIs (only for Tier-A misses), cached
# ===========================================================================

_last_nominatim = [0.0]   # mutable timestamp for the >=1 req/sec rule


def q_nominatim(cand):
    requests = _requests()
    gap = time.time() - _last_nominatim[0]
    if gap < 1.1:
        time.sleep(1.1 - gap)
    _last_nominatim[0] = time.time()
    try:
        r = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": cand, "format": "json", "addressdetails": 1,
                    "extratags": 1, "limit": 1},
            headers={"User-Agent": NOMINATIM_UA}, timeout=30,
        )
        r.raise_for_status()
        js = r.json()
        if not js:
            return {"found": False}
        top = js[0]
        addr = top.get("address", {})
        return {
            "found": True,
            "display_name": top.get("display_name", ""),
            "class": top.get("class", ""),
            "type": top.get("type", ""),
            "country": addr.get("country", ""),
            "country_code": (addr.get("country_code") or "").upper(),
            "lat": top.get("lat"), "lon": top.get("lon"),
            "importance": top.get("importance"),
        }
    except Exception as e:
        return {"found": False, "error": str(e)[:120]}


def q_photon(cand):
    requests = _requests()
    try:
        r = requests.get("https://photon.komoot.io/api/",
                         params={"q": cand, "limit": 1, "lang": "en"}, timeout=30)
        r.raise_for_status()
        feats = r.json().get("features", [])
        if not feats:
            return {"found": False}
        p = feats[0].get("properties", {})
        return {"found": True, "name": p.get("name", ""),
                "country": p.get("country", ""), "country_code": (p.get("countrycode") or "").upper(),
                "type": p.get("type", ""), "osm_key": p.get("osm_key", "")}
    except Exception as e:
        return {"found": False, "error": str(e)[:120]}


def q_overpass(cand):
    requests = _requests()
    safe = cand.replace('"', '\\"')
    lo_lon, lo_lat, hi_lon, hi_lat = AU_BBOX
    bbox = f"({lo_lat},{lo_lon},{hi_lat},{hi_lon})"
    query = (
        f'[out:json][timeout:25];('
        f'node["name"="{safe}"]["tourism"]{bbox};'
        f'node["name"="{safe}"]["natural"]{bbox};'
        f'node["name"="{safe}"]["leisure"]{bbox};'
        f'way["name"="{safe}"]["tourism"]{bbox};'
        f'way["name"="{safe}"]["natural"]{bbox};'
        f');out tags 1;'
    )
    try:
        r = requests.post("https://overpass-api.de/api/interpreter",
                          data={"data": query}, headers={"User-Agent": NOMINATIM_UA}, timeout=40)
        r.raise_for_status()
        els = r.json().get("elements", [])
        if not els:
            return {"found": False}
        tags = els[0].get("tags", {})
        return {"found": True,
                "kind": tags.get("tourism") or tags.get("natural") or tags.get("leisure", ""),
                "name": tags.get("name", "")}
    except Exception as e:
        return {"found": False, "error": str(e)[:120]}


def q_wikipedia(cand):
    requests = _requests()
    try:
        r = requests.get(
            "https://en.wikipedia.org/w/api.php",
            params={"action": "query", "titles": cand, "prop": "extracts",
                    "exintro": 1, "explaintext": 1, "redirects": 1, "format": "json"},
            headers={"User-Agent": NOMINATIM_UA}, timeout=30,
        )
        r.raise_for_status()
        pages = r.json().get("query", {}).get("pages", {})
        page = next(iter(pages.values()), {})
        if "missing" in page or page.get("pageid") is None:
            return {"found": False}
        extract = (page.get("extract") or "")[:300]
        return {"found": True, "title": page.get("title", ""), "extract": extract}
    except Exception as e:
        return {"found": False, "error": str(e)[:120]}


# ===========================================================================
# STAGE 4 — LLM adjudication (n8n cloud / ollama)
# ===========================================================================

CLASSIFY_SYSTEM = (
    "You validate candidate place names extracted from Australian travel itineraries. "
    "Given a CANDIDATE string and EVIDENCE gathered from geocoding databases, classify it. "
    "Return ONLY compact JSON: {\"label\":\"PLACE|BLACKLIST|SYNONYM\",\"canonical\":\"\",\"reason\":\"\"}. "
    "PLACE = a real geographic place or named attraction in Australia (city, town, suburb, "
    "national park, beach, gorge, mountain, cape, bay, island, river, lake, waterfall, lookout, "
    "landmark, museum, gallery, theatre, precinct, cultural centre, heritage site). "
    "BLACKLIST = NOT an Australian geographic place: hotel/motel/resort/lodge names, restaurant or "
    "bar names, tour-company or marketing phrases, activity verbs, generic words, or any place "
    "located outside Australia. "
    "SYNONYM = an alternate or former name of a canonical Australian place — put the canonical "
    "name in 'canonical' (e.g. 'Ayers Rock' -> 'Uluru', 'Fraser Island' -> \"K'gari\"). "
    "Weigh the evidence: if every database locates it outside Australia, choose BLACKLIST; if the "
    "evidence shows a hotel/resort POI, choose BLACKLIST; if databases confirm an Australian "
    "feature, choose PLACE."
)

_JSON_RE = re.compile(r"\{.*\}", re.S)


def _evidence_text(ev):
    lines = []

    ga = ev.get("ga")
    if ga is not None:
        if ga != "__miss__":
            lines.append("GeoscienceAU gazetteer: MATCH (type={})".format(ga or "n/a"))
        else:
            lines.append("GeoscienceAU gazetteer: no match")

    if ev.get("wikidata") is not None:
        lines.append("Wikidata AU places: MATCH" if ev["wikidata"]
                     else "Wikidata AU places: no match")

    atp = ev.get("alltheplaces")
    if atp is not None:
        if atp != "__miss__":
            lines.append("All the Places (AU POI): MATCH (brand/cat={})".format(atp))
        else:
            lines.append("All the Places (AU POI): no match")

    if ev.get("capad") is not None:
        lines.append("CAPAD (AU protected areas): MATCH" if ev["capad"]
                     else "CAPAD (AU protected areas): no match")

    n = ev.get("nominatim")
    if n:
        if n.get("found"):
            lines.append("Nominatim: FOUND {} [class={}/{}, country={}({})]".format(
                n.get("display_name", ""), n.get("class"), n.get("type"),
                n.get("country"), n.get("country_code")))
        else:
            lines.append("Nominatim: not found")

    p = ev.get("photon")
    if p:
        if p.get("found"):
            lines.append("Photon: FOUND {} [country={}({}), {}]".format(
                p.get("name", ""), p.get("country"), p.get("country_code"), p.get("osm_key")))
        else:
            lines.append("Photon: not found")

    o = ev.get("overpass")
    if o:
        if o.get("found"):
            lines.append("Overpass OSM (AU bbox): FOUND {} ({})".format(
                o.get("name", ""), o.get("kind")))
        else:
            lines.append("Overpass OSM (AU bbox): not found")

    w = ev.get("wikipedia")
    if w:
        if w.get("found"):
            lines.append("Wikipedia: EXISTS - {}".format(w.get("extract", "")))
        else:
            lines.append("Wikipedia: no article")

    return "\n".join(lines) if lines else "(no evidence gathered)"


def _parse_verdict(text):
    m = _JSON_RE.search(text or "")
    if not m:
        return None
    try:
        js = json.loads(m.group(0))
    except json.JSONDecodeError:
        return None
    label = str(js.get("label", "")).upper().strip()
    if label not in ("PLACE", "BLACKLIST", "SYNONYM"):
        return None
    return {"label": label,
            "canonical": str(js.get("canonical", "") or "").strip(),
            "reason": str(js.get("reason", "") or "").strip()[:200]}


def classify_n8n(cand, ev):
    requests = _requests()
    user = f"CANDIDATE: {cand}\nEVIDENCE:\n{_evidence_text(ev)}"
    for attempt in range(3):
        try:
            r = requests.post(N8N_URL,
                              headers={"N8N-KEY": N8N_KEY, "Content-Type": "application/json"},
                              json={"system": CLASSIFY_SYSTEM, "user": user}, timeout=60)
            r.raise_for_status()
            text = r.json()["output"][0]["content"][0]["text"]
            v = _parse_verdict(text)
            if v:
                return v
        except Exception:
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
    return None


def classify_ollama(cand, ev):
    requests = _requests()
    user = (f"CANDIDATE: {cand}\nEVIDENCE:\n{_evidence_text(ev)}\n\n"
            "Return ONLY the JSON object described in the system prompt.")
    body = {"model": OLLAMA_MODEL, "temperature": 0.1, "stream": False, "keep_alive": "60m",
            "messages": [{"role": "system", "content": CLASSIFY_SYSTEM},
                         {"role": "user", "content": user}]}
    for attempt in range(3):
        try:
            r = requests.post(OLLAMA_HOST + "/v1/chat/completions", json=body, timeout=120)
            r.raise_for_status()
            v = _parse_verdict(r.json()["choices"][0]["message"]["content"])
            if v:
                return v
        except Exception:
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
    return None


# ===========================================================================
# Cache helpers
# ===========================================================================

def load_cache(path):
    if os.path.exists(path):
        try:
            return json.load(open(path, encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    json.dump(cache, open(tmp, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    os.replace(tmp, path)


# ===========================================================================
# Comparison inputs (Ollama benchmark cache + NotebookLM extraction)
# ===========================================================================

def load_notebooklm(path):
    """Parse 'ROW N: place1, place2, ...' -> {row_number: [places]}."""
    out = {}
    if not os.path.exists(path):
        print(f"  [skip] NotebookLM file not found at {path}")
        return out
    for line in open(path, encoding="utf-8"):
        m = re.match(r"\s*ROW\s+(\d+)\s*:\s*(.*)", line, re.I)
        if not m:
            continue
        n = int(m.group(1))
        out[n] = [p.strip() for p in m.group(2).split(",") if p.strip()]
    return out


def ollama_places_for(cache, loc, act):
    """Look up what Ollama (qwen2.5:7b) extracted for this row in benchmark_ner_cache.json."""
    key = bn._cache_key(OLLAMA_BENCH_MODEL, loc, act)
    return cache.get(key, [])


# ===========================================================================
# Main
# ===========================================================================

def main():
    # Windows consoles default to cp1252 and choke on Unicode in our output
    # (em-dashes, the set-intersection sign in summary labels, place names).
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--backend", choices=["n8n", "ollama"], default="n8n",
                    help="LLM used for candidate adjudication (default: n8n cloud, free)")
    ap.add_argument("--rows", type=int, default=SAMPLE_SIZE, help="sample size (default 50)")
    ap.add_argument("--no-shortcut", action="store_true",
                    help="run the LLM even when a local gazetteer already confirms the place")
    ap.add_argument("--dry-run", action="store_true",
                    help="generate + print candidates only; no API/LLM calls")
    ap.add_argument("--max-live", type=int, default=400,
                    help="safety cap on candidates sent to live geocoding APIs (default 400)")
    ap.add_argument("--use-overpass", action="store_true",
                    help="also query Overpass on Nominatim+Photon misses (slow; off by default)")
    args = ap.parse_args()

    classify = classify_n8n if args.backend == "n8n" else classify_ollama

    # --- Stage 1: same 50-row sample -------------------------------------
    print("Stage 1 — loading sample")
    df = pd.read_csv(CSV_PATH)
    sample = bn._pick_sample(df, args.rows)
    print(f"  {len(sample)} rows across {sample['state'].nunique()} states")

    # --- Stage 2: candidates ---------------------------------------------
    print("Stage 2 — generating candidates")
    row_candidates = []          # list[list[str]] aligned to sample rows
    unique = {}                  # normalised -> display form
    for _, row in sample.iterrows():
        cands = extract_candidates(row.get("location", ""), row.get("activity", ""))
        row_candidates.append(cands)
        for c in cands:
            unique.setdefault(_norm(c), c)
    print(f"  {sum(len(c) for c in row_candidates)} candidate mentions; "
          f"{len(unique)} unique candidates")

    if args.dry_run:
        print("\n[dry-run] candidates per row:")
        for idx, cands in enumerate(row_candidates, 1):
            print(f"  ROW {idx:>2}: {cands}")
        print(f"\n[dry-run] {len(unique)} unique candidates — no API/LLM calls made.")
        return

    # --- Stage 3a: local gazetteers --------------------------------------
    print("Stage 3a — loading local gazetteers")
    ga   = load_ga_gazetteer(GA_GAZ)
    wd   = load_wikidata(WIKIDATA)
    atp  = load_alltheplaces(ALLTHEPLACES)
    capad = load_capad(CAPAD_DIR)

    # --- Stage 3b + 4: ground + adjudicate each unique candidate ---------
    print(f"Stage 3b/4 — grounding + LLM adjudication ({args.backend})")
    cache = load_cache(GEO_CACHE)
    total = len(unique)
    live_used = 0
    live_capped = 0
    for i, (key, disp) in enumerate(sorted(unique.items()), 1):
        entry = cache.get(key)
        if entry and "verdict" in entry and entry.get("_backend") == args.backend:
            continue  # already done with this backend

        ev = (entry or {}).get("evidence") or {}
        # Tier A — local matches (cheap, always recorded)
        ga_hit = ga.get(key, "__miss__") if ga else None
        ev["ga"] = ga_hit
        ev["wikidata"] = (key in wd) if wd else None
        atp_hit = atp.get(key, "__miss__") if atp else None
        ev["alltheplaces"] = atp_hit
        ev["capad"] = (key in capad) if capad else None

        tier_a_hit = (ga_hit not in (None, "__miss__")) or \
                     (ev["wikidata"] is True) or \
                     (atp_hit not in (None, "__miss__")) or \
                     (ev["capad"] is True)

        # Safety cap: once we've sent --max-live candidates to live APIs, stop calling
        # them. Remaining ungrounded candidates are recorded as such (no silent coverage).
        already_live = any(k in ev for k in ("nominatim", "photon", "overpass", "wikipedia"))
        if not tier_a_hit and not already_live and live_used >= args.max_live:
            live_capped += 1
            cache[key] = {"display": disp, "evidence": ev,
                          "verdict": {"label": "BLACKLIST", "canonical": "",
                                      "reason": "over --max-live cap; not grounded"},
                          "_backend": args.backend}
            continue

        # Tier B — live APIs only when Tier A missed (long tail)
        if not tier_a_hit:
            if not already_live:
                live_used += 1
            if "nominatim" not in ev:
                ev["nominatim"] = q_nominatim(disp)
            if not ev["nominatim"].get("found") and "photon" not in ev:
                ev["photon"] = q_photon(disp)
            # Overpass is opt-in: it only fires on Nominatim+Photon misses (mostly noise),
            # and hundreds of single-name queries against the shared endpoint are slow/impolite.
            if (args.use_overpass and not ev["nominatim"].get("found")
                    and not ev.get("photon", {}).get("found") and "overpass" not in ev):
                ev["overpass"] = q_overpass(disp)
            if (not ev["nominatim"].get("found")
                    and not ev.get("photon", {}).get("found")
                    and not ev.get("overpass", {}).get("found") and "wikipedia" not in ev):
                ev["wikipedia"] = q_wikipedia(disp)

        # Adjudicate
        if tier_a_hit and not args.no_shortcut:
            src = "GA gazetteer" if ga_hit not in (None, "__miss__") else \
                  "Wikidata" if ev["wikidata"] else \
                  "CAPAD (protected area)" if ev.get("capad") else "All the Places"
            verdict = {"label": "PLACE", "canonical": "",
                       "reason": f"local gazetteer hit ({src})"}
        else:
            verdict = classify(disp, ev) or {"label": "BLACKLIST", "canonical": "",
                                             "reason": "no confident evidence / LLM unavailable"}

        cache[key] = {"display": disp, "evidence": ev, "verdict": verdict,
                      "_backend": args.backend}
        if i % 10 == 0 or i == total:
            save_cache(cache, GEO_CACHE)
            print(f"  {i}/{total} candidates resolved (live API used: {live_used})")
    save_cache(cache, GEO_CACHE)
    if live_capped:
        print(f"  [warn] {live_capped} candidates skipped live grounding (--max-live={args.max_live}); "
              f"raise --max-live or add local gazetteers to cover them.")

    # --- Stage 5: aggregate + compare ------------------------------------
    print("Stage 5 — building comparison report")
    ollama_cache = load_cache(OLLAMA_CACHE)
    nlm = load_notebooklm(NLM_FILE)

    def synonym_canonical(key, disp):
        v = cache.get(key, {}).get("verdict", {})
        if v.get("label") == "SYNONYM" and v.get("canonical"):
            return v["canonical"]
        return disp

    rows_out, cand_rows = [], []
    for idx, (_, row) in enumerate(sample.iterrows()):
        loc = row.get("location", "")
        act = row.get("activity", "") if isinstance(row.get("activity", ""), str) else ""
        geo_places = []
        for c in row_candidates[idx]:
            key = _norm(c)
            v = cache.get(key, {}).get("verdict", {})
            if v.get("label") == "PLACE":
                geo_places.append(c)
            elif v.get("label") == "SYNONYM":
                geo_places.append(synonym_canonical(key, c))
            # per-candidate detail row
            ev = cache.get(key, {}).get("evidence", {})
            cand_rows.append({
                "row": idx + 1, "candidate": c,
                "label": v.get("label", ""), "canonical": v.get("canonical", ""),
                "reason": v.get("reason", ""),
                "ga": "" if ev.get("ga") in (None, "__miss__") else "hit",
                "wikidata": "hit" if ev.get("wikidata") else "",
                "alltheplaces": "" if ev.get("alltheplaces") in (None, "__miss__") else "hit",
                "capad": "hit" if ev.get("capad") else "",
                "nominatim_country": ev.get("nominatim", {}).get("country_code", ""),
                "photon_country": ev.get("photon", {}).get("country_code", ""),
                "wiki": "exists" if ev.get("wikipedia", {}).get("found") else "",
            })
        # de-dup geo_places preserving order
        seen, gp = set(), []
        for p in geo_places:
            if p.lower() not in seen:
                seen.add(p.lower()); gp.append(p)

        ollama_p = ollama_places_for(ollama_cache, loc, act)
        nlm_p = nlm.get(idx + 1, [])
        gp_set = {p.lower() for p in gp}
        ol_set = {p.lower() for p in ollama_p}
        nl_set = {p.lower() for p in nlm_p}
        rows_out.append({
            "row": idx + 1,
            "location": loc,
            "activity": act[:200],
            "geo+llm places": "; ".join(gp),
            "ollama places": "; ".join(ollama_p),
            "notebooklm places": "; ".join(nlm_p),
            "geo∩ollama": len(gp_set & ol_set),
            "geo∩nlm": len(gp_set & nl_set),
            "geo only": "; ".join(sorted(gp_set - ol_set - nl_set)),
            "missed (in others, not geo)": "; ".join(sorted((ol_set | nl_set) - gp_set)),
        })

    comp_df = pd.DataFrame(rows_out)
    cand_df = pd.DataFrame(cand_rows)

    # Summary
    def _uni(col):
        s = set()
        for v in comp_df[col]:
            s |= {x.strip().lower() for x in str(v).split(";") if x.strip()}
        return s
    geo_u, ol_u, nl_u = _uni("geo+llm places"), _uni("ollama places"), _uni("notebooklm places")
    label_counts = cand_df["label"].value_counts().to_dict() if not cand_df.empty else {}
    summary = pd.DataFrame([
        {"metric": "rows", "value": len(comp_df)},
        {"metric": "unique candidates", "value": len(unique)},
        {"metric": "candidates -> PLACE", "value": label_counts.get("PLACE", 0)},
        {"metric": "candidates -> BLACKLIST", "value": label_counts.get("BLACKLIST", 0)},
        {"metric": "candidates -> SYNONYM", "value": label_counts.get("SYNONYM", 0)},
        {"metric": "unique places (geo+llm)", "value": len(geo_u)},
        {"metric": "unique places (ollama)", "value": len(ol_u)},
        {"metric": "unique places (notebooklm)", "value": len(nl_u)},
        {"metric": "geo ∩ ollama", "value": len(geo_u & ol_u)},
        {"metric": "geo ∩ notebooklm", "value": len(geo_u & nl_u)},
        {"metric": "in ollama/nlm but NOT geo", "value": len((ol_u | nl_u) - geo_u)},
        {"metric": "in geo but NOT ollama/nlm", "value": len(geo_u - ol_u - nl_u)},
    ])

    def _build(p):
        with pd.ExcelWriter(p, engine="openpyxl") as xl:
            summary.to_excel(xl, sheet_name="summary", index=False)
            comp_df.to_excel(xl, sheet_name="comparison", index=False)
            cand_df.to_excel(xl, sheet_name="candidates", index=False)
    written = write_xlsx_safe(_build, OUT_XLSX)
    cand_df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    print(f"\nDone.\n  Excel: {written}\n  CSV:   {OUT_CSV}")
    print("  Summary:")
    for _, r in summary.iterrows():
        print(f"    {r['metric']:<32} {r['value']}")


if __name__ == "__main__":
    main()
