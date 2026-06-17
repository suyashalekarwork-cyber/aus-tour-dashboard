"""
City Review Tool — curate the city/state data quality (Track 1).

Two commands:

  python analysis/city_review_tool.py export
      Builds data/config/city_review.xlsx with these working sheets:
        • Cities      — every city; type a new name in 'Change To' to rename/merge it
        • States      — same, for the 'state' column
        • City Types  — classify each canonical city (city/region/national_park/route/island)
        • Null Audit  — every blank-city day, with a candidate city + fixable/generic flag

  python analysis/city_review_tool.py apply
      Reads the sheets and writes:
        • Cities  'Change To' -> data/config/city_synonyms.txt
        • States  'Change To' -> data/config/state_synonyms.txt
        • City Types 'Type'   -> data/config/city_types.txt
      (+ a timestamped block in data/config/city_changelog.txt). Re-run the pipeline
      (python run_pipeline.py) to propagate.

City/state renames are config-driven: fill 'Change To' on the Cities/States sheet and
run apply; the pipeline applies them via combine_sources.apply_city_synonyms /
apply_state_synonyms. The canonical city vocabulary is owned by _PLACES_RAW in
webscraping/pipeline/combine_sources.py; gazetteer-miss additions surfaced in the Null
Audit are still applied by editing _PLACES_RAW by hand.
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    raise SystemExit("pandas not installed. Run: pip install pandas")

ROOT        = Path(__file__).resolve().parent.parent
PIPELINE    = ROOT / "webscraping" / "pipeline"
CONFIG_DIR  = ROOT / "data" / "config"
KEYWORDS    = ROOT / "data" / "keywords"
REVIEW_XLSX   = CONFIG_DIR / "city_review.xlsx"
CITY_TYPES     = CONFIG_DIR / "city_types.txt"
CITY_SYNONYMS  = CONFIG_DIR / "city_synonyms.txt"
STATE_SYNONYMS = CONFIG_DIR / "state_synonyms.txt"
CHANGELOG      = CONFIG_DIR / "city_changelog.txt"

VALID_TYPES = ["city", "region", "national_park", "route", "island"]

# Import the gazetteer so the city vocabulary stays single-sourced.
sys.path.insert(0, str(PIPELINE))
import combine_sources as cs   # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from xlsx_io import write_xlsx_safe   # noqa: E402

# ── Colour palette (matches keyword_review_tool.py) ───────────────────────────
HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
ROW_ALT   = "F2F7FF"
TYPE_FILL = {
    "city":          "FFFFFF",
    "region":        "FFF3CD",   # amber
    "national_park": "D6F0D6",   # green
    "route":         "E2D9F3",   # purple
    "island":        "D7E9FF",   # blue
}
FIXABLE_FILL = "FFF3CD"
GENERIC_FILL = "E7E7E7"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, text):
    cell.value     = text
    cell.font      = Font(bold=True, color=HEADER_FG, size=10)
    cell.fill      = _fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()


# ── Shared helpers ────────────────────────────────────────────────────────────

_REGION_WORDS = re.compile(
    r"\b(ranges|range|valley|coast|peninsula|region|outback|gorge|reef|"
    r"promontory|hinterland|plateau|hills)\b", re.IGNORECASE)
_ROUTE_WORDS  = re.compile(r"\b(road|drive|way|highway|track|trail|loop)\b", re.IGNORECASE)
_GENERIC_DAY  = re.compile(
    r"\b(day at sea|free day|farewell|in transit|depart|arrive home|at leisure|"
    r"day at your own|own arrangements|check out|free time)\b", re.IGNORECASE)


def _heuristic_type(name: str) -> str:
    """Best-guess type for a city not yet in city_types.txt."""
    n = name.lower()
    if "national park" in n or n.endswith(" np"):
        return "national_park"
    if _ROUTE_WORDS.search(n):
        return "route"
    if "island" in n:
        return "island"
    if _REGION_WORDS.search(n):
        return "region"
    return "city"


def _load_dataset() -> pd.DataFrame:
    """Load the newest keyword_dataset_*.csv (has city + place_tags + tour_url)."""
    candidates = sorted(KEYWORDS.glob("keyword_dataset_*.csv"))
    if not candidates:
        raise SystemExit(
            f"No keyword_dataset_*.csv found in {KEYWORDS}.\n"
            "Run the pipeline first: python run_pipeline.py")
    path = candidates[-1]
    print(f"[data] reading {path.name}")
    df = pd.read_csv(path)
    for col in ("city", "place_tags", "location", "activity", "tour_name", "tour_url", "source"):
        if col not in df.columns:
            df[col] = ""
    return df


# ── EXPORT ────────────────────────────────────────────────────────────────────

def _sheet_city_types(wb, df, existing):
    ws = wb.create_sheet("City Types")
    cities = sorted({c.strip() for _, _, c in cs._PLACES_RAW if c and c.strip()}, key=str.lower)

    # Frequencies from the dataset
    city_norm = df["city"].fillna("").astype(str).str.strip()
    days  = city_norm.value_counts()
    tours = df.assign(_c=city_norm).groupby("_c")["tour_url"].nunique()

    headers = ["City", "State", "Freq (days)", "Freq (tours)", "Suggested Type", "Type"]
    widths  = [26, 8, 12, 13, 16, 16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws.cell(1, col), h)
        ws.column_dimensions[ws.cell(1, col).coordinate[0]].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(cities) + 1}"

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(VALID_TYPES),
                        allow_blank=True, showErrorMessage=True,
                        errorTitle="Invalid type",
                        error="Choose: " + ", ".join(VALID_TYPES))
    ws.add_data_validation(dv)
    dv.sqref = f"F2:F{len(cities) + 1}"

    # Colour the Type cell by its value
    for t, hexc in TYPE_FILL.items():
        if t == "city":
            continue
        ws.conditional_formatting.add(
            f"F2:F{len(cities) + 1}",
            FormulaRule(formula=[f'$F2="{t}"'], fill=_fill(hexc)))

    for i, city in enumerate(cities):
        r = i + 2
        suggested = existing.get(city.lower()) or _heuristic_type(city)
        values = [
            city,
            cs._CITY_STATE.get(city, ""),
            int(days.get(city, 0)),
            int(tours.get(city, 0)),
            suggested,
            suggested,                # pre-fill Type with the current best guess
        ]
        bg = ROW_ALT if i % 2 else "FFFFFF"
        for col, val in enumerate(values, 1):
            c = ws.cell(r, col, value=val)
            c.fill = _fill(bg)
            c.border = _border()
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col in (2, 3, 4) else "left")
    print(f"[export]   City Types     : {len(cities)} cities")


def _sheet_null_audit(wb, df):
    ws = wb.create_sheet("Null Audit")
    blanks = df[df["city"].fillna("").astype(str).str.strip() == ""].copy()

    def candidate(tags):
        tags = str(tags) if tags and str(tags) != "nan" else ""
        if not tags:
            return ""
        first = ""
        for tag in tags.split("; "):
            tag = tag.strip()
            if not tag:
                continue
            first = first or tag
            _, c = cs._find_place(tag)
            if c:
                return c            # a known gazetteer place -> recoverable city
        return first                # unknown -> gazetteer-miss candidate

    def classify(loc, act):
        text = f"{loc} {act}"
        return "generic_day" if _GENERIC_DAY.search(text) else "fixable"

    blanks["candidate_city"] = blanks["place_tags"].map(candidate)
    blanks["classification"] = [classify(l, a) for l, a in
                                zip(blanks["location"].fillna(""), blanks["activity"].fillna(""))]

    # Summary by source at the top
    summary = blanks.groupby("source").size().sort_values(ascending=False)
    fixable = (blanks["classification"] == "fixable").sum()
    ws.cell(1, 1, value="BLANK-CITY AUDIT").font = Font(bold=True, size=12)
    ws.cell(2, 1, value=f"{len(blanks)} blank-city rows  |  {fixable} fixable  |  "
                        f"{len(blanks) - fixable} generic_day").font = Font(size=10)
    r = 3
    for src, n in summary.items():
        ws.cell(r, 1, value=f"   {src}").font = Font(size=10)
        ws.cell(r, 2, value=int(n)).font = Font(size=10)
        r += 1
    table_start = r + 1

    headers = ["source", "tour_name", "day_number", "location", "place_tags",
               "candidate_city", "classification"]
    widths  = [16, 34, 11, 34, 40, 18, 14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws.cell(table_start, col), h)
        ws.column_dimensions[ws.cell(table_start, col).coordinate[0]].width = w
    ws.freeze_panes = ws.cell(table_start + 1, 1).coordinate

    cls_col = len(headers)  # 'classification' column index
    rng = f"A{table_start + 1}:G{table_start + len(blanks)}"
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'$G{table_start + 1}="fixable"'], fill=_fill(FIXABLE_FILL)))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'$G{table_start + 1}="generic_day"'], fill=_fill(GENERIC_FILL)))

    blanks = blanks.sort_values(["classification", "source"])
    for i, (_, row) in enumerate(blanks.iterrows()):
        rr = table_start + 1 + i
        values = [
            row.get("source", ""), str(row.get("tour_name", ""))[:120],
            row.get("day_number", ""), str(row.get("location", ""))[:120],
            str(row.get("place_tags", ""))[:200],
            row.get("candidate_city", ""), row.get("classification", ""),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(rr, col, value=val)
            c.border = _border()
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 4, 5)))
    print(f"[export]   Null Audit     : {len(blanks)} blank-city rows ({fixable} fixable)")


def _dup_suggestion(value, freq):
    """Return the higher-frequency value that `value` looks like a duplicate of
    (substring or >0.85 similar), or '' when there's no good candidate."""
    vl = value.lower()
    best, best_freq = "", -1
    for other, ofreq in freq.items():
        if other == value:
            continue
        ol = other.lower()
        if vl == ol:
            continue
        is_dup = (vl in ol or ol in vl) or SequenceMatcher(None, vl, ol).ratio() > 0.85
        # only suggest merging the smaller value INTO the larger one
        if is_dup and ofreq >= freq.get(value, 0) and ofreq > best_freq:
            best, best_freq = other, ofreq
    return best


def _sheet_rename(wb, sheet_name, label, df, column):
    """A flat 'rename' sheet: every distinct value in df[column] with its frequency,
    a blank 'Change To' column to fill, and a 'Reason' note (pre-filled with a hint for
    likely duplicates). Used for both the Cities and States sheets."""
    ws = wb.create_sheet(sheet_name)
    norm = df[column].fillna("").astype(str).str.strip()
    norm = norm[norm != ""]
    freq = {v: int(n) for v, n in norm.value_counts().items()}   # sorted desc
    values = list(freq.keys())

    headers = [label, "Days", "Change To", "Reason / hint"]
    widths  = [28, 8, 22, 52]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws.cell(1, col), h)
        ws.column_dimensions[ws.cell(1, col).coordinate[0]].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(values) + 1}"

    hints = 0
    for i, v in enumerate(values):
        r = i + 2
        sug = _dup_suggestion(v, freq)
        reason = f"possible duplicate of {sug} — set Change To = {sug} to merge" if sug else ""
        if sug:
            hints += 1
        bg = ROW_ALT if i % 2 else "FFFFFF"
        for col, val in enumerate([v, freq[v], "", reason], 1):
            c = ws.cell(r, col, value=val)
            c.fill = _fill(bg)
            c.border = _border()
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col == 2 else "left",
                                    wrap_text=(col == 4))
    print(f"[export]   {sheet_name:<14}: {len(values)} {label.lower()} values ({hints} duplicate hints)")


def _sheet_help(wb):
    ws = wb.create_sheet("How to Use", 0)
    lines = [
        ("How to use this workbook", True),
        ("", False),
        ("CITIES sheet  ← rename / merge cities here", True),
        ("  • Every city, most frequent first. To RENAME or MERGE one, type the name you", False),
        ("    want in the 'Change To' column (e.g. on the Kimberley row, write Broome).", False),
        ("  • Leave 'Change To' blank to keep a city unchanged.", False),
        ("  • 'Reason / hint' flags likely duplicates as a suggestion — it is only a hint;", False),
        ("    nothing happens unless you fill 'Change To'.", False),
        ("", False),
        ("STATES sheet  ← rename / standardise states here", True),
        ("  • Same as the Cities sheet, for the 'state' column.", False),
        ("", False),
        ("Then save and run:  python analysis/city_review_tool.py apply", True),
        ("  • Cities 'Change To' -> data/config/city_synonyms.txt", False),
        ("  • States 'Change To' -> data/config/state_synonyms.txt", False),
        ("  • Then re-run the pipeline to propagate:  python run_pipeline.py", False),
        ("", False),
        ("CITY TYPES sheet", True),
        ("  • Set the 'Type' for every city: city, region, national_park, route, island.", False),
        ("  • 'Suggested Type' is a starting guess — correct the 'Type' column where wrong.", False),
        ("  • On apply, writes data/config/city_types.txt.", False),
        ("", False),
        ("NULL AUDIT sheet  (read-only worklist)", True),
        ("  • Every tour-day with a blank city.", False),
        ("  • 'fixable' = a place was recoverable; 'generic_day' = legitimately city-less.", False),
        ("  • 'candidate_city' = the place we'd assign. If it is NOT already a gazetteer", False),
        ("    city, add it to _PLACES_RAW in webscraping/pipeline/combine_sources.py.", False),
    ]
    ws.column_dimensions["A"].width = 90
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(i, 1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(wrap_text=True)


def export():
    df = _load_dataset()
    existing = cs._load_city_types()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)            # drop default sheet
    _sheet_help(wb)
    _sheet_rename(wb, "Cities", "City", df, "city")
    _sheet_rename(wb, "States", "State", df, "state")
    _sheet_city_types(wb, df, existing)
    _sheet_null_audit(wb, df)
    wb.active = wb["Cities"]

    saved = write_xlsx_safe(lambda p: wb.save(p), REVIEW_XLSX)
    print(f"[export] saved -> {saved}")
    print("         Fill the 'Type' column, then run: python analysis/city_review_tool.py apply")


# ── APPLY ─────────────────────────────────────────────────────────────────────

def _write_changelog(type_changes, synonym_changes=()):
    """synonym_changes: list of (label, variant, canonical), e.g. ('city', 'kimberley', 'Broome')."""
    if not type_changes and not synonym_changes:
        return
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(CHANGELOG, "a", encoding="utf-8") as f:
        f.write(f"\n=== {ts} ===\n")
        for city, old, new in type_changes:
            f.write(f"  [city_type]      {city}: {old or '(none)'} -> {new}\n")
        for label, variant, canonical in synonym_changes:
            f.write(f"  [{label}_synonym]  {variant} -> {canonical}\n")
        total = len(type_changes) + len(synonym_changes)
        f.write(f"  --- {total} changes ---\n")
    print(f"[apply] Changelog updated -> {CHANGELOG.name}")


def _write_synonyms(path, existing, new_pairs, header_lines):
    """Merge new {variant_lower: canonical} pairs into `existing`, rewrite `path` with the
    given header, and return the list of (variant, canonical) pairs newly added or changed."""
    changed = []
    for variant_lower, canonical in new_pairs.items():
        if existing.get(variant_lower) != canonical:
            changed.append((variant_lower, canonical))
        existing[variant_lower] = canonical

    lines = list(header_lines) + [""]
    for variant_lower in sorted(existing, key=str.lower):
        lines.append(f"{variant_lower} -> {existing[variant_lower]}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return changed


def _write_city_synonyms(new_pairs):
    """Merge city rename pairs into city_synonyms.txt. Returns changed (variant, canonical)."""
    return _write_synonyms(
        CITY_SYNONYMS, cs._load_city_synonyms(), new_pairs,
        ["# City name canonicalization — Variant -> Canonical City (case-insensitive)",
         "# Curated via analysis/city_review_tool.py (Cities sheet) or edited by hand.",
         "# Applied to the 'city' column in combine_sources.py before the city_type lookup.",
         "# Empty / all-commented = no-op. Re-run the pipeline after editing."])


def _write_state_synonyms(new_pairs):
    """Merge state rename pairs into state_synonyms.txt. Returns changed (variant, canonical)."""
    return _write_synonyms(
        STATE_SYNONYMS, cs._load_state_synonyms(), new_pairs,
        ["# State name canonicalization — Variant -> Canonical State (case-insensitive)",
         "# Curated via analysis/city_review_tool.py (States sheet) or edited by hand.",
         "# Applied to the 'state' column in combine_sources.py.",
         "# Empty / all-commented = no-op. Re-run the pipeline after editing."])


def _write_city_types(mapping):
    """Rewrite city_types.txt grouped by type, preserving the file's header."""
    by_type = defaultdict(list)
    for city, ctype in mapping.items():
        by_type[ctype].append(city)
    labels = {"city": "Cities & towns",
              "region": "Regions / valleys / ranges / coasts",
              "national_park": "National parks & natural-feature parks",
              "island": "Islands",
              "route": "Touring routes / tracks / drives"}
    lines = ["# City type classification — Canonical City -> type",
             "# Types: " + " | ".join(VALID_TYPES),
             "# Curated via analysis/city_review_tool.py",
             "# Source of truth for the city vocabulary is _PLACES_RAW in combine_sources.py",
             ""]
    for t in VALID_TYPES:
        cities = sorted(by_type.get(t, []), key=str.lower)
        if not cities:
            continue
        lines.append(f"# --- {labels.get(t, t)} ---")
        lines.extend(f"{c} -> {t}" for c in cities)
        lines.append("")
    CITY_TYPES.write_text("\n".join(lines), encoding="utf-8")


def _read_rename_sheet(wb, sheet_name, label):
    """Read a rename sheet ('Cities'/'States'): return {variant_lower: canonical} from the
    'Change To' column, skipping blanks and no-op (Change To == name) rows."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    try:
        col_name = header.index(label)
        col_to   = header.index("Change To")
    except ValueError as e:
        print(f"  [warn] '{sheet_name}' sheet missing column ({e}) — skipped")
        return {}
    pairs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (str(row[col_name]).strip() if row[col_name] else "")
        to   = (str(row[col_to]).strip() if col_to < len(row) and row[col_to] else "")
        if name and to and name.lower() != to.lower():
            pairs[name.lower()] = to
    return pairs


def apply():
    if not REVIEW_XLSX.exists():
        raise SystemExit(f"Not found: {REVIEW_XLSX}\nRun 'export' first.")
    print(f"[apply] reading {REVIEW_XLSX}")
    wb = openpyxl.load_workbook(REVIEW_XLSX, data_only=True)
    if "City Types" not in wb.sheetnames:
        raise SystemExit("'City Types' sheet not found.")
    ws = wb["City Types"]
    header = [c.value for c in ws[1]]
    try:
        col_city = header.index("City")
        col_type = header.index("Type")
    except ValueError as e:
        raise SystemExit(f"Expected column not found: {e}")

    # Start from the canonical city set so unfilled rows keep a sensible default.
    existing = cs._load_city_types()                       # lower -> type
    display  = {}                                          # CanonicalCity -> type
    for _, _, c in cs._PLACES_RAW:
        c = (c or "").strip()
        if c:
            display.setdefault(c, existing.get(c.lower(), _heuristic_type(c)))

    changes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        city = (str(row[col_city]).strip() if row[col_city] else "")
        ctype = (str(row[col_type]).strip().lower() if row[col_type] else "")
        if not city or not ctype:
            continue
        if ctype not in VALID_TYPES:
            print(f"  [warn] '{city}': invalid type '{ctype}' — skipped")
            continue
        old = display.get(city, existing.get(city.lower(), ""))
        if old != ctype:
            changes.append((city, old, ctype))
        display[city] = ctype

    _write_city_types({c: t for c, t in display.items()})
    print(f"[apply] wrote {len(display)} entries -> {CITY_TYPES.name} ({len(changes)} changed)")

    # Cities / States sheets — apply 'Change To' renames to the synonym files.
    syn_changes = []
    city_pairs = _read_rename_sheet(wb, "Cities", "City")
    if city_pairs:
        chg = _write_city_synonyms(city_pairs)
        syn_changes += [("city", v, c) for v, c in chg]
        print(f"[apply] wrote {len(chg)} city rename(s) -> {CITY_SYNONYMS.name}")
    state_pairs = _read_rename_sheet(wb, "States", "State")
    if state_pairs:
        chg = _write_state_synonyms(state_pairs)
        syn_changes += [("state", v, c) for v, c in chg]
        print(f"[apply] wrote {len(chg)} state rename(s) -> {STATE_SYNONYMS.name}")

    _write_changelog(changes, syn_changes)
    if changes or syn_changes:
        print("        Re-run the pipeline to propagate changes: python run_pipeline.py")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="City review helper (Track 1 data quality).")
    ap.add_argument("command", choices=["export", "apply"],
                    help="'export' builds city_review.xlsx | 'apply' writes city_types.txt")
    args = ap.parse_args()
    export() if args.command == "export" else apply()


if __name__ == "__main__":
    main()
