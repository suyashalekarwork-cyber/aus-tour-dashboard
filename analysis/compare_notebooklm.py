"""
compare_notebooklm.py — Compare NotebookLM place classification vs Ollama NER.

STEP 1 (run this first):
    python analysis/compare_notebooklm.py

    Generates: data/keywords/notebooklm_input_50rows.txt
    This file contains the same 50 rows that were fed to Ollama, formatted
    for pasting into NotebookLM.

STEP 2 (you do this):
    Open NotebookLM, paste the contents of notebooklm_input_50rows.txt,
    copy the response, and save it to:
        Gemini_Noteboolm/notebooklm_extraction_50rows.txt

STEP 3 (run this again after saving the file):
    python analysis/compare_notebooklm.py

    Generates: data/keywords/notebooklm_comparison.xlsx
    Side-by-side Excel comparing what Ollama found vs what NotebookLM found.
"""

import os, re, json, hashlib
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH    = os.path.join(BASE, "data", "raw", "combined_itineraries_2026-06.csv")
CACHE_FILE  = os.path.join(BASE, "data", "keywords", "benchmark_ner_cache.json")
REPORT_FILE = os.path.join(BASE, "Gemini_Noteboolm", "notebooklm_report.txt")
CHAT_FILE   = os.path.join(BASE, "Gemini_Noteboolm", "notebooklm_chat.txt")
INPUT_50    = os.path.join(BASE, "data", "keywords", "notebooklm_input_50rows.txt")
EXTRACT_OUT = os.path.join(BASE, "Gemini_Noteboolm", "notebooklm_extraction_50rows.txt")
AIP_FILE    = os.path.join(BASE, "Gemini_Noteboolm", "ai_playground.txt")
FLASH_FILE  = os.path.join(BASE, "Gemini_Noteboolm", "ai_playground_3.5flash.txt")
OUT_XLSX    = os.path.join(BASE, "data", "keywords", "notebooklm_comparison.xlsx")

OLLAMA_MODEL = "qwen2.5:7b"
SAMPLE_SIZE  = 50

# Conflict overrides: items where the two NotebookLM sources disagree.
# These are our manual verdicts.
FORCE_PLACE = {
    "queenstown",      # Tasmania has a Queenstown (not just NZ)
    "wintjiri wiru",   # Indigenous cultural experience at Uluru
    "yulara",          # Town next to Uluru
}
FORCE_BLACKLIST = {
    "saffire freycinet",  # Luxury hotel/lodge name, not a geographic place
}


# ---------------------------------------------------------------------------
# Parse NotebookLM report file
# ---------------------------------------------------------------------------

def parse_report(path):
    """Parse notebooklm_report.txt → {places, blacklist, synonyms, uncertain}"""
    text = open(path, encoding="utf-8").read()

    def extract_list(label):
        m = re.search(rf"{label}:\s*(.+?)(?=\n\n|\Z)", text, re.S | re.I)
        if not m:
            return []
        raw = m.group(1).strip()
        return [p.strip() for p in raw.split(",") if p.strip()]

    places    = extract_list("PLACE")
    blacklist = extract_list("BLACKLIST")
    uncertain = extract_list("UNCERTAIN")

    # Synonym pairs: "* Name -> Name" or "* Name -> Name"
    synonyms = []
    for m in re.finditer(r"\*\s*(.+?)\s*->\s*(.+)", text):
        synonyms.append((m.group(1).strip(), m.group(2).strip()))

    return {"places": places, "blacklist": blacklist, "synonyms": synonyms, "uncertain": uncertain}


# ---------------------------------------------------------------------------
# Parse NotebookLM chat file
# ---------------------------------------------------------------------------

def parse_chat(path):
    """Parse notebooklm_chat.txt → {places, blacklist, uncertain} (all rounds merged)"""
    text = open(path, encoding="utf-8").read()

    all_places    = []
    all_blacklist = []
    all_uncertain = []

    def extract_after(label, block):
        m = re.search(rf"{label}:\s*(.+?)(?=\n(?:PLACE|BLACKLIST|UNCERTAIN):|$)", block, re.S | re.I)
        if not m:
            return []
        return [p.strip() for p in m.group(1).split(",") if p.strip()]

    # Split into rounds by blank lines or PLACE: markers at the start of a line
    blocks = re.split(r"\n{2,}", text)
    for block in blocks:
        if re.search(r"^PLACE:", block, re.M | re.I):
            all_places    += extract_after("PLACE", block)
            all_blacklist += extract_after("BLACKLIST", block)
            all_uncertain += extract_after("UNCERTAIN", block)

    return {
        "places":    all_places,
        "blacklist": all_blacklist,
        "uncertain": all_uncertain,
    }


# ---------------------------------------------------------------------------
# Merge both sources into one canonical PLACE set
# ---------------------------------------------------------------------------

def build_place_set(report, chat):
    """
    Merge report PLACE + chat PLACE into one set.
    Apply manual conflict overrides (FORCE_PLACE, FORCE_BLACKLIST).
    Returns a set of lowercase place names.
    """
    combined = set()
    for p in report["places"] + chat["places"]:
        combined.add(p.lower())

    # Remove anything in report/chat blacklists (unless overridden to PLACE)
    all_blacklist = set(b.lower() for b in report["blacklist"] + chat["blacklist"])
    combined -= (all_blacklist - FORCE_PLACE)
    combined |= FORCE_PLACE
    combined -= FORCE_BLACKLIST
    return combined


# ---------------------------------------------------------------------------
# Replicate the same sample logic as benchmark_ner.py
# ---------------------------------------------------------------------------

def pick_same_sample(df, n=SAMPLE_SIZE):
    """Same logic as benchmark_ner.py _pick_sample() — must match exactly."""
    df = df[df["activity"].notna() & (df["activity"].str.len() > 80)].copy()
    df["_act_len"] = df["activity"].str.len()
    groups = []
    for state, grp in df.groupby("state"):
        groups.append(grp.nlargest(max(1, n // df["state"].nunique() + 2), "_act_len"))
    sampled = pd.concat(groups).drop_duplicates().nlargest(n, "_act_len")
    if len(sampled) < n:
        remaining = df[~df.index.isin(sampled.index)].sample(
            min(n - len(sampled), len(df) - len(sampled)), random_state=42
        )
        sampled = pd.concat([sampled, remaining])
    return sampled.head(n).reset_index(drop=True)


def cache_key(loc, act):
    return hashlib.md5(f"bench|{OLLAMA_MODEL}|{loc}|||{act[:1500]}".encode()).hexdigest()


# ---------------------------------------------------------------------------
# Export 50 rows as text for NotebookLM
# ---------------------------------------------------------------------------

NOTEBOOKLM_PROMPT = """\
TASK: Extract Australian place names from tour itinerary rows.
======================================================================

You will be given 50 rows of Australian tour itinerary text.
For EACH row, carefully read the Location and Activity fields and extract
EVERY Australian place name mentioned — no matter how briefly.

======================================================================
WHAT TO INCLUDE (extract these):
======================================================================
- Cities and towns: Sydney, Cairns, Alice Springs, Broome
- Suburbs and precincts: Paddington, Circular Quay, The Rocks, Surry Hills
- National parks and reserves: Kakadu National Park, Cape Le Grand National Park
- Beaches: Wineglass Bay, Bondi Beach, Lucky Bay, Whitehaven Beach
- Gorges and canyons: Simpsons Gap, Ormiston Gorge, Standley Chasm
- Mountains and peaks: Mount Wellington, Mount Isa, kunanyi
- Islands: Rottnest Island, Bruny Island, Lord Howe Island
- Rivers and lakes: Barron River, Lake Argyle, Daintree River
- Waterfalls: Millstream Falls, Russell Falls, Edith Falls
- Lookouts and viewpoints: Wineglass Bay Lookout, Oxer Lookout
- Landmarks and monuments: Sydney Opera House, Sydney Harbour Bridge
- Museums, galleries, theatres: Questacon, Australian War Memorial, MONA
- Tourist attractions and experiences: Crocosaurus Cove, Healesville Sanctuary
- Walks and tracks: Larapinta Trail, Bibbulmun Track, Organ Pipes Track
- Roads and drives: Great Ocean Road, Gibb River Road, Savannah Way
- Bays, capes, peninsulas: Mornington Peninsula, Cape Jervis, Geographe Bay
- Markets and precincts: Salamanca Market, Queen Victoria Market
- Historic sites and heritage areas: Port Arthur, Sovereign Hill
- Regions and areas: Fleurieu Peninsula, Atherton Tablelands, The Kimberley
- Indigenous place names: K'gari, Uluru, Nitmiluk, Purnululu

IMPORTANT — expand conjunctions:
If the text says "Joffre and Knox gorges" → extract both "Joffre Gorge" AND "Knox Gorge"
If the text says "Oxer and Junction Pool lookouts" → extract both separately

======================================================================
WHAT TO EXCLUDE (do NOT extract these):
======================================================================
- Hotel / resort / lodge names: Novotel, Voco, Saffire Freycinet, Capella Lodge, Peppers
- Restaurant and café names: Kettle & Tin, Tarte Bakery, Opera Bar
- Tour company names: Raging Thunder, Whales in Paradise, BridgeClimb (the company)
- Activity verbs and descriptions: Kayak, Snorkel, Hike, Swim, Scenic Flight
- Marketing phrases: Follow the Flavours, Fall in Love with Sydney, Epic Adventure
- Days and times: Day 1, Morning, Afternoon, Overnight
- Non-Australian places: Singapore, New Zealand, Queenstown NZ, Bali, Dubai
- Generic words alone: Area, Region, Reserve, Metro, CBD (unless part of a named place)
- People's names: Captain Cook, Matthew Flinders

======================================================================
WORKED EXAMPLE:
======================================================================
Input:
Location: Hobart to Bruny Island
Activity: A 40min drive from Hobart brings you to Kettering, where you board
the ferry across the D'Entrecasteaux Channel to North Bruny Island. At Great Bay,
stop for tastings at the Bruny Island Cheese Co. Explore South Bruny National Park,
flanked by cliffs with views over Cloudy Bay Beach and Cape Bruny Lighthouse.

Correct output:
ROW X: Hobart, Kettering, D'Entrecasteaux Channel, North Bruny Island, Great Bay, South Bruny National Park, Cloudy Bay Beach, Cape Bruny Lighthouse

Explanation:
- Hobart ✓ city
- Kettering ✓ town (ferry departure point)
- D'Entrecasteaux Channel ✓ waterway
- North Bruny Island ✓ island
- Great Bay ✓ bay
- Bruny Island Cheese Co ✗ EXCLUDED — food business name
- South Bruny National Park ✓ national park
- Cloudy Bay Beach ✓ beach
- Cape Bruny Lighthouse ✓ landmark

======================================================================
OUTPUT FORMAT — strict, one line per row:
======================================================================
ROW 1: Place A, Place B, Place C
ROW 2: Place D
ROW 3: Place E, Place F, Place G, Place H
...ROW 50: Place X, Place Y

Rules:
- One line per row, starting with "ROW N:"
- Comma-separated place names, no bullet points
- If truly no places found, write: ROW N: (none)
- Do NOT write any explanation, headings, or extra text
- Do NOT skip any ROW number — output all 50 lines

======================================================================
NOW EXTRACT FROM THESE 50 ROWS:
======================================================================

"""

def export_input_file(sample, ollama_results):
    """Write the 50 rows to a file for the user to paste into NotebookLM."""
    lines = [NOTEBOOKLM_PROMPT]
    for i, (_, row) in enumerate(sample.iterrows(), 1):
        loc = str(row.get("location", "") or "").strip()
        act = str(row.get("activity", "") or "").strip()
        ollama = ollama_results[i - 1]
        lines.append(f"=== ROW {i} ===")
        lines.append(f"Location: {loc}")
        lines.append(f"Activity: {act[:1500]}")
        if ollama:
            lines.append(f"[Ollama found for reference — do NOT copy this]: {', '.join(ollama)}")
        lines.append("")

    os.makedirs(os.path.dirname(INPUT_50), exist_ok=True)
    with open(INPUT_50, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Written: {INPUT_50}")
    print(f"  ({len(sample)} rows)")


# ---------------------------------------------------------------------------
# Parse NotebookLM extraction output
# ---------------------------------------------------------------------------

def parse_extraction_output(path):
    """
    Parse NotebookLM extraction output from notebooklm_extraction_50rows.txt.
    Expected format:
        ROW 1: Place A, Place B
        ROW 2: Place C
        ...
    Returns list of 50 lists (one per row). Missing rows get empty list.
    """
    results = {}
    text = open(path, encoding="utf-8").read()
    for m in re.finditer(r"ROW\s*(\d+)\s*:\s*(.*)$", text, re.M | re.I):
        row_num = int(m.group(1))
        raw = m.group(2).strip()
        if raw:
            places = [p.strip() for p in raw.split(",") if p.strip()]
        else:
            places = []
        results[row_num] = places

    # Build list of SAMPLE_SIZE entries
    return [results.get(i, []) for i in range(1, SAMPLE_SIZE + 1)]


# ---------------------------------------------------------------------------
# Parse AIP (JSON array) and Flash (line-wrapped plain text) extraction outputs
# ---------------------------------------------------------------------------

def parse_extraction_json(path):
    """Parse JSON array format: ["ROW 1: A, B", "ROW 2: C", ...]"""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    results = {}
    for item in data:
        m = re.match(r"ROW\s*(\d+)\s*:\s*(.+)", item.strip(), re.I)
        if m:
            results[int(m.group(1))] = [p.strip() for p in m.group(2).split(",") if p.strip()]
    return [results.get(i, []) for i in range(1, SAMPLE_SIZE + 1)]


def parse_extraction_wrapped(path):
    """Parse plain text with possible line-wrapped ROW entries (Flash format)."""
    with open(path, encoding="utf-8") as f:
        raw = " ".join(f.read().split())  # flatten line wraps
    results = {}
    for m in re.finditer(r"ROW\s*(\d+)\s*:\s*(.*?)(?=ROW\s*\d+\s*:|$)", raw, re.I):
        places = [p.strip() for p in m.group(2).split(",") if p.strip()]
        results[int(m.group(1))] = places
    return [results.get(i, []) for i in range(1, SAMPLE_SIZE + 1)]


# ---------------------------------------------------------------------------
# Build comparison Excel
# ---------------------------------------------------------------------------

def build_excel(sample, ollama_results, notebooklm_list_set,
                notebooklm_extract=None, aip_extract=None, flash_extract=None):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    rows = []
    summary_total    = 0
    summary_covered  = 0
    summary_missed   = 0

    for i, (_, row) in enumerate(sample.iterrows()):
        loc  = str(row.get("location", "") or "")
        act  = str(row.get("activity",  "") or "")
        ollama = ollama_results[i]

        ollama_lower = [p.lower() for p in ollama]
        in_list  = [p for p in ollama if p.lower() in notebooklm_list_set]
        not_in   = [p for p in ollama if p.lower() not in notebooklm_list_set]
        pct      = (len(in_list) / len(ollama) * 100) if ollama else None

        summary_total   += len(ollama)
        summary_covered += len(in_list)
        summary_missed  += len(not_in)

        r = {
            "row":           i + 1,
            "state":         row.get("state", ""),
            "location":      loc,
            "activity":      act[:300],
            "ollama_found":  " | ".join(ollama) if ollama else "(none)",
            "in_NLM_list":   " | ".join(in_list) if in_list else "(none)",
            "missing_from_NLM": " | ".join(not_in) if not_in else "(none)",
            "coverage_%":    f"{pct:.0f}%" if pct is not None else "N/A",
        }

        if notebooklm_extract is not None:
            nlm_ext   = notebooklm_extract[i]
            aip_ext   = aip_extract[i]   if aip_extract   else []
            fla_ext   = flash_extract[i] if flash_extract  else []

            nlm_set   = {p.lower() for p in nlm_ext}
            aip_set   = {p.lower() for p in aip_ext}
            fla_set   = {p.lower() for p in fla_ext}
            oll_set   = {p.lower() for p in ollama}

            # Per-place vote count (how many of the 3 extraction models found it)
            all_place_votes = {}
            for p in nlm_ext + aip_ext + fla_ext:
                k = p.lower()
                if k not in all_place_votes:
                    all_place_votes[k] = {"canonical": p, "votes": 0}
                all_place_votes[k]["votes"] += (k in nlm_set) + (k in aip_set) + (k in fla_set)
            # De-dup within each list first so votes aren't doubled
            all_place_votes = {}
            for lst, flag_set in [(nlm_ext, nlm_set), (aip_ext, aip_set), (fla_ext, fla_set)]:
                seen_in_this_list = set()
                for p in lst:
                    k = p.lower()
                    if k in seen_in_this_list:
                        continue
                    seen_in_this_list.add(k)
                    if k not in all_place_votes:
                        all_place_votes[k] = {"canonical": p, "votes": 0}
                    all_place_votes[k]["votes"] += 1

            in_all_3   = [v["canonical"] for v in all_place_votes.values() if v["votes"] == 3]
            hall_risk  = [v["canonical"] for v in all_place_votes.values() if v["votes"] == 1]

            # Merged union (preserve first-seen canonical casing)
            merged_order = []
            seen_merged  = set()
            for p in nlm_ext + aip_ext + fla_ext:
                k = p.lower()
                if k not in seen_merged:
                    seen_merged.add(k)
                    merged_order.append(p)

            r["NLM_extracted"]    = " | ".join(nlm_ext)      if nlm_ext      else "(none)"
            r["AIP_extracted"]    = " | ".join(aip_ext)      if aip_ext      else "(none)"
            r["Flash_extracted"]  = " | ".join(fla_ext)      if fla_ext      else "(none)"
            r["in_all_3"]         = " | ".join(in_all_3)     if in_all_3     else "(none)"
            r["merged_3way"]      = " | ".join(merged_order) if merged_order else "(none)"
            r["hallucination_risk"] = " | ".join(hall_risk)  if hall_risk    else ""
            r["only_ollama"]      = " | ".join(p for p in ollama if p.lower() not in seen_merged)
            r["in_both_oll_nlm"]  = " | ".join(p for p in ollama if p.lower() in nlm_set)

        rows.append(r)

    df_main = pd.DataFrame(rows)

    # Summary data
    overall_pct = (summary_covered / summary_total * 100) if summary_total else 0
    summary_rows = [
        ("Total Ollama places (50 rows)", summary_total),
        ("Covered by NotebookLM PLACE list", summary_covered),
        ("Missing from NotebookLM PLACE list", summary_missed),
        ("Overall coverage %", f"{overall_pct:.1f}%"),
        ("", ""),
        ("Verdict", "GOOD (>= 80%)" if overall_pct >= 80 else "NEEDS MORE CHUNKS (< 80%)"),
        ("", ""),
        ("--- 3-Way Extraction Comparison ---", ""),
    ]
    if notebooklm_extract and aip_extract and flash_extract:
        from collections import Counter as _Counter
        nlm_total  = sum(len(r) for r in notebooklm_extract)
        aip_total  = sum(len(r) for r in aip_extract)
        fla_total  = sum(len(r) for r in flash_extract)
        agree_3way = 0
        union_3way = 0
        hall_count = 0
        for i in range(SAMPLE_SIZE):
            ns = {p.lower() for p in notebooklm_extract[i]}
            as_ = {p.lower() for p in aip_extract[i]}
            fs = {p.lower() for p in flash_extract[i]}
            agree_3way += len(ns & as_ & fs)
            union_3way += len(ns | as_ | fs)
            all_votes = _Counter()
            for p in notebooklm_extract[i] + aip_extract[i] + flash_extract[i]:
                k = p.lower()
                all_votes[k] = (k in ns) + (k in as_) + (k in fs)
            hall_count += sum(1 for v in all_votes.values() if v == 1)
        summary_rows += [
            ("NLM extracted (50 rows)", nlm_total),
            ("AIP extracted (50 rows)", aip_total),
            ("Flash extracted (50 rows)", fla_total),
            ("All 3 agreed", agree_3way),
            ("Union of all 3", union_3way),
            ("NLM vs AIP agreement %", f"{(sum(len({p.lower() for p in notebooklm_extract[i]} & {p.lower() for p in aip_extract[i]}) for i in range(SAMPLE_SIZE)) / union_3way * 100 if union_3way else 0):.1f}%"),
            ("Single-model only (hallucination risk)", hall_count),
        ]
    df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    # Color scheme
    green_fill  = PatternFill("solid", fgColor="375623")
    blue_fill   = PatternFill("solid", fgColor="1F4E79")
    orange_fill = PatternFill("solid", fgColor="C55A11")
    red_fill    = PatternFill("solid", fgColor="9B2335")
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    white_bold  = Font(bold=True, color="FFFFFF")

    col_colors = {
        "ollama_found":        blue_fill,
        "in_NLM_list":         green_fill,
        "missing_from_NLM":    red_fill,
        "coverage_%":          blue_fill,
        "NLM_extracted":       orange_fill,
        "AIP_extracted":       blue_fill,
        "Flash_extracted":     blue_fill,
        "in_all_3":            green_fill,
        "merged_3way":         green_fill,
        "hallucination_risk":  red_fill,
        "only_ollama":         blue_fill,
        "in_both_oll_nlm":     green_fill,
    }

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:

        df_main.to_excel(writer, sheet_name="Comparison", index=False)
        ws = writer.sheets["Comparison"]

        col_widths = {
            "row": 5, "state": 12, "location": 28, "activity": 45,
            "ollama_found": 40, "in_NLM_list": 35, "missing_from_NLM": 35,
            "coverage_%": 12,
            "NLM_extracted": 45, "AIP_extracted": 45, "Flash_extracted": 45,
            "in_all_3": 40, "merged_3way": 50, "hallucination_risk": 40,
            "only_ollama": 35, "in_both_oll_nlm": 35,
        }

        for ci, col in enumerate(df_main.columns, 1):
            letter = get_column_letter(ci)
            cell = ws.cell(row=1, column=ci)
            cell.fill  = col_colors.get(col, blue_fill)
            cell.font  = white_bold
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[letter].width = col_widths.get(col, 25)
            for data_cell in ws[letter][1:]:
                data_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        for ri in range(2, len(df_main) + 2):
            if ri % 2 == 0:
                for cell in ws[ri]:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = alt_fill

        # Summary sheet
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        ws2 = writer.sheets["Summary"]
        ws2.column_dimensions["A"].width = 40
        ws2.column_dimensions["B"].width = 20
        ws2.cell(row=1, column=1).font = Font(bold=True)
        ws2.cell(row=1, column=2).font = Font(bold=True)

    print(f"  Saved: {OUT_XLSX}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading NotebookLM files...")
    report = parse_report(REPORT_FILE)
    chat   = parse_chat(CHAT_FILE)
    nlm_places = build_place_set(report, chat)
    print(f"  Report: {len(report['places'])} places, {len(report['blacklist'])} blacklist, {len(report['synonyms'])} synonyms")
    print(f"  Chat:   {len(chat['places'])} places, {len(chat['blacklist'])} blacklist")
    print(f"  Merged PLACE set: {len(nlm_places)} unique places (after conflict overrides)")

    print(f"\nLoading CSV: {CSV_PATH} ...")
    df = pd.read_csv(CSV_PATH)
    print(f"  {len(df):,} rows total")

    sample = pick_same_sample(df, SAMPLE_SIZE)
    print(f"  Sampled {len(sample)} rows across {sample['state'].nunique()} states")

    print(f"\nLooking up Ollama benchmark cache...")
    cache = {}
    if os.path.exists(CACHE_FILE):
        cache = json.load(open(CACHE_FILE, encoding="utf-8"))
    print(f"  Cache has {len(cache)} entries")

    ollama_results = []
    cache_hits = 0
    for _, row in sample.iterrows():
        loc = str(row.get("location", "") or "")
        act = str(row.get("activity", "") or "")
        k = cache_key(loc, act)
        places = cache.get(k, [])
        if places:
            cache_hits += 1
        ollama_results.append(places)

    print(f"  Cache hits: {cache_hits}/{SAMPLE_SIZE} rows")

    print(f"\nExporting 50 rows for NotebookLM...")
    export_input_file(sample, ollama_results)

    print(f"\nPASTE THIS FILE INTO NOTEBOOKLM:")
    print(f"  {INPUT_50}")
    print(f"\nSave the NotebookLM response to:")
    print(f"  {EXTRACT_OUT}")

    # Load AIP and Flash extraction outputs if they exist
    aip_extract   = None
    flash_extract = None
    if os.path.exists(AIP_FILE):
        aip_extract = parse_extraction_json(AIP_FILE)
        print(f"  AIP (JSON): {sum(len(r) for r in aip_extract)} places across {sum(1 for r in aip_extract if r)} rows")
    if os.path.exists(FLASH_FILE):
        flash_extract = parse_extraction_wrapped(FLASH_FILE)
        print(f"  Flash (text): {sum(len(r) for r in flash_extract)} places across {sum(1 for r in flash_extract if r)} rows")

    # Step 3: if NLM extraction output exists, do full 3-way comparison
    if os.path.exists(EXTRACT_OUT):
        print(f"\nFound NLM extraction output: {EXTRACT_OUT}")
        notebooklm_extract = parse_extraction_output(EXTRACT_OUT)
        found = sum(1 for r in notebooklm_extract if r)
        print(f"  NLM parsed: {found}/{SAMPLE_SIZE} rows | {sum(len(r) for r in notebooklm_extract)} places")
        print(f"\nBuilding 3-way comparison Excel...")
        build_excel(sample, ollama_results, nlm_places, notebooklm_extract, aip_extract, flash_extract)
    else:
        print(f"\nNo NLM extraction output yet — building coverage-only Excel...")
        build_excel(sample, ollama_results, nlm_places,
                    notebooklm_extract=None, aip_extract=aip_extract, flash_extract=flash_extract)

    print("\nDone.")


if __name__ == "__main__":
    main()
